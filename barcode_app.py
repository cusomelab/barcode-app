# ╔══════════════════════════════════════════════════════╗
# ║         [쿠썸] 바코드 라벨 생성기 - Streamlit         ║
# ╚══════════════════════════════════════════════════════╝
import os, io, re, urllib.request, csv, zipfile
from collections import OrderedDict
from datetime import datetime, timedelta
import streamlit as st
import streamlit.components.v1 as components_v1
import time
from PIL import Image, ImageDraw, ImageFont
import barcode
from barcode.writer import ImageWriter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                 TableStyle, HRFlowable, PageBreak)
from reportlab.lib.styles import ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from pypdf import PdfWriter, PdfReader
import pdfplumber
import pypdfium2 as pdfium

# ── 폰트 준비 ──────────────────────────────────────────
FONT_PATH = 'NanumGothicBold.ttf'
FONT_REG_PATH = 'NanumGothic.ttf'
if not os.path.exists(FONT_PATH):
    urllib.request.urlretrieve(
        'https://github.com/google/fonts/raw/main/ofl/nanumgothic/NanumGothic-Bold.ttf',
        FONT_PATH
    )
if not os.path.exists(FONT_REG_PATH):
    urllib.request.urlretrieve(
        'https://github.com/google/fonts/raw/main/ofl/nanumgothic/NanumGothic-Regular.ttf',
        FONT_REG_PATH
    )

# reportlab 한국어 폰트 등록
try:
    pdfmetrics.registerFont(TTFont('NanumBold', FONT_PATH))
    pdfmetrics.registerFont(TTFont('NanumReg', FONT_REG_PATH))
except Exception:
    pass

# ══════════════════════════════════════════════════════
# ── 공통 헬퍼 (바코드 라벨용) ──────────────────────────
# ══════════════════════════════════════════════════════
def wrap_text(text, font, max_w, draw):
    lines, cur = [], []
    for word in text.split(' '):
        test = ' '.join(cur + [word])
        bbox = draw.textbbox((0,0), test, font=font)
        if bbox[2]-bbox[0] <= max_w: cur.append(word)
        else:
            if cur: lines.append(' '.join(cur)); cur=[word]
            else:
                tmp=''
                for ch in word:
                    tb=draw.textbbox((0,0),tmp+ch,font=font)
                    if tb[2]-tb[0]<=max_w: tmp+=ch
                    else:
                        if tmp: lines.append(tmp)
                        tmp=ch
                if tmp: lines.append(tmp)
                cur=[]
    if cur: lines.append(' '.join(cur))
    return lines

def get_barcode_img(barcode_number, write_text=False):
    bc = barcode.get('code128', barcode_number, writer=ImageWriter())
    bc.writer.set_options({'module_height':80,'module_width':1.6,
                           'quiet_zone':6,'write_text':write_text})
    raw = bc.render()
    g=raw.convert('L'); w_r,h_r=g.size; p=g.load()
    l =next(x for x in range(w_r)          if any(p[x,yy]<255 for yy in range(h_r)))
    rr=next(x for x in range(w_r-1,-1,-1)   if any(p[x,yy]<255 for yy in range(h_r)))+1
    t =next(yy for yy in range(h_r)         if any(p[xx,yy]<255 for xx in range(w_r)))
    if write_text:
        bar_end=t
        for yy in range(t,h_r):
            if sum(1 for xx in range(l,rr) if p[xx,yy]<100)/(rr-l)>=0.25: bar_end=yy
        return raw.crop((l,t,rr,bar_end+2))
    else:
        b=next(yy for yy in range(h_r-1,-1,-1) if any(p[xx,yy]<255 for xx in range(w_r)))+1
        return raw.crop((l,t,rr,b))

# ── 소형 라벨 생성 ─────────────────────────────────────
def create_small(product_name, barcode_number, material, fixed_origin, fixed_age):
    CANVAS_W, CANVAS_H = 650, 450; PAD = 30
    img=Image.new('RGB',(CANVAS_W,CANVAS_H),'white'); draw=ImageDraw.Draw(img)
    font_big=ImageFont.truetype(FONT_PATH,26)
    font_mid=ImageFont.truetype(FONT_PATH,20)

    y=PAD
    for line in wrap_text(product_name,font_big,CANVAS_W-PAD*2,draw)[:2]:
        bb=draw.textbbox((0,0),line,font=font_big)
        draw.text(((CANVAS_W-(bb[2]-bb[0]))//2,y),line,font=font_big,fill='black')
        y+=bb[3]-bb[1]+6

    bc_img=get_barcode_img(barcode_number,write_text=False)
    c1b=draw.textbbox((0,0),fixed_origin,font=font_mid)
    c2b=draw.textbbox((0,0),fixed_age,font=font_mid)
    mat_h=font_mid.size+8 if material else 0
    fixed_h=(c1b[3]-c1b[1])+(c2b[3]-c2b[1])+mat_h+18
    bc_y=y+10; cue_y=CANVAS_H-fixed_h-PAD
    BAR_W=CANVAS_W-PAD*2; BAR_H=cue_y-bc_y-6
    if BAR_H<60: BAR_H=60
    img.paste(bc_img.resize((BAR_W,BAR_H),Image.LANCZOS),(PAD,bc_y))
    cur_y=bc_y+BAR_H+8

    if material:
        mt=f'재질 : {material}'; mb=draw.textbbox((0,0),mt,font=font_mid)
        draw.text(((CANVAS_W-(mb[2]-mb[0]))//2,cur_y),mt,font=font_mid,fill='black')
        cur_y+=mb[3]-mb[1]+6
    for txt in (fixed_origin,fixed_age):
        tb=draw.textbbox((0,0),txt,font=font_mid)
        draw.text(((CANVAS_W-(tb[2]-tb[0]))//2,cur_y),txt,font=font_mid,fill='black')
        cur_y+=tb[3]-tb[1]+5
    return img

# ── 대형 라벨 생성 ─────────────────────────────────────
def fit_font(text, max_w, draw, max_size=42, min_size=8):
    for size in range(max_size,min_size-1,-1):
        f=ImageFont.truetype(FONT_PATH,size)
        bb=draw.textbbox((0,0),text,font=f)
        if bb[2]-bb[0]<=max_w: return f
    return ImageFont.truetype(FONT_PATH,min_size)

def create_large(product_name, barcode_number, material, fix_list):
    CANVAS_W,CANVAS_H=450,640; PAD=22
    img=Image.new('RGB',(CANVAS_W,CANVAS_H),'white'); draw=ImageDraw.Draw(img)
    fn=ImageFont.truetype(FONT_PATH,33)
    fm=ImageFont.truetype(FONT_PATH,20)
    ff=ImageFont.truetype(FONT_PATH,16)

    y=PAD
    for line in wrap_text(product_name,fn,CANVAS_W-PAD*2,draw)[:3]:
        bb=draw.textbbox((0,0),line,font=fn)
        draw.text(((CANVAS_W-(bb[2]-bb[0]))//2,y),line,font=fn,fill='black')
        y+=bb[3]-bb[1]+4
    y+=8

    bc_img=get_barcode_img(barcode_number,write_text=False)
    fix_h=0
    for txt in fix_list:
        for ln in wrap_text(txt,ff,CANVAS_W-PAD*2,draw):
            bb=draw.textbbox((0,0),ln,font=ff); fix_h+=bb[3]-bb[1]+3
        fix_h+=6

    BAR_W=CANVAS_W-PAD*2
    BAR_H=CANVAS_H-y-30-14-fix_h-PAD
    if BAR_H<60: BAR_H=60
    img.paste(bc_img.resize((BAR_W,BAR_H),Image.LANCZOS),(PAD,y)); y+=BAR_H+6

    if material:
        mt=f'재질 : {material}'; mb=draw.textbbox((0,0),mt,font=fm)
        draw.text(((CANVAS_W-(mb[2]-mb[0]))//2,y),mt,font=fm,fill='black')
        y+=mb[3]-mb[1]+8

    draw.line([(PAD,y),(CANVAS_W-PAD,y)],fill=(160,160,160),width=1); y+=10
    for txt in fix_list:
        for ln in wrap_text(txt,ff,CANVAS_W-PAD*2,draw):
            bb=draw.textbbox((0,0),ln,font=ff)
            draw.text((PAD,y),ln,font=ff,fill='black'); y+=bb[3]-bb[1]+3
        y+=6

    draw.rectangle([1,1,CANVAS_W-2,CANVAS_H-2],outline=(180,180,180),width=1)
    return img.rotate(90,expand=True)

# ── 엑셀 처리 공통 ─────────────────────────────────────
def process_excel(uploaded_file, mode, settings):
    wb=load_workbook(uploaded_file); ws=wb.active
    col_insert=settings['col_insert']
    ws.column_dimensions[get_column_letter(col_insert)].width=settings['col_width']

    temp_dir='_temp'; os.makedirs(temp_dir,exist_ok=True)
    ok=0; errors=[]
    progress=st.progress(0)
    status=st.empty()

    total=sum(1 for r in range(settings['start_row'],ws.max_row+1)
              if ws.cell(r,settings['col_barcode']).value)

    for r in range(settings['start_row'],ws.max_row+1):
        bv=ws.cell(r,settings['col_barcode']).value
        if not bv: continue
        bv=str(bv).strip()
        nm=str(ws.cell(r,settings['col_name']).value or '').strip()
        mt=str(ws.cell(r,settings['col_material']).value or '').strip()
        img_path=f'{temp_dir}/label_{r}.png'

        try:
            if mode=='소형':
                img=create_small(nm,bv,mt,settings['origin'],settings['age'])
            else:
                img=create_large(nm,bv,mt,settings['fix_list'])
            img.save(img_path)
        except Exception as e:
            errors.append(f'{r}행 실패: {e}'); continue

        xl=XLImage(img_path)
        xl.width=settings['insert_w']; xl.height=settings['insert_h']
        ws.add_image(xl,f'{get_column_letter(col_insert)}{r}')
        ws.row_dimensions[r].height=settings['row_height']
        ok+=1
        progress.progress(ok/max(total,1))
        status.text(f'✅ {r}행 처리 중... ({ok}/{total})')

    output=io.BytesIO(); wb.save(output); output.seek(0)
    progress.progress(1.0); status.text(f'🎉 완료! {ok}개 생성')
    return output, ok, errors


# ══════════════════════════════════════════════════════
# ── 출고 작업 지시서 PDF 생성 함수들 ──────────────────
# ══════════════════════════════════════════════════════

def parse_date(date_str):
    """날짜 문자열을 파싱해서 datetime 반환. 실패하면 None."""
    if not date_str or not date_str.strip():
        return None
    s = date_str.strip()
    nums = [n for n in __import__('re').findall(r'\d+', s)]
    try:
        if len(nums) == 1 and len(nums[0]) == 8:
            ymd = nums[0]
            return datetime(int(ymd[:4]), int(ymd[4:6]), int(ymd[6:8]))
        elif len(nums) >= 3:
            y, m, d = int(nums[0]), int(nums[1]), int(nums[2])
            if y < 100: y += 2000
            return datetime(y, m, d)
    except Exception:
        pass
    return None


def calc_deadline(date_str):
    """입고예정일 + 20일 → 입고마감일 문자열 반환"""
    dt = parse_date(date_str)
    if dt is None:
        return '날짜 없음'
    deadline = dt + timedelta(days=20)
    return deadline.strftime('%Y-%m-%d')


def parse_csv_to_items(file_bytes):
    """CSV 바이트 → WorkOrderItem 리스트 반환"""
    text = file_bytes.decode('utf-8-sig', errors='replace')
    reader = csv.reader(text.splitlines())
    rows = list(reader)
    if not rows:
        return []

    # 헤더 행 건너뛰기 (첫 행)
    data_rows = rows[1:]
    items = []
    for row in data_rows:
        if len(row) < 11 or not (row[1] if len(row) > 1 else '').strip():
            continue
        def safe(idx, default=''):
            return row[idx].strip() if idx < len(row) else default
        try:
            qty = int(safe(7, '0') or '0')
        except ValueError:
            qty = 0
        items.append({
            'logisticsCenter': safe(1),
            'expectedDate':    safe(3),
            'productBarcode':  safe(5),
            'productName':     safe(6),
            'quantity':        qty,
            'shipmentNumber':  safe(8),
            'orderDate':       safe(9),
            'boxNumber':       safe(10),
            'location':        safe(12),
        })
    return items


def group_items(items, grouping_keys):
    """그룹화 기준에 따라 dict로 묶음"""
    grouped = {}
    for item in items:
        key_parts = [item.get(k, '') for k in grouping_keys if item.get(k)]
        key = '_'.join(key_parts) if key_parts else '(미분류)'
        grouped.setdefault(key, []).append(item)

    # 각 그룹 내부 정렬: 박스번호 → 상품명
    import locale
    for key in grouped:
        grouped[key].sort(key=lambda x: (
            [int(c) if c.isdigit() else c.lower()
             for c in __import__('re').split(r'(\d+)', x.get('boxNumber',''))],
            x.get('productName','')
        ))
    return grouped


def create_work_order_pdf(group_key, items, shipment_id=None, box_number=None):
    """reportlab으로 출고 작업 지시서 PDF 생성 → BytesIO 반환
    shipment_id/box_number가 있으면 상단 우측에 표시"""
    buf = io.BytesIO()
    PAGE_W, PAGE_H = A4
    MARGIN = 18 * mm

    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=MARGIN, bottomMargin=MARGIN
    )

    # 스타일 정의
    s_title   = ParagraphStyle('title',   fontName='NanumBold', fontSize=18, leading=22, textColor=colors.HexColor('#111111'))
    s_sub     = ParagraphStyle('sub',     fontName='NanumBold', fontSize=9,  leading=12, textColor=colors.HexColor('#888888'), spaceAfter=2)
    s_card_lbl= ParagraphStyle('cardlbl', fontName='NanumBold', fontSize=8,  leading=10, textColor=colors.HexColor('#888888'))
    s_card_val= ParagraphStyle('cardval', fontName='NanumBold', fontSize=12, leading=15, textColor=colors.HexColor('#111111'))
    s_card_big= ParagraphStyle('cardbig', fontName='NanumBold', fontSize=22, leading=26, textColor=colors.HexColor('#1a56db'))
    s_th      = ParagraphStyle('th',      fontName='NanumBold', fontSize=9,  leading=11, textColor=colors.white)
    s_td      = ParagraphStyle('td',      fontName='NanumReg',  fontSize=9,  leading=12, textColor=colors.HexColor('#111111'), wordWrap='CJK')
    s_td_bold = ParagraphStyle('tdbold',  fontName='NanumBold', fontSize=10, leading=12, textColor=colors.HexColor('#1a56db'))
    s_mono    = ParagraphStyle('mono',    fontName='NanumReg',  fontSize=9,  leading=12, textColor=colors.HexColor('#111111'))
    s_footer  = ParagraphStyle('footer',  fontName='NanumReg',  fontSize=8,  leading=10, textColor=colors.HexColor('#888888'))
    s_shipment= ParagraphStyle('shipment',fontName='NanumBold', fontSize=14, leading=18, textColor=colors.HexColor('#1a56db'), alignment=2)
    s_boxnum_huge = ParagraphStyle('boxnumhuge', fontName='NanumBold', fontSize=48, leading=54, textColor=colors.HexColor('#dc2626'), alignment=0)

    total_qty   = sum(i['quantity'] for i in items)
    first       = items[0]
    deadline    = calc_deadline(first.get('expectedDate',''))
    created_at  = datetime.now().strftime('%Y-%m-%d %H:%M')
    usable_w    = PAGE_W - MARGIN * 2

    story = []

    # ── 송장번호 바코드 (피킹검증 스캔용) ───────────────
    # CSV에 있는 송장번호(shipmentNumber)를 바코드로 추가
    invoice_for_barcode = first.get('shipmentNumber', '') or ''
    barcode_flowable = None
    if invoice_for_barcode:
        try:
            # 벡터 기반 Code128 (PNG보다 훨씬 선명)
            from reportlab.graphics.barcode.code128 import Code128
            from reportlab.graphics.shapes import Drawing
            from reportlab.graphics import renderPDF
            from reportlab.platypus import Flowable

            class _BarcodeFlowable(Flowable):
                def __init__(self, value, width_mm=70, height_mm=14):
                    Flowable.__init__(self)
                    self.value = str(value)
                    self.width = width_mm * mm
                    self.height = height_mm * mm
                    self.hAlign = 'RIGHT'

                def draw(self):
                    # barWidth를 원하는 폭에 맞춰 계산 (여백 포함)
                    # Code128은 가변 길이라 렌더 후 스케일링
                    bc = Code128(self.value, barHeight=self.height, humanReadable=False)
                    bc_w = bc.width
                    scale = self.width / bc_w if bc_w > 0 else 1
                    self.canv.saveState()
                    self.canv.scale(scale, 1)
                    bc.drawOn(self.canv, 0, 0)
                    self.canv.restoreState()

            barcode_flowable = _BarcodeFlowable(invoice_for_barcode, width_mm=70, height_mm=14)
        except Exception:
            # Fallback: 기존 PNG 방식
            try:
                from reportlab.platypus import Image as RLImage
                bc_img = get_barcode_img(str(invoice_for_barcode), write_text=False)
                bc_buf = io.BytesIO()
                bc_img.save(bc_buf, format='PNG')
                bc_buf.seek(0)
                barcode_flowable = RLImage(bc_buf, width=70*mm, height=14*mm)
                barcode_flowable.hAlign = 'RIGHT'
            except Exception:
                barcode_flowable = None

    # ── 상단 큰 박스번호 표시 ──
    # 호출 시 box_number 인자로 전달된 값을 사용 (자동 부여된 송장별 박스번호)
    # 송장 전체가 국내재고인 경우 box_number=None이 전달되어 표시되지 않음
    # 배대지 박스별 수량 요약도 함께 표시 (예: "58번 M2(10),W11(93)")
    def _extract_box_key(bn_raw):
        """boxNumber 값에서 배대지 박스 키(M2, W11 등)만 추출.
        실제 포맷 예: '★M12(1)' → 'M12', 'W3(5)' → 'W3', 'M2' → 'M2'
        국내재고/부족/RAW 등은 제외."""
        import re as _re_ek
        s = str(bn_raw or '').strip()
        if not s or s.lower() == 'nan':
            return ''
        if any(kw in s for kw in ('국내', '부족', '재고', 'RAW')):
            return ''
        # 기호(★●■ 등) 프리픽스와 (N) 수량 무시하고 박스키만 추출
        m = _re_ek.search(r'([A-Za-z]+\d+)', s)
        if m:
            return m.group(1).upper()
        m = _re_ek.search(r'(\d+)', s)
        return m.group(1) if m else ''

    dapae_summary_str = ''
    if box_number:
        from collections import Counter as _Counter
        dapae_counts = _Counter()  # {배대지박스키: 수량합}
        for it in items:
            bkey = _extract_box_key(it.get('boxNumber', ''))
            if bkey:
                dapae_counts[bkey] += it.get('quantity', 0)
        if dapae_counts:
            parts = [f'{k}({v})' for k, v in sorted(dapae_counts.items())]
            # '1번'은 48pt, 배대지 구성은 18pt(타이틀 크기)로 축소
            dapae_summary_str = (
                f' <font size="18" color="#111111">'
                f'{",".join(parts)}</font>'
            )
    big_box_label = f'{box_number}번{dapae_summary_str}' if box_number else ''

    # ── 헤더: 좌측 타이틀 + 우측 쉽먼트(+바코드) ──────────
    # 박스번호는 좌측 상단에 이미 표시되므로 우측에는 쉽먼트 ID만
    if shipment_id:
        shipment_text = f'쉽먼트 {shipment_id}'
    else:
        shipment_text = ''

    if shipment_text or barcode_flowable:
        right_cell = []
        if barcode_flowable:
            right_cell.append(barcode_flowable)
            right_cell.append(Spacer(1, 1*mm))
        if shipment_text:
            right_cell.append(Paragraph(shipment_text, s_shipment))
        right_tbl = Table([[c] for c in right_cell], colWidths=[usable_w * 0.5])
        right_tbl.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ]))
        # 좌측: 큰 박스번호 + 타이틀
        if big_box_label:
            left_cell = [
                Paragraph(f'📦 {big_box_label}', s_boxnum_huge),
                Paragraph('출고 작업 지시서', s_title),
            ]
            left_tbl = Table([[c] for c in left_cell], colWidths=[usable_w * 0.5])
            left_tbl.setStyle(TableStyle([
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
                ('TOPPADDING', (0,0), (-1,-1), 0),
                ('BOTTOMPADDING', (0,0), (-1,-1), 0),
            ]))
            left_content = left_tbl
        else:
            left_content = Paragraph('출고 작업 지시서', s_title)

        header_data = [[left_content, right_tbl]]
        header_tbl = Table(header_data, colWidths=[usable_w * 0.5, usable_w * 0.5])
        header_tbl.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'BOTTOM'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ]))
        story.append(header_tbl)
    else:
        if big_box_label:
            story.append(Paragraph(f'📦 {big_box_label}', s_boxnum_huge))
        story.append(Paragraph('출고 작업 지시서', s_title))

    story.append(Paragraph(group_key, s_sub))
    story.append(Spacer(1, 1*mm))
    story.append(HRFlowable(width='100%', thickness=2, color=colors.HexColor('#1a56db')))
    story.append(Spacer(1, 4*mm))

    # ── 정보 카드 (3열 테이블) ─────────────────────────
    col_w = usable_w / 3

    def info_card(label, value, big=False):
        lbl = Paragraph(label, s_card_lbl)
        val = Paragraph(str(value), s_card_big if big else s_card_val)
        return [lbl, val]

    card_data = [[
        info_card('물류센터',  first.get('logisticsCenter','-')),
        info_card('입고예정일', first.get('expectedDate','-')),
        info_card('총 수량',   total_qty, big=True),
    ],[
        info_card('송장번호',  first.get('shipmentNumber','-')),
        info_card('입고마감일', deadline),
        info_card('품목 수',   f'{len(items)}개'),
    ]]

    def make_card_cell(label_val_list):
        """[lbl_para, val_para] → 테이블 셀용 nested table"""
        t = Table([[label_val_list[0]], [label_val_list[1]]], colWidths=[col_w - 6*mm])
        t.setStyle(TableStyle([
            ('LEFTPADDING',  (0,0),(-1,-1), 0),
            ('RIGHTPADDING', (0,0),(-1,-1), 0),
            ('TOPPADDING',   (0,0),(-1,-1), 1),
            ('BOTTOMPADDING',(0,0),(-1,-1), 1),
        ]))
        return t

    card_table_data = []
    for row in card_data:
        card_table_data.append([make_card_cell(cell) for cell in row])

    card_bg = [colors.HexColor('#f8fafc'), colors.HexColor('#eff6ff')]
    card_tbl = Table(card_table_data, colWidths=[col_w]*3)
    card_style = [
        ('BACKGROUND', (0,0),(2,0), card_bg[0]),
        ('BACKGROUND', (0,1),(2,1), card_bg[1]),
        ('BOX',        (0,0),(2,1), 1, colors.HexColor('#e2e8f0')),
        ('INNERGRID',  (0,0),(2,1), 0.5, colors.HexColor('#e2e8f0')),
        ('LEFTPADDING',  (0,0),(-1,-1), 4*mm),
        ('RIGHTPADDING', (0,0),(-1,-1), 2*mm),
        ('TOPPADDING',   (0,0),(-1,-1), 3*mm),
        ('BOTTOMPADDING',(0,0),(-1,-1), 3*mm),
        ('VALIGN',     (0,0),(-1,-1), 'MIDDLE'),
        ('ROUNDEDCORNERS', [4]),
    ]
    card_tbl.setStyle(TableStyle(card_style))
    story.append(card_tbl)
    story.append(Spacer(1, 5*mm))

    # ── 상품 테이블 ────────────────────────────────────
    # 컬럼 폭: 바코드 27%, 상품명 35%, 수량 8%, 위치 14%, 박스 16%
    cw = [usable_w*p for p in [0.25, 0.35, 0.08, 0.16, 0.16]]

    header_row = [
        Paragraph('바코드', s_th),
        Paragraph('상품명', s_th),
        Paragraph('수량', s_th),
        Paragraph('위치', s_th),
        Paragraph('박스', s_th),
    ]
    table_data = [header_row]

    for i, item in enumerate(items):
        # ReportLab Paragraph는 XML 파서를 사용하므로 특수문자 이스케이프 필수
        from xml.sax.saxutils import escape as _xml_escape
        row = [
            Paragraph(_xml_escape(item.get('productBarcode','')), s_mono),
            Paragraph(_xml_escape(item.get('productName','')),    s_td),
            Paragraph(str(item.get('quantity',0)),   s_td_bold),
            Paragraph(_xml_escape(item.get('location','')),       s_td),
            Paragraph(_xml_escape(item.get('boxNumber','')),      s_td),
        ]
        table_data.append(row)

    # 합계 행
    table_data.append([
        Paragraph('합  계', ParagraphStyle('sum', fontName='NanumBold', fontSize=9, textColor=colors.HexColor('#111111'))),
        Paragraph('', s_td),
        Paragraph(str(total_qty), ParagraphStyle('sumqty', fontName='NanumBold', fontSize=12, textColor=colors.HexColor('#1a56db'))),
        Paragraph('', s_td),
        Paragraph('', s_td),
    ])

    tbl = Table(table_data, colWidths=cw, repeatRows=1)

    row_colors = []
    for i in range(1, len(table_data)-1):
        bg = colors.white if i % 2 == 1 else colors.HexColor('#f8fafc')
        row_colors.append(('BACKGROUND', (0,i),(4,i), bg))

    tbl_style = [
        # 헤더
        ('BACKGROUND', (0,0),(4,0), colors.HexColor('#1e293b')),
        ('TEXTCOLOR',  (0,0),(4,0), colors.white),
        ('ALIGN',      (0,0),(4,0), 'CENTER'),
        # 합계 행
        ('BACKGROUND', (0,-1),(4,-1), colors.HexColor('#eff6ff')),
        ('LINEABOVE',  (0,-1),(4,-1), 1, colors.HexColor('#93c5fd')),
        # 전체
        ('FONTSIZE',   (0,0),(-1,-1), 9),
        ('TOPPADDING', (0,0),(-1,-1), 4),
        ('BOTTOMPADDING',(0,0),(-1,-1), 4),
        ('LEFTPADDING',(0,0),(-1,-1), 3*mm),
        ('RIGHTPADDING',(0,0),(-1,-1), 2*mm),
        ('VALIGN',     (0,0),(-1,-1), 'MIDDLE'),
        ('ALIGN',      (2,1),(2,-1), 'CENTER'),  # 수량 가운데
        ('GRID',       (0,0),(-1,-1), 0.4, colors.HexColor('#e2e8f0')),
        ('LINEBELOW',  (0,0),(4,0), 1, colors.HexColor('#1a56db')),
    ] + row_colors

    tbl.setStyle(TableStyle(tbl_style))
    story.append(tbl)
    story.append(Spacer(1, 5*mm))

    # ── 푸터 ──────────────────────────────────────────
    story.append(HRFlowable(width='100%', thickness=0.5, color=colors.HexColor('#e2e8f0')))
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph(
        f'※ 바코드와 수량을 작업 전 반드시 대조해 주세요. (자동 생성 문서) · 생성일시: {created_at}',
        s_footer
    ))

    doc.build(story)
    buf.seek(0)
    return buf


def merge_pdfs(pdf_buffers):
    """여러 PDF BytesIO를 하나로 병합 → BytesIO 반환"""
    writer = PdfWriter()
    for buf in pdf_buffers:
        reader = PdfReader(buf)
        for page in reader.pages:
            writer.add_page(page)
    out = io.BytesIO()
    writer.write(out)
    out.seek(0)
    return out


def assign_box_numbers(items):
    """items에서 쉽먼트(송장)별로 박스번호 자동 부여.
    정렬 기준: 입고예정일 > 물류센터 > 송장번호
    송장 전체가 국내재고/부족일 때만 제외 (일부만 국내재고면 부여).
    반환: {송장번호: 박스번호(int)} 딕셔너리
    """
    if not items:
        return {}
    ship_info = {}   # ship → (expectedDate, center, ship)
    ship_valid = {}  # ship → 하나라도 유효한 박스가 있으면 True
    for it in items:
        ship = str(it.get('shipmentNumber', '') or '').strip()
        if not ship:
            continue
        bn_raw = str(it.get('boxNumber', '') or '').strip()
        is_domestic = bool(bn_raw) and any(kw in bn_raw for kw in ('국내', '부족', '재고', 'RAW'))
        has_valid_box = bool(bn_raw) and not is_domestic
        if ship not in ship_info:
            center = str(it.get('logisticsCenter', '') or '').strip()
            edate = str(it.get('expectedDate', '') or '').strip()
            ship_info[ship] = (edate, center, ship)
            ship_valid[ship] = has_valid_box
        else:
            # 하나라도 유효하면 True로 유지
            if has_valid_box:
                ship_valid[ship] = True
    valid_ships = {s: info for s, info in ship_info.items() if ship_valid.get(s, False)}
    # 정렬: 입고예정일 → 물류센터 → 송장번호
    sorted_ships = sorted(valid_ships.values(), key=lambda x: (x[0], x[1], x[2]))
    return {s[2]: idx + 1 for idx, s in enumerate(sorted_ships)}


def assign_box_numbers_with_existing(items, existing_box_map):
    """기존 매핑(시트 M열)을 보존하고, 그 외 송장만 신규 번호 부여.
    - existing_box_map에 있는 송장은 기존 번호 그대로 유지 (발주 취소돼도 고정)
    - 신규 송장은 max(기존)+1 부터 (입고예정일>물류센터>송장번호) 순으로 부여
    반환: {송장번호: 박스번호(int)}
    """
    result = {}
    for s, v in (existing_box_map or {}).items():
        s = str(s).strip()
        try:
            n = int(str(v).strip())
            if s and n > 0:
                result[s] = n
        except (ValueError, TypeError):
            continue
    ship_info = {}
    ship_valid = {}
    for it in items or []:
        ship = str(it.get('shipmentNumber', '') or '').strip()
        if not ship or ship in result:
            continue
        bn_raw = str(it.get('boxNumber', '') or '').strip()
        is_domestic = bool(bn_raw) and any(kw in bn_raw for kw in ('국내', '부족', '재고', 'RAW'))
        has_valid_box = bool(bn_raw) and not is_domestic
        if ship not in ship_info:
            center = str(it.get('logisticsCenter', '') or '').strip()
            edate = str(it.get('expectedDate', '') or '').strip()
            ship_info[ship] = (edate, center, ship)
            ship_valid[ship] = has_valid_box
        else:
            if has_valid_box:
                ship_valid[ship] = True
    valid_new = {s: info for s, info in ship_info.items() if ship_valid.get(s, False)}
    sorted_new = sorted(valid_new.values(), key=lambda x: (x[0], x[1], x[2]))
    next_num = max(result.values(), default=0) + 1
    for s in sorted_new:
        result[s[2]] = next_num
        next_num += 1
    return result


def create_shipment_barcodes_pdf(shipment_numbers):
    """송장번호 리스트 → 바코드 PDF (한 페이지에 여러 송장 배치)"""
    from reportlab.platypus import Image as RLImage
    buf = io.BytesIO()
    PAGE_W, PAGE_H = A4
    MARGIN = 15 * mm

    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=MARGIN, bottomMargin=MARGIN
    )

    s_title = ParagraphStyle('btitle', fontName='NanumBold', fontSize=14, leading=18, alignment=1, spaceAfter=8)
    s_num = ParagraphStyle('bnum', fontName='NanumReg', fontSize=10, leading=12, alignment=1, spaceAfter=4)

    story = [Paragraph('📦 송장번호 바코드 (피킹검증 스캔용)', s_title), Spacer(1, 4*mm)]

    # 2열 그리드로 배치
    rows_data = []
    pair = []
    for sn in shipment_numbers:
        try:
            img = get_barcode_img(str(sn), write_text=False)
            img_buf = io.BytesIO()
            img.save(img_buf, format='PNG')
            img_buf.seek(0)
            rl_img = RLImage(img_buf, width=80*mm, height=22*mm)
            rl_img.hAlign = 'CENTER'
            cell = [Paragraph(f'<b>{sn}</b>', s_num), rl_img]
            cell_tbl = Table([[c] for c in cell], colWidths=[85*mm])
            cell_tbl.setStyle(TableStyle([
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
                ('TOPPADDING', (0,0), (-1,-1), 1),
                ('BOTTOMPADDING', (0,0), (-1,-1), 1),
            ]))
            pair.append(cell_tbl)
            if len(pair) == 2:
                rows_data.append(pair)
                pair = []
        except Exception:
            continue
    if pair:
        pair.append(Paragraph('', s_num))
        rows_data.append(pair)

    if rows_data:
        grid = Table(rows_data, colWidths=[90*mm, 90*mm])
        grid.setStyle(TableStyle([
            ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#cccccc')),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cccccc')),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('LEFTPADDING', (0,0), (-1,-1), 4),
            ('RIGHTPADDING', (0,0), (-1,-1), 4),
        ]))
        story.append(grid)

    doc.build(story)
    buf.seek(0)
    return buf


def create_box_labels_pdf(box_entries):
    """폼텍 3100 (38.1 x 21.2mm, 5열 13행 = 65칸/페이지) 라벨 PDF 생성

    box_entries: [(box_num, total_qty, size_label), ...] 또는 [(box_num, info_text), ...]
                 각 라벨에 표시할 박스 정보 리스트
    """
    from reportlab.pdfgen import canvas as _canvas
    from reportlab.lib.pagesizes import A4 as _A4
    from reportlab.lib.units import mm as _mm

    # 폼텍 3100 사양 (38.1 × 21.2mm, 5×13 = 65칸)
    LABEL_W = 38.1 * _mm
    LABEL_H = 21.2 * _mm
    X_GAP = 2.54 * _mm      # 라벨 사이 가로 간격 (폼텍 3100 표준)
    Y_GAP = 0 * _mm         # 세로 간격 없음 (라벨이 붙어있음)
    COLS = 5
    ROWS = 13
    PER_PAGE = COLS * ROWS  # 65
    # A4: 210 x 297mm
    # 좌우 여백: (210 - (5*38.1 + 4*2.54)) / 2 = (210 - 200.66) / 2 = 4.67mm
    # 상하 여백: (297 - 13*21.2) / 2 = 10.7mm
    total_w = COLS * LABEL_W + (COLS - 1) * X_GAP
    total_h = ROWS * LABEL_H + (ROWS - 1) * Y_GAP
    LEFT_MARGIN = (210 * _mm - total_w) / 2
    TOP_MARGIN = (297 * _mm - total_h) / 2

    from reportlab.lib.utils import ImageReader as _ImageReader

    buf = io.BytesIO()
    c = _canvas.Canvas(buf, pagesize=_A4)
    page_w, page_h = _A4

    for idx, entry in enumerate(box_entries):
        page_idx = idx // PER_PAGE
        slot_idx = idx % PER_PAGE
        if slot_idx == 0 and idx > 0:
            c.showPage()

        col = slot_idx % COLS
        row = slot_idx // COLS
        # 프린터 오프셋 보정:
        # - 1, 2열만 왼쪽으로 10% 이동 (3~5열은 그대로)
        # - 전체 위로 25% 이동 (기존 10% + 추가 15%)
        x_offset = -LABEL_W * 0.1 if col < 2 else 0
        y_offset = LABEL_H * 0.25  # PDF는 위로 갈수록 y 증가
        x = LEFT_MARGIN + col * (LABEL_W + X_GAP) + x_offset
        # 좌표 변환: PDF는 좌하단 원점, 라벨은 좌상단부터 채움
        y_top = page_h - TOP_MARGIN - row * (LABEL_H + Y_GAP) + y_offset
        y = y_top - LABEL_H

        # 라벨 내용 파싱
        if isinstance(entry, (tuple, list)):
            box_num = str(entry[0])
            qty = entry[1] if len(entry) > 1 else None
            size_label = entry[2] if len(entry) > 2 else None
        else:
            box_num = str(entry)
            qty = None
            size_label = None

        # 사이즈 라벨에서 한글만 추출 (이모지 제거): 🟢대 → 대
        size_char = ''
        if size_label:
            for ch in size_label:
                if ch in ('대', '중', '소'):
                    size_char = ch
                    break

        # 라벨 레이아웃:
        # ┌─────────────────────┐
        # │ 대   1번    65개    │  ← 상단: 정보
        # │ |||||||||||||||||   │  ← 중하단: 바코드
        # └─────────────────────┘

        # 상단 좌측: 사이즈
        if size_char:
            c.setFont('NanumBold', 11)
            c.drawString(x + LABEL_W * 0.05, y + LABEL_H * 0.68, size_char)

        # 상단 중앙: 박스 번호 (가장 큰 글씨)
        c.setFont('NanumBold', 16)
        c.drawCentredString(x + LABEL_W * 0.48, y + LABEL_H * 0.66,
                            f'{box_num}번')

        # 상단 우측: 수량
        if qty is not None:
            c.setFont('NanumReg', 7)
            c.drawRightString(x + LABEL_W - LABEL_W * 0.05,
                              y + LABEL_H * 0.72, f'{qty}개')

        # 중하단: Code128 바코드 (#N 형식)
        try:
            barcode_text = f'#{box_num}'
            bc_img = get_barcode_img(barcode_text, write_text=False)
            bc_buf = io.BytesIO()
            bc_img.save(bc_buf, format='PNG')
            bc_buf.seek(0)
            bc_w = LABEL_W * 0.85
            bc_h = LABEL_H * 0.42
            c.drawImage(
                _ImageReader(bc_buf),
                x + (LABEL_W - bc_w) / 2,
                y + LABEL_H * 0.1,
                width=bc_w,
                height=bc_h,
                preserveAspectRatio=False,
                mask='auto',
            )
            # 바코드 아래 텍스트
            c.setFont('NanumReg', 6)
            c.drawCentredString(x + LABEL_W / 2, y + LABEL_H * 0.02, barcode_text)
        except Exception:
            pass

    c.save()
    buf.seek(0)
    return buf


def create_multi_trigger_label_pdf():
    """다량 입력 트리거 바코드(#MULTI) 폼텍 3100 형식 A4 1장 (65칸 동일 바코드)"""
    # create_box_labels_pdf 재활용: 65개 동일 엔트리
    entries = [('MULTI', '', '다량')] * 65
    return create_box_labels_pdf(entries)


# ══════════════════════════════════════════════════════
# Streamlit UI
# ══════════════════════════════════════════════════════
st.set_page_config(page_title='쿠썸 로켓배송 운영 관리', page_icon='🚀', layout='centered')
st.markdown("""<style>
    .block-container { max-width: 58rem !important; }
    /* 피킹 스캔 결과 피드백 */
    .scan-ok {
        background: #d4edda; border-left: 6px solid #28a745;
        padding: 1.2rem 1.5rem; border-radius: 8px; margin: 0.5rem 0; color: #155724;
    }
    .scan-error {
        background: #f8d7da; border-left: 6px solid #dc3545;
        padding: 1.2rem 1.5rem; border-radius: 8px; margin: 0.5rem 0; color: #721c24;
        animation: shake 0.5s ease-in-out;
    }
    .scan-warning {
        background: #fff3cd; border-left: 6px solid #ffc107;
        padding: 1.2rem 1.5rem; border-radius: 8px; margin: 0.5rem 0; color: #856404;
    }
    .scan-complete {
        background: #cce5ff; border-left: 6px solid #007bff;
        padding: 1.2rem 1.5rem; border-radius: 8px; margin: 0.5rem 0; color: #004085;
    }
    .scan-shortage {
        background: #e2e3f1; border-left: 6px solid #6c63ff;
        padding: 1.2rem 1.5rem; border-radius: 8px; margin: 0.5rem 0; color: #383467;
    }
    @keyframes shake {
        0%, 100% { transform: translateX(0); }
        20% { transform: translateX(-10px); }
        40% { transform: translateX(10px); }
        60% { transform: translateX(-6px); }
        80% { transform: translateX(6px); }
    }
    .shipment-input {
        background: #f0f2f6; padding: 2rem; border-radius: 12px; text-align: center;
    }
</style>""", unsafe_allow_html=True)
st.title('🚀 쿠썸 로켓배송 운영 관리')
st.caption('엑셀 파일을 업로드하면 바코드 이미지를 자동으로 삽입합니다')

# ══════════════════════════════════════════════════════
# 피킹 검증 시스템 — 헬퍼 함수 & 설정
# ══════════════════════════════════════════════════════
PICKING_CONFIG = {
    "SERVICE_ACCOUNT_FILE": "service_account.json",
}

def _extract_sheet_id(url_or_id):
    """구글 시트 URL 또는 ID에서 스프레드시트 ID만 추출"""
    m = re.search(r'/d/([a-zA-Z0-9_-]+)', url_or_id)
    if m:
        return m.group(1)
    return url_or_id.strip()

@st.cache_resource(ttl=600)
def get_gsheet_client():
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        # 1순위: Streamlit Secrets (Cloud 배포용)
        if "gcp_service_account" in st.secrets:
            creds = Credentials.from_service_account_info(
                dict(st.secrets["gcp_service_account"]), scopes=scopes
            )
        # 2순위: 로컬 JSON 파일
        else:
            creds = Credentials.from_service_account_file(
                PICKING_CONFIG["SERVICE_ACCOUNT_FILE"], scopes=scopes
            )
        return gspread.authorize(creds)
    except FileNotFoundError:
        return None
    except Exception as e:
        st.warning(f"구글 시트 연결 실패: {e}")
        return None

def pick_load_sheet_as_df(client, sheet_url, tab_name):
    try:
        import pandas as _pd
        sheet_id = _extract_sheet_id(sheet_url)
        spreadsheet = client.open_by_key(sheet_id)
        worksheet = spreadsheet.worksheet(tab_name)
        # 중복 헤더 대응: get_all_values로 읽고 직접 DataFrame 생성
        all_values = worksheet.get_all_values()
        if len(all_values) < 2:
            return _pd.DataFrame()
        headers = all_values[0]
        # 중복 컬럼명 처리: 같은 이름이면 _2, _3 붙임
        seen = {}
        unique_headers = []
        for h in headers:
            h = str(h).strip()
            if h == '':
                h = f'_unnamed_{len(unique_headers)}'
            if h in seen:
                seen[h] += 1
                unique_headers.append(f"{h}_{seen[h]}")
            else:
                seen[h] = 1
                unique_headers.append(h)
        df = _pd.DataFrame(all_values[1:], columns=unique_headers)
        return df
    except Exception as e:
        st.error(f"시트 '{tab_name}' 로드 실패: {e}")
        return None

def pick_append_log(client, sheet_url, log_entry):
    try:
        sheet_id = _extract_sheet_id(sheet_url)
        spreadsheet = client.open_by_key(sheet_id)
        try:
            ws = spreadsheet.worksheet("피킹로그")
        except Exception:
            ws = spreadsheet.add_worksheet(title="피킹로그", rows=1000, cols=10)
            ws.append_row(["시간","송장번호","바코드","상품명","결과","스캔수량","필요수량","회차기호","박스번호"])
        ws.append_row(log_entry)
        return True
    except Exception:
        return False

def pick_update_sheet_inventory(client, sheet_url, tab_name, barcode, decrement=1, box_number=None):
    """배대지 시트 스캔 수량 기록.
    매칭: A열(박스번호, 0) + E열(바코드, 4) 둘 다 같은 행.
    기록:
      - U열(스캔수량, 20): 기존값에 decrement 만큼 누적
      - V열(남은수량, 21): G열 수량 - U열 스캔수량 (안 온 상품 체크용)
    G열(수량)은 건드리지 않음 — 원본 주문 수량으로 보존.
    box_number 없으면 레거시 동작(바코드만 매칭) 유지.
    """
    try:
        sheet_id = _extract_sheet_id(sheet_url)
        spreadsheet = client.open_by_key(sheet_id)
        ws = spreadsheet.worksheet(tab_name)
        all_values = ws.get_all_values()
        if len(all_values) < 2:
            return False
        _BOX_COL = 0   # A열
        _BC_COL = 4    # E열 (바코드)
        _QTY_COL = 6   # G열 (원본 수량)
        _SCAN_COL = 20 # U열 (누적 스캔)
        _REM_COL = 21  # V열 (남은 수량)

        _norm_box = str(box_number or '').strip().upper()
        for row_idx in range(1, len(all_values)):
            row = all_values[row_idx]
            if len(row) <= _BC_COL:
                continue
            row_bc = str(row[_BC_COL]).strip()
            if row_bc != barcode:
                continue
            if _norm_box:
                row_box = str(row[_BOX_COL]).strip().upper() if len(row) > _BOX_COL else ''
                if row_box != _norm_box:
                    continue
            # G열 수량 (원본)
            qty_raw = row[_QTY_COL] if len(row) > _QTY_COL else ''
            try:
                qty_orig = int(float(str(qty_raw).strip() or '0'))
            except (ValueError, TypeError):
                qty_orig = 0
            # U열 기존 스캔 수량
            prev_raw = row[_SCAN_COL] if len(row) > _SCAN_COL else ''
            try:
                prev_scan = int(float(str(prev_raw).strip() or '0'))
            except (ValueError, TypeError):
                prev_scan = 0
            new_scan = max(0, prev_scan + int(decrement))
            remaining = qty_orig - new_scan  # 안 온 상품 = 음수 가능(초과스캔 표식)
            ws.update_cell(row_idx + 1, _SCAN_COL + 1, new_scan)
            ws.update_cell(row_idx + 1, _REM_COL + 1, remaining)
            return True
        return False
    except Exception:
        return False


def pick_update_check_qty(client, sheet_url, tab_name, barcode, ship_num, scanned_qty):
    """출고확인 시트의 해당 행 L열(확인 수량)에 스캔된 수량 기록.
    매칭: F열(바코드) + I열(송장번호)
    """
    try:
        sheet_id = _extract_sheet_id(sheet_url)
        spreadsheet = client.open_by_key(sheet_id)
        ws = spreadsheet.worksheet(tab_name)
        all_values = ws.get_all_values()
        if len(all_values) < 2:
            return False
        # F열(5) = 바코드, I열(8) = 송장번호, L열(11) = 확인 수량 → 12번째 열
        for row_idx in range(1, len(all_values)):
            row = all_values[row_idx]
            row_bc = str(row[5]).strip() if len(row) > 5 else ''
            row_ship = str(row[8]).strip() if len(row) > 8 else ''
            if row_bc == barcode and (not ship_num or row_ship == ship_num):
                ws.update_cell(row_idx + 1, 12, scanned_qty)
                return True
        return False
    except Exception:
        return False


_SHEET_SHIP_COL_IDX = 8   # I열 (송장번호), 0-based
_SHEET_QTY_COL_IDX = 11   # L열 (확인수량), 0-based → update_cell에는 +1=12
_SHEET_BOX_COL_IDX = 12   # M열 (출고박스번호), 0-based → update_cell에는 +1=13
_SHEET_BOX_COL_LETTER = 'M'


def pick_read_box_numbers(client, sheet_url, tab_name):
    """출고확인 시트에서 송장번호(I열) → 출고박스번호(M열) 매핑을 읽음.
    M열이 비어있거나 숫자가 아닌 행은 제외.
    반환: {송장번호: 박스번호(int)} (성공, 빈 dict 포함 가능)
         / None (API 호출 실패 — 읽기 자체가 안 됨)
    """
    try:
        sheet_id = _extract_sheet_id(sheet_url)
        spreadsheet = client.open_by_key(sheet_id)
        ws = spreadsheet.worksheet(tab_name)
        all_values = ws.get_all_values()
    except Exception:
        return None
    if len(all_values) < 2:
        return {}
    mapping = {}
    for row in all_values[1:]:
        ship = str(row[_SHEET_SHIP_COL_IDX]).strip() if len(row) > _SHEET_SHIP_COL_IDX else ''
        box = str(row[_SHEET_BOX_COL_IDX]).strip() if len(row) > _SHEET_BOX_COL_IDX else ''
        if not ship or not box:
            continue
        try:
            # "36" / " 36 " / "36.0" (엑셀 float 표기) 모두 허용
            n = int(float(box))
            if n > 0:
                mapping[ship] = n
        except (ValueError, TypeError):
            continue
    return mapping


def pick_write_box_numbers(client, sheet_url, tab_name, ship_to_box, only_empty=True):
    """출고확인 시트 M열에 송장별 박스번호 기록 (batch).
    only_empty=True이면 M열이 비어있는 행만 쓰기(기존 값 보존).
    반환: 기록된 셀 수 / -1 (API 실패 시)
    """
    if not ship_to_box:
        return 0
    try:
        sheet_id = _extract_sheet_id(sheet_url)
        spreadsheet = client.open_by_key(sheet_id)
        ws = spreadsheet.worksheet(tab_name)
        all_values = ws.get_all_values()
        if len(all_values) < 2:
            return 0
        updates = []
        for row_idx in range(1, len(all_values)):
            row = all_values[row_idx]
            ship = str(row[_SHEET_SHIP_COL_IDX]).strip() if len(row) > _SHEET_SHIP_COL_IDX else ''
            if not ship or ship not in ship_to_box:
                continue
            current_m = str(row[_SHEET_BOX_COL_IDX]).strip() if len(row) > _SHEET_BOX_COL_IDX else ''
            if only_empty and current_m:
                continue
            updates.append({
                'range': f'{_SHEET_BOX_COL_LETTER}{row_idx + 1}',
                'values': [[str(ship_to_box[ship])]],
            })
        if updates:
            ws.batch_update(updates)
        return len(updates)
    except Exception:
        return -1


def stock_update_barcode(client, sheet_url, tab_name, barcode, qty, location):
    """등록상품정보 시트에서 바코드(D열, index 3) 매칭 후 재고/위치 업데이트.
    - AB열(index 27): 기존 재고 + qty 누적
    - AC열(index 28): 위치 추가 (중복이면 스킵, 여러 위치는 ", " 구분)
    - C열(index 2): 상품명 반환용
    반환: dict {ok: bool, name: str, new_stock: int, error: str, row_idx: int}
    """
    _BC_COL = 3       # D열
    _NAME_COL = 2     # C열
    _STOCK_COL = 27   # AB열 (28번째)
    _LOC_COL = 28     # AC열 (29번째)

    try:
        qty_int = int(qty)
    except (ValueError, TypeError):
        qty_int = 1
    if qty_int <= 0:
        return {'ok': False, 'name': '', 'new_stock': 0, 'error': '수량은 1 이상이어야 함'}

    # 시트 데이터 캐시 (첫 호출에서만 전체 읽기, 이후 캐시 사용)
    _cache_key = '_stock_sheet_cache'
    all_values = st.session_state.get(_cache_key)
    if all_values is None:
        try:
            sheet_id = _extract_sheet_id(sheet_url)
            spreadsheet = client.open_by_key(sheet_id)
            ws = spreadsheet.worksheet(tab_name)
            all_values = ws.get_all_values()
            st.session_state[_cache_key] = all_values
            st.session_state['_stock_ws'] = ws
        except Exception as e:
            return {'ok': False, 'name': '', 'new_stock': 0, 'error': f'시트 열기 실패: {e}'}
    if len(all_values) < 2:
        return {'ok': False, 'name': '', 'new_stock': 0, 'error': '시트가 비어있음'}

    ws = st.session_state.get('_stock_ws')
    if ws is None:
        try:
            sheet_id = _extract_sheet_id(sheet_url)
            ws = client.open_by_key(sheet_id).worksheet(tab_name)
            st.session_state['_stock_ws'] = ws
        except Exception as e:
            return {'ok': False, 'name': '', 'new_stock': 0, 'error': f'워크시트 열기 실패: {e}'}

    for row_idx in range(1, len(all_values)):
        row = all_values[row_idx]
        if len(row) <= _BC_COL:
            continue
        if str(row[_BC_COL]).strip() != str(barcode).strip():
            continue

        # AB열 기존 재고 + qty
        prev_raw = row[_STOCK_COL] if len(row) > _STOCK_COL else ''
        try:
            prev = int(float(str(prev_raw).strip() or '0'))
        except (ValueError, TypeError):
            prev = 0
        new_stock = prev + qty_int

        # AC열 위치 (중복 방지)
        existing_loc = str(row[_LOC_COL]).strip() if len(row) > _LOC_COL else ''
        loc_new = str(location or '').strip()
        updated_loc = existing_loc
        if loc_new:
            if not existing_loc:
                updated_loc = loc_new
            else:
                parts = [s.strip() for s in existing_loc.split(',') if s.strip()]
                if loc_new not in parts:
                    parts.append(loc_new)
                    updated_loc = ', '.join(parts)

        # 캐시 즉시 반영 (다음 스캔에서 누적 정확하게)
        while len(all_values[row_idx]) <= max(_STOCK_COL, _LOC_COL):
            all_values[row_idx].append('')
        all_values[row_idx][_STOCK_COL] = str(new_stock)
        if updated_loc != existing_loc:
            all_values[row_idx][_LOC_COL] = updated_loc

        # 백그라운드 시트 쓰기 (UI 블로킹 없음)
        import threading
        _ws = ws
        _r = row_idx + 1
        _ns = new_stock
        _ul = updated_loc
        _loc_changed = (updated_loc != existing_loc)
        def _bg():
            try:
                _ws.update_cell(_r, _STOCK_COL + 1, _ns)
                if _loc_changed:
                    _ws.update_cell(_r, _LOC_COL + 1, _ul)
            except Exception:
                pass
        threading.Thread(target=_bg, daemon=True).start()

        name = str(row[_NAME_COL]).strip() if len(row) > _NAME_COL else ''
        return {'ok': True, 'name': name, 'new_stock': new_stock, 'error': ''}
    return {'ok': False, 'name': '', 'new_stock': 0, 'error': f'미등록 바코드: {barcode}'}


def pick_parse_box(box_str):
    import pandas as _pd
    if _pd.isna(box_str) or str(box_str).strip() == "":
        return {"기호": None, "박스": None, "수량": None, "상태": "알수없음"}
    box_str = str(box_str).strip()
    match = re.match(r"((?:국내)?부족)\((-?\d+)\)", box_str)
    if match:
        return {"기호": match.group(1), "박스": None, "수량": int(match.group(2)), "상태": "부족"}
    # 영문+숫자 형식 (예: W1, M3) + 수량
    match = re.match(r"([●★■▲◆◇○□△▼♦♠♣♥☆※·]+)([A-Za-z]*\d+)\((\d+)\)", box_str)
    if match:
        return {"기호": match.group(1), "박스": match.group(2).upper(), "수량": int(match.group(3)), "상태": "피킹가능"}
    match = re.match(r"([●★■▲◆◇○□△▼♦♠♣♥☆※·]+)([A-Za-z]*\d+)", box_str)
    if match:
        return {"기호": match.group(1), "박스": match.group(2).upper(), "수량": None, "상태": "피킹가능"}
    match = re.match(r"(국내재고)\((\d+)\)", box_str)
    if match:
        return {"기호": match.group(1), "박스": None, "수량": int(match.group(2)), "상태": "피킹가능"}
    if box_str == "국내재고":
        return {"기호": "국내재고", "박스": None, "수량": None, "상태": "피킹가능"}
    return {"기호": box_str, "박스": None, "수량": None, "상태": "알수없음"}

def pick_clean_출고(df):
    import pandas as _pd
    df = df.copy()
    required = ["바코드", "상품명", "수량", "쉽먼트운송장번호"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        st.error(f"출고지시서 필수 컬럼 누락: {missing}")
        return None
    df["수량"] = _pd.to_numeric(df["수량"], errors="coerce").fillna(0).astype(int)
    df["쉽먼트운송장번호"] = df["쉽먼트운송장번호"].astype(str).str.replace(r"\.0$", "", regex=True)
    df["바코드"] = df["바코드"].astype(str).str.strip()
    # 확인 수량(L열) 보존 - 시트 재로드 시 진행 상태 복원용
    check_col = None
    for c in df.columns:
        c_norm = str(c).strip().replace(' ', '')
        if c_norm in ('확인수량', '확인량'):
            check_col = c
            break
    if check_col is not None:
        df["확인수량"] = _pd.to_numeric(df[check_col], errors="coerce").fillna(0).astype(int)
    elif len(df.columns) >= 12:
        # 헤더가 비어있거나 다른 이름이면 L열(12번째) 위치로 시도
        df["확인수량"] = _pd.to_numeric(df.iloc[:, 11], errors="coerce").fillna(0).astype(int)
    else:
        df["확인수량"] = 0
    if "박스번호" in df.columns:
        parsed = df["박스번호"].apply(pick_parse_box)
        df["회차기호"] = parsed.apply(lambda x: x["기호"])
        df["박스넘버"] = parsed.apply(lambda x: x["박스"])
        df["박스내수량"] = parsed.apply(lambda x: x["수량"])
        df["피킹상태"] = parsed.apply(lambda x: x["상태"])
    return df

def pick_clean_배대지(df):
    import pandas as _pd
    df = df.copy()
    if "바코드" not in df.columns:
        st.error("배대지 시트에 '바코드' 컬럼이 없습니다")
        return None
    df["바코드"] = df["바코드"].astype(str).str.strip()
    if "수량" in df.columns:
        df["수량"] = _pd.to_numeric(df["수량"], errors="coerce").fillna(0).astype(int)
    if "배대지주문수량" in df.columns:
        df["배대지주문수량"] = _pd.to_numeric(df["배대지주문수량"], errors="coerce").fillna(0).astype(int)
    if "박스번호" in df.columns:
        parsed = df["박스번호"].apply(pick_parse_box)
        df["회차기호"] = parsed.apply(lambda x: x["기호"])
        df["박스넘버"] = parsed.apply(lambda x: x["박스"])
    return df

def pick_init_session():
    defaults = {
        "pick_df_출고": None, "pick_df_배대지": None,
        "pick_selected_shipment": None, "pick_selected_shipments": [], "pick_show_add_input": False, "pick_picking_state": {},
        "pick_inventory_state": {}, "pick_scan_log": [],
        "pick_last_scan_result": None, "pick_scan_counter": 0,
        "pick_completed_shipments": set(), "pick_shortage_items": [],
        "pick_data_loaded": False, "pick_gsheet_client": None,
        "pick_use_gsheet": False,
        "pick_sheet_url_출고": "", "pick_sheet_tab_출고": "",
        "pick_sheet_url_배대지": "", "pick_sheet_tab_배대지": "",
    }
    for key, val in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = val

pick_init_session()

def pick_load_all_data(url_출고, tab_출고, url_배대지="", tab_배대지=""):
    import pandas as _pd
    client = get_gsheet_client()
    if client:
        st.session_state.pick_gsheet_client = client
        st.session_state.pick_use_gsheet = True
        # 시트 재로드 시 M열 캐시 초기화 (발주 변동 반영)
        for _k in list(st.session_state.keys()):
            if _k.startswith("_pick_existing_box_") or _k.startswith("_pick_box_written_"):
                del st.session_state[_k]
        # 출고지시서(쉽먼트시트) 로드
        if url_출고 and tab_출고:
            df_출고 = pick_load_sheet_as_df(client, url_출고, tab_출고)
            if df_출고 is not None and not df_출고.empty:
                st.session_state.pick_df_출고 = pick_clean_출고(df_출고)
                st.session_state.pick_sheet_url_출고 = url_출고
                st.session_state.pick_sheet_tab_출고 = tab_출고
        # 배대지 입고 로드
        if url_배대지 and tab_배대지:
            df_배대지 = pick_load_sheet_as_df(client, url_배대지, tab_배대지)
            if df_배대지 is not None and not df_배대지.empty:
                st.session_state.pick_df_배대지 = pick_clean_배대지(df_배대지)
                st.session_state.pick_sheet_url_배대지 = url_배대지
                st.session_state.pick_sheet_tab_배대지 = tab_배대지
        if st.session_state.pick_df_출고 is not None:
            st.session_state.pick_data_loaded = True
            return True
    return False

def pick_init_inventory():
    df = st.session_state.pick_df_배대지
    if df is None or df.empty:
        return
    inventory = {}
    for _, row in df.iterrows():
        barcode = row["바코드"]
        symbol = row.get("회차기호", "기타")
        qty = row.get("수량", 0)
        key = (symbol, barcode)
        inventory[key] = inventory.get(key, 0) + qty
    st.session_state.pick_inventory_state = inventory

def pick_init_picking(shipment_ids):
    """단일 또는 다중 쉽먼트 ID를 받아 피킹 초기화.
    다중일 경우 각 바코드별로 어느 쉽먼트(박스)에 속하는지 추적."""
    if isinstance(shipment_ids, str):
        shipment_ids = [shipment_ids]

    df = st.session_state.pick_df_출고
    picking = {}
    shortage_items = []

    for ship_idx, shipment_id in enumerate(shipment_ids, start=1):
        shipment_df = df[df["쉽먼트운송장번호"] == shipment_id]
        if shipment_df.empty:
            st.error(f"쉽먼트 {shipment_id}를 찾을 수 없습니다")
            continue
        ship_label = f"{ship_idx}번박스"
        for _, row in shipment_df.iterrows():
            bc = row["바코드"]
            symbol = row.get("회차기호", "")
            qty = row["수량"]
            pick_status = row.get("피킹상태", "피킹가능")
            if pick_status == "부족":
                shortage_items.append({
                    "바코드": bc, "상품명": row["상품명"],
                    "부족수량": abs(row.get("박스내수량", 0) or 0),
                    "박스번호": row.get("박스번호", ""),
                    "쉽먼트박스": ship_label,
                })
                continue
            if bc in picking:
                picking[bc]["필요수량"] += qty
                # 같은 바코드가 여러 쉽먼트에 있으면 박스 라벨을 합침
                if ship_label not in picking[bc]["쉽먼트박스목록"]:
                    picking[bc]["쉽먼트박스목록"].append(ship_label)
                # 송장별 필요수량 추적
                picking[bc]["송장별수량"][ship_label] = picking[bc]["송장별수량"].get(ship_label, 0) + qty
            else:
                inv_key = (symbol, bc)
                inv_qty = st.session_state.pick_inventory_state.get(inv_key, None)
                picking[bc] = {
                    "상품명": row["상품명"], "필요수량": qty, "스캔수량": 0,
                    "회차기호": symbol if symbol else "N/A",
                    "박스번호": row.get("박스번호", ""), "박스넘버": row.get("박스넘버", ""),
                    "박스내수량": row.get("박스내수량", None), "배대지잔여": inv_qty,
                    "SKU_ID": row.get("SKU ID", ""), "물류센터": row.get("물류센터(FC)", ""),
                    "쉽먼트박스목록": [ship_label],
                    "송장별수량": {ship_label: qty},
                }

    st.session_state.pick_picking_state = picking
    st.session_state.pick_shortage_items = shortage_items
    st.session_state.pick_selected_shipment = " + ".join(shipment_ids) if len(shipment_ids) > 1 else shipment_ids[0]
    st.session_state.pick_selected_shipments = shipment_ids
    st.session_state.pick_scan_log = []
    st.session_state.pick_last_scan_result = None
    st.session_state.pick_scan_counter = 0
    # 다량 모드 상태 초기화 (이전 쉽먼트의 모드가 끌려오지 않도록)
    st.session_state.pick_next_qty = 1
    st.session_state.pick_qty_input_mode = False

def pick_process_scan(barcode, qty=1):
    barcode = barcode.strip()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    state = st.session_state.pick_picking_state
    inventory = st.session_state.pick_inventory_state

    # #MULTI 트리거 → 수량 입력 모드 진입
    if barcode.upper() == "#MULTI":
        st.session_state.pick_qty_input_mode = True
        st.session_state.pick_next_qty = 1
        result = {"status": "multi_trigger", "message": "🔢 다량 입력 모드",
                  "detail": "수량을 입력한 후 상품 바코드를 스캔하세요",
                  "barcode": barcode, "상품명": "", "시간": now}
        st.session_state.pick_last_scan_result = result
        st.session_state.pick_scan_counter += 1
        return result

    if barcode not in state:
        df = st.session_state.pick_df_출고
        hint = ""
        if df is not None:
            match = df[df["바코드"] == barcode]
            if not match.empty:
                name = match["상품명"].iloc[0][:25]
                others = match["쉽먼트운송장번호"].unique()[:3]
                hint = f" → [{name}] 다른 쉽먼트에 있음: {', '.join(s[-6:] for s in others)}"
            else:
                hint = " → 출고지시서에 없는 바코드"
        result = {"status": "error", "message": "🚨 오피킹! 이 쉽먼트에 없는 바코드",
                  "detail": f"{barcode}{hint}", "barcode": barcode, "상품명": "", "시간": now}
        st.session_state.pick_scan_log.append(result)
        st.session_state.pick_last_scan_result = result
        st.session_state.pick_scan_counter += 1
        return result

    item = state[barcode]
    qty = max(1, int(qty))
    needed = item["필요수량"]
    prev_scanned = item["스캔수량"]

    # 이미 다 찬 상태에서 추가 스캔 → 전량 초과
    if prev_scanned >= needed:
        item["스캔수량"] += qty
        box_label = ""
        if "송장별수량" in item and item["송장별수량"]:
            box_label = list(item["송장별수량"].keys())[-1]
        result = {"status": "over", "message": f"⚠️ 수량 초과! {item['상품명'][:35]}",
                  "detail": f"필요 {needed}개 전부 스캔됨 (+{qty} 추가)",
                  "barcode": barcode, "상품명": item["상품명"], "시간": now, "박스": box_label}
        st.session_state.pick_scan_log.append(result)
        st.session_state.pick_last_scan_result = result
        st.session_state.pick_scan_counter += 1
        # 다량 모드 1회성 해제
        st.session_state.pick_next_qty = 1
        st.session_state.pick_qty_input_mode = False
        return result

    # 정상 처리: qty 만큼 증가 (초과분 포함)
    item["스캔수량"] += qty
    over_qty = max(0, item["스캔수량"] - needed)
    effective_inc = qty - over_qty  # 실제 유효 스캔 수

    # 박스 라벨: 마지막 유효 유닛 기준
    box_label = ""
    if "송장별수량" in item:
        cumulative = 0
        last_valid_scan = min(item["스캔수량"], needed)
        for ship_label, ship_qty in item["송장별수량"].items():
            cumulative += ship_qty
            if last_valid_scan <= cumulative:
                box_label = ship_label
                break

    # 재고 차감 (유효분만)
    symbol = item["회차기호"]
    inv_key = (symbol, barcode)
    shortage_warning = ""
    sheet_decrement = 0
    if inv_key in inventory and effective_inc > 0:
        available = inventory[inv_key]
        if available >= effective_inc:
            inventory[inv_key] -= effective_inc
            item["배대지잔여"] = inventory[inv_key]
            sheet_decrement = effective_inc
        else:
            sheet_decrement = available
            inventory[inv_key] = 0
            item["배대지잔여"] = 0
            shortage_warning = f" | ⚠ {symbol}회차 배대지 재고 소진!"
    elif inv_key in inventory:
        item["배대지잔여"] = inventory[inv_key]

    remaining = max(0, needed - item["스캔수량"])
    box_msg = f" → {box_label}" if box_label else ""
    qty_label = f" ×{qty}" if qty > 1 else ""
    if over_qty > 0:
        status = "over"
        message = f"⚠️ 수량 초과! {item['상품명'][:35]}{box_msg}"
        detail = f"스캔 {item['스캔수량']}/{needed} (+{over_qty} 초과){shortage_warning}"
    else:
        status = "shortage" if shortage_warning else "ok"
        message = f"✅ {item['상품명'][:35]}{box_msg}{qty_label}"
        detail = f"스캔 {item['스캔수량']}/{needed} (남은: {remaining}){shortage_warning}"

    result = {
        "status": status,
        "message": message,
        "detail": detail,
        "barcode": barcode, "상품명": item["상품명"], "시간": now, "박스": box_label,
        "처리수량": qty,
    }
    st.session_state.pick_scan_log.append(result)
    st.session_state.pick_last_scan_result = result
    st.session_state.pick_scan_counter += 1

    # 다량 모드는 1회성 → 1개 모드로 복귀
    st.session_state.pick_next_qty = 1
    st.session_state.pick_qty_input_mode = False

    if st.session_state.pick_use_gsheet and st.session_state.pick_gsheet_client:
        import threading
        log_row = [now, st.session_state.pick_selected_shipment or "", barcode,
                   item["상품명"][:40], result["status"], item["스캔수량"],
                   item["필요수량"], item.get("회차기호",""), item.get("박스번호","")]
        log_url = st.session_state.pick_sheet_url_출고
        _dec = sheet_decrement
        # 배대지 시트 매칭용 박스번호 (배대지박스 식별자: M1/W3 등)
        _bg_box = str(item.get("박스넘버") or '').strip().upper()
        # 백그라운드 스레드로 구글 시트 업데이트 (스캔 속도 향상)
        def _bg_sheet_update():
            try:
                pick_append_log(st.session_state.pick_gsheet_client, log_url, log_row)
                if result["status"] in ("ok", "shortage") and _dec > 0 and st.session_state.pick_sheet_url_배대지 and st.session_state.pick_sheet_tab_배대지:
                    pick_update_sheet_inventory(
                        st.session_state.pick_gsheet_client,
                        st.session_state.pick_sheet_url_배대지,
                        st.session_state.pick_sheet_tab_배대지,
                        barcode, decrement=_dec, box_number=_bg_box
                    )
            except Exception:
                pass
        threading.Thread(target=_bg_sheet_update, daemon=True).start()
    return result

def pick_get_progress():
    state = st.session_state.pick_picking_state
    if not state:
        return {"total":0,"scanned":0,"skus":0,"done_skus":0,"pct":0.0,"is_complete":False,"over":0,"shortage":0}
    total = sum(v["필요수량"] for v in state.values())
    scanned = sum(min(v["스캔수량"], v["필요수량"]) for v in state.values())
    over = sum(max(0, v["스캔수량"] - v["필요수량"]) for v in state.values())
    skus = len(state)
    done_skus = sum(1 for v in state.values() if v["스캔수량"] >= v["필요수량"])
    shortage = sum(1 for v in state.values()
                   if v.get("배대지잔여") is not None and v["배대지잔여"] == 0 and v["스캔수량"] < v["필요수량"])
    pct = scanned / total if total > 0 else 0
    return {"total":total,"scanned":scanned,"skus":skus,"done_skus":done_skus,
            "pct":pct,"is_complete":scanned>=total,"over":over,"shortage":shortage}

tab1, tab2, tab_shipment, tab8, tab_stock, tab3, tab4, tab5 = st.tabs([
    '📦 소형 라벨',
    '📋 대형 라벨 (90도 회전)',
    '🚛 쉽먼트 관리',
    '📦 피킹 & 분류',
    '📥 재고 확인',
    '📄 출고 작업 지시서 PDF',
    '📎 PDF 병합',
    '📝 발주중단 공문',
])

# 쉽먼트 관리 탭 안에 "통합 관리"와 "재출력"을 서브탭으로 구성
with tab_shipment:
    tab6, tab7 = st.tabs(['🚛 쉽먼트 통합 관리', '🔄 쉽먼트 재출력'])

# ── 소형 탭 ────────────────────────────────────────────
with tab1:
    st.subheader('📋 열 번호 설정')
    st.caption('A=1, B=2, C=3, D=4 ... L=12, M=13')
    c1,c2 = st.columns(2)
    with c1:
        s_col_name     = st.number_input('상품명 열',     min_value=1, max_value=50, value=4,  key='s_name')
        s_col_barcode  = st.number_input('바코드 열',     min_value=1, max_value=50, value=12, key='s_barcode')
        s_col_material = st.number_input('재질 열',       min_value=1, max_value=50, value=13, key='s_material')
    with c2:
        s_col_insert   = st.number_input('이미지 삽입 열', min_value=1, max_value=50, value=13, key='s_insert')
        s_start_row    = st.number_input('시작 행',       min_value=1, max_value=10,  value=2,  key='s_startrow')

    st.divider()
    st.subheader('✍️ 고정 문구')
    s_origin = st.text_input('제조국',   value='제조국 Made in China',             key='s_origin')
    s_age    = st.text_input('사용연령', value='본 제품은 14세 이상 사용가능합니다', key='s_age')

    st.divider()
    s_file = st.file_uploader('📂 엑셀 파일 업로드', type=['xlsx'], key='s_file')

    if st.button('🚀 소형 라벨 생성 시작', type='primary', key='s_btn'):
        if not s_file:
            st.warning('⚠️ 엑셀 파일을 먼저 업로드해주세요!')
        else:
            with st.spinner('처리 중...'):
                settings={
                    'col_name':s_col_name,'col_barcode':s_col_barcode,
                    'col_material':s_col_material,'col_insert':s_col_insert,
                    'start_row':s_start_row,'origin':s_origin,'age':s_age,
                    'insert_w':244,'insert_h':157,'row_height':120,'col_width':38
                }
                output, ok, errors = process_excel(s_file, '소형', settings)
            if errors:
                for e in errors: st.error(e)
            fname = s_file.name.replace('.xlsx','_완성.xlsx')
            st.download_button('⬇️ 완성 파일 다운로드', output, file_name=fname,
                             mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── 대형 탭 ────────────────────────────────────────────
with tab2:
    st.subheader('📋 열 번호 설정')
    st.caption('A=1, B=2, C=3, D=4 ... L=12, M=13')
    c1,c2 = st.columns(2)
    with c1:
        l_col_name     = st.number_input('상품명 열',     min_value=1, max_value=50, value=4,  key='l_name')
        l_col_barcode  = st.number_input('바코드 열',     min_value=1, max_value=50, value=12, key='l_barcode')
        l_col_material = st.number_input('재질 열',       min_value=1, max_value=50, value=13, key='l_material')
    with c2:
        l_col_insert   = st.number_input('이미지 삽입 열', min_value=1, max_value=50, value=13, key='l_insert')
        l_start_row    = st.number_input('시작 행',       min_value=1, max_value=10,  value=2,  key='l_startrow')

    st.divider()
    st.subheader('✍️ 고정 문구')
    l_caution = st.text_area('취급주의', value='취급 상 주의 사항 : 화기에 주의 하세요. 의류의 경우 단독 세탁 권장, 표백제 사용 금지, 다림질 주의, 착용시 손톱이나 날카로운 곳에 긁히지 않도록 주의', height=80, key='l_caution')
    l_addr    = st.text_input('주소/전화', value='표시자 주소 및 전화번호 : (주) 폰이지 서울시 영등포구 영등포로 109, 749호 0507-1311-1108', key='l_addr')
    l_origin  = st.text_input('제조국',   value='제조국 : Made in China',  key='l_origin')
    l_age     = st.text_input('사용연령', value='사용연령 : 만14세이상',    key='l_age')

    st.divider()
    l_file = st.file_uploader('📂 엑셀 파일 업로드', type=['xlsx'], key='l_file')

    if st.button('🚀 대형 라벨 생성 시작', type='primary', key='l_btn'):
        if not l_file:
            st.warning('⚠️ 엑셀 파일을 먼저 업로드해주세요!')
        else:
            with st.spinner('처리 중...'):
                settings={
                    'col_name':l_col_name,'col_barcode':l_col_barcode,
                    'col_material':l_col_material,'col_insert':l_col_insert,
                    'start_row':l_start_row,
                    'fix_list':[l_caution,l_addr,l_origin,l_age],
                    'insert_w':298,'insert_h':208,'row_height':160,'col_width':58
                }
                output, ok, errors = process_excel(l_file, '대형', settings)
            if errors:
                for e in errors: st.error(e)
            fname = l_file.name.replace('.xlsx','_완성.xlsx')
            st.download_button('⬇️ 완성 파일 다운로드', output, file_name=fname,
                             mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── 재고 확인(반출 재고 채우기) 탭 ──────────────────────────
with tab_stock:
    st.header('📥 반출 재고 채우기')
    st.caption('위치를 먼저 지정하고 바코드를 스캔하면 등록상품정보 시트의 AB열(국내재고)에 누적, AC열에 위치 기록')

    _STOCK_SHEET_URL = "https://docs.google.com/spreadsheets/d/1M-r5BfuVRh2dunBsR_6lZZ7f4sH7NSI0B9rbBMuMdTc/edit?gid=0#gid=0"
    _STOCK_SHEET_TAB = "등록상품정보"

    # 세션 초기화
    for _k, _v in [
        ('stock_location', ''),
        ('stock_scan_log', []),
        ('stock_next_qty', 1),
        ('stock_qty_input_mode', False),
        ('stock_scan_counter', 0),
        ('stock_last_result', None),
    ]:
        if _k not in st.session_state:
            st.session_state[_k] = _v

    # ── 시트 연결 ──
    _sclient = st.session_state.get('pick_gsheet_client') or get_gsheet_client()
    if _sclient is None:
        st.error('❌ Google Sheets 인증 실패 — 서비스 계정 키를 확인하세요')
        st.stop()
    st.session_state['pick_gsheet_client'] = _sclient

    _sinfo_c1, _sinfo_c2, _sinfo_c3 = st.columns([3, 1, 1])
    with _sinfo_c1:
        st.text_input('구글 시트', value=_STOCK_SHEET_URL, disabled=True, key='stock_url_display')
    with _sinfo_c2:
        st.text_input('탭 이름', value=_STOCK_SHEET_TAB, disabled=True, key='stock_tab_display')
    with _sinfo_c3:
        st.markdown('<br>', unsafe_allow_html=True)
        if st.button('🔄 시트 새로고침', key='stock_reload', use_container_width=True):
            for _sk in ['_stock_sheet_cache', '_stock_ws']:
                st.session_state.pop(_sk, None)
            st.rerun()

    # ── 1단계: 위치 입력 ──
    st.divider()
    st.markdown('### 📍 1단계: 담을 위치 지정')
    _loc_c1, _loc_c2, _loc_c3 = st.columns([3, 1, 1])
    with _loc_c1:
        _loc_input = st.text_input(
            '위치/박스 (예: G박스, A-1구역)',
            value=st.session_state.stock_location,
            key='stock_location_input',
            placeholder='예: G박스, A-1구역',
        )
    with _loc_c2:
        if st.button('✅ 위치 설정', use_container_width=True, key='stock_set_loc'):
            _cleaned = _loc_input.strip()
            if _cleaned:
                st.session_state.stock_location = _cleaned
                st.success(f'위치: {_cleaned}')
                st.rerun()
            else:
                st.error('위치를 입력하세요')
    with _loc_c3:
        if st.button('🔄 위치 변경', use_container_width=True, key='stock_reset_loc'):
            st.session_state.stock_location = ''
            st.rerun()

    if not st.session_state.stock_location:
        st.info('👆 먼저 위치를 입력하고 "✅ 위치 설정"을 눌러주세요')
        st.stop()

    st.success(f'📍 현재 위치: **{st.session_state.stock_location}**')

    # ── 2단계: 바코드 스캔 (fragment로 감싸서 전체 rerun 없이 스캔 부분만 리렌더) ──
    st.divider()
    _stock_use_fragment = getattr(st, 'fragment', lambda f: f)

    @_stock_use_fragment
    def _stock_scan_fragment():
        def _stock_rerun():
            try:
                st.rerun(scope='fragment')
            except TypeError:
                st.rerun()

        if st.session_state.stock_qty_input_mode:
            st.warning(f'🔢 다량 입력 모드 — 다음 스캔은 **{st.session_state.stock_next_qty}개**로 처리됩니다. `#MULTI` 다시 찍으면 해제.')

        if st.session_state.stock_next_qty > 1:
            st.markdown(
                f'<div style="background:#f59e0b;color:white;padding:0.4rem;border-radius:6px;text-align:center;font-weight:bold">'
                f'📦 다음 스캔: {st.session_state.stock_next_qty}개</div>',
                unsafe_allow_html=True)

        def _stock_process_scan(raw):
            _now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            barcode = str(raw or '').strip()
            if not barcode:
                return None
            if barcode.upper() == '#MULTI':
                st.session_state.stock_qty_input_mode = not st.session_state.stock_qty_input_mode
                if not st.session_state.stock_qty_input_mode:
                    st.session_state.stock_next_qty = 1
                return {
                    'status': 'multi_trigger',
                    'message': '🔢 다량 입력 모드 ON' if st.session_state.stock_qty_input_mode else '1개 모드 복귀',
                    'barcode': barcode, 'name': '', 'qty': 0, 'stock': 0, 'time': _now,
                }
            if st.session_state.stock_qty_input_mode and barcode.isdigit():
                n = int(barcode)
                if n > 0:
                    st.session_state.stock_next_qty = n
                    return {
                        'status': 'qty_set',
                        'message': f'🔢 다음 스캔 수량 = {n}개',
                        'barcode': barcode, 'name': '', 'qty': n, 'stock': 0, 'time': _now,
                    }
            _qty = st.session_state.stock_next_qty if st.session_state.stock_qty_input_mode else 1
            _sclient = st.session_state.get('pick_gsheet_client')
            _res = stock_update_barcode(
                _sclient, _STOCK_SHEET_URL, _STOCK_SHEET_TAB,
                barcode, _qty, st.session_state.stock_location,
            )
            if st.session_state.stock_qty_input_mode:
                st.session_state.stock_next_qty = 1
                st.session_state.stock_qty_input_mode = False
            if _res['ok']:
                return {
                    'status': 'ok',
                    'message': f"✅ [{_res['name'][:30]}] +{_qty} → 재고 {_res['new_stock']}",
                    'barcode': barcode, 'name': _res['name'], 'qty': _qty,
                    'stock': _res['new_stock'], 'time': _now,
                }
            else:
                return {
                    'status': 'error',
                    'message': f"❌ {_res['error']}",
                    'barcode': barcode, 'name': '', 'qty': 0, 'stock': 0, 'time': _now,
                }

        _scan_key = f"stock_scan_{st.session_state.stock_scan_counter}"
        _scan_val = st.text_input(
            '바코드 입력',
            key=_scan_key,
            placeholder='바코드를 스캔하거나 #MULTI 입력...',
            label_visibility='collapsed',
        )
        if _scan_val:
            _result = _stock_process_scan(_scan_val)
            if _result:
                st.session_state.stock_last_result = _result
                st.session_state.stock_scan_log.append(_result)
            st.session_state.stock_scan_counter += 1
            _stock_rerun()

        # 최근 결과 + 음성 안내
        _last = st.session_state.stock_last_result
        _tts_msg = None
        if _last:
            if _last['status'] == 'ok':
                st.success(_last['message'])
                _tts_msg = '재고완료'
            elif _last['status'] == 'error':
                st.error(_last['message'])
                _tts_msg = '다시 찍어주세요'
            else:
                st.info(_last['message'])

        # 최근 5건 간략 로그
        _recent = st.session_state.stock_scan_log[-5:]
        if _recent:
            for _e in reversed(_recent):
                _icon = {'ok': '✅', 'error': '❌', 'multi_trigger': '🔢', 'qty_set': '🔢'}.get(_e['status'], '?')
                st.caption(f"{_icon} {_e['time'][-8:]} | {_e['barcode']} | {_e['message']}")

        # 음성 안내(TTS) + 자동 포커스 (fragment rerun마다 실행)
        # scan_counter를 JS에 삽입해서 매번 고유 HTML 생성 → 캐시 방지
        from streamlit.components.v1 import html as _stock_html
        _tts_js = ''
        if _tts_msg:
            from html import escape as _html_esc
            _tts_js = f"""
            try{{
                window.speechSynthesis.cancel();
                setTimeout(function(){{
                    var u = new SpeechSynthesisUtterance('{_html_esc(_tts_msg)}');
                    u.lang = 'ko-KR'; u.rate = 1.2; u.volume = 1.0;
                    var voices = window.speechSynthesis.getVoices();
                    var koVoice = voices.find(v => v.lang && v.lang.startsWith('ko'));
                    if (koVoice) u.voice = koVoice;
                    window.speechSynthesis.speak(u);
                }}, 80);
            }}catch(e){{}}
            """
            st.session_state.stock_last_result = None

        _uid = st.session_state.stock_scan_counter
        _stock_html(f"""
        <script>
        /* scan_{_uid} */
        (function(){{
            const doc = window.parent.document;
            {_tts_js}

            function findScan(){{
                const inputs = doc.querySelectorAll('input[type="text"]');
                for (const inp of inputs){{
                    if (inp.placeholder && inp.placeholder.includes('#MULTI')) return inp;
                }}
                return null;
            }}
            function focusScan(){{
                const inp = findScan();
                if (inp && doc.activeElement !== inp) {{ inp.focus(); inp.select(); }}
                return inp;
            }}

            // 1) 즉시 + 여러 타이밍에서 포커스 시도
            focusScan();
            [30,80,150,300,500,800,1200,1800].forEach(d => setTimeout(focusScan, d));

            // 2) DOM 변화 감지 (rerun으로 새 input 생성 시 즉시 포커스)
            if (!window.__stockScanObsAttached) {{
                const obs = new MutationObserver(focusScan);
                obs.observe(doc.body, {{childList: true, subtree: true}});
                window.__stockScanObsAttached = true;
            }}

            // 3) 입력이 blur되면 즉시 다시 포커스 (실수로 다른 곳 클릭해도 복귀)
            if (!window.__stockScanBlurAttached) {{
                doc.addEventListener('focusout', function(ev){{
                    const inp = findScan();
                    if (!inp) return;
                    if (ev.target === inp) {{
                        setTimeout(function(){{
                            if (doc.activeElement !== inp) focusScan();
                        }}, 50);
                    }}
                }}, true);
                window.__stockScanBlurAttached = true;
            }}

            // 4) 문자 주입 — 스캔 input 밖에서 입력된 글자를 캡처해 input에 append
            //    "R"이 포커스 잡히기 전에 날아가는 문제 해결
            if (!window.__stockScanKeyAttached) {{
                const win = window.parent;
                const nativeSetter = Object.getOwnPropertyDescriptor(
                    win.HTMLInputElement.prototype, 'value'
                ).set;
                doc.addEventListener('keydown', function(ev){{
                    const inp = findScan();
                    if (!inp) return;
                    if (doc.activeElement === inp) return;  // 이미 포커스면 자연 입력
                    const ae = doc.activeElement;
                    const tag = (ae && ae.tagName || '').toLowerCase();
                    if (tag === 'input' || tag === 'textarea' || (ae && ae.isContentEditable)) return;
                    if (ev.ctrlKey || ev.metaKey || ev.altKey) return;
                    if (ev.key && ev.key.length === 1) {{
                        // 프린터블 문자: input에 직접 주입 + React 감지 이벤트 발송
                        ev.preventDefault();
                        ev.stopPropagation();
                        nativeSetter.call(inp, (inp.value || '') + ev.key);
                        inp.dispatchEvent(new Event('input', {{bubbles: true}}));
                        inp.focus();
                    }} else if (ev.key === 'Enter') {{
                        // Enter: 포커스 이동 후 keydown 재발송
                        ev.preventDefault();
                        ev.stopPropagation();
                        inp.focus();
                        setTimeout(function(){{
                            inp.dispatchEvent(new KeyboardEvent('keydown', {{
                                key: 'Enter', code: 'Enter', keyCode: 13, which: 13, bubbles: true
                            }}));
                        }}, 5);
                    }}
                }}, true);
                window.__stockScanKeyAttached = true;
            }}

            // 5) 탭 다시 활성화 시 포커스
            if (!window.__stockScanVisAttached) {{
                doc.addEventListener('visibilitychange', function(){{
                    if (doc.visibilityState === 'visible') {{
                        [50,200,500].forEach(d => setTimeout(focusScan, d));
                    }}
                }}, true);
                window.__stockScanVisAttached = true;
            }}
        }})();
        </script>
        """, height=0)

    st.markdown('### 🔍 바코드 스캔')
    st.caption('💡 일반 스캔 = +1 / `#MULTI` → 숫자 → 바코드 = 그 수량만큼 +누적')
    _stock_scan_fragment()

    # ── 스캔 로그 ──
    if st.session_state.stock_scan_log:
        st.divider()
        with st.expander(f"📜 스캔 로그 ({len(st.session_state.stock_scan_log)}건)", expanded=True):
            import pandas as _pds
            _log_rows = []
            for e in reversed(st.session_state.stock_scan_log[-100:]):
                _icon = {'ok': '✅', 'error': '❌', 'multi_trigger': '🔢', 'qty_set': '🔢'}.get(e['status'], '?')
                _log_rows.append({
                    '시간': e['time'][-8:],
                    '결과': _icon,
                    '바코드': e['barcode'],
                    '수량': e['qty'] if e['qty'] else '',
                    '누적재고': e['stock'] if e['stock'] else '',
                    '상품명': e['name'][:40],
                })
            st.dataframe(_pds.DataFrame(_log_rows), use_container_width=True, hide_index=True)

            if st.button('🗑️ 로그 초기화', key='stock_clear_log'):
                st.session_state.stock_scan_log = []
                st.session_state.stock_last_result = None
                st.rerun()

    # 자동 포커스 (스캐너 연사 대응 — MutationObserver로 input 생성 즉시 포커스)
    from streamlit.components.v1 import html as _stock_html
    _stock_html("""
    <script>
    (function(){
        const doc = window.parent.document;
        function findScan(){
            const inputs = doc.querySelectorAll('input[type="text"]');
            for (const inp of inputs){
                if (inp.placeholder && inp.placeholder.includes('#MULTI')) return inp;
            }
            return null;
        }
        function focusScan(){
            const inp = findScan();
            if (inp && doc.activeElement !== inp) { inp.focus(); inp.select(); }
        }
        focusScan();
        [100,200,400,800,1200].forEach(d => setTimeout(focusScan, d));
        // DOM 변경 감지하여 새 input 생성 시 즉시 포커스
        const obs = new MutationObserver(function(){
            focusScan();
        });
        obs.observe(doc.body, {childList: true, subtree: true});
        // 5초 후 observer 정리 (성능)
        setTimeout(function(){ obs.disconnect(); }, 5000);
    })();
    </script>
    """, height=0)


# ── 출고 작업 지시서 탭 ───────────────────────────────
with tab3:
    st.subheader('📄 출고 작업 지시서 PDF 생성')
    st.caption('CSV 파일을 업로드하면 그룹별 지시서를 PDF로 생성하고 하나로 병합합니다')

    # CSV 컬럼 안내
    with st.expander('📌 CSV 컬럼 형식 안내', expanded=False):
        st.markdown("""
| 열 번호 | 내용 |
|--------|------|
| B (2번째) | 물류센터 |
| D (4번째) | 입고예정일 |
| F (6번째) | 바코드 |
| G (7번째) | 상품명 |
| H (8번째) | 수량 |
| I (9번째) | 송장번호 |
| J (10번째) | 발주일 |
| K (11번째) | 박스번호 |
| M (13번째) | 위치 |
        """)
        st.info('첫 번째 행은 헤더로 자동 스킵됩니다.')

    st.divider()

    # 그룹화 기준
    st.subheader('🗂️ 그룹화 기준 선택')
    grouping_options = {
        '물류센터': 'logisticsCenter',
        '송장번호': 'shipmentNumber',
        '박스번호': 'boxNumber',
    }
    selected_labels = st.multiselect(
        '그룹화 기준 (복수 선택 가능)',
        options=list(grouping_options.keys()),
        default=['물류센터', '송장번호'],
        key='p_grouping'
    )
    selected_keys = [grouping_options[lbl] for lbl in selected_labels]

    st.divider()

    # 다운로드 옵션
    st.subheader('⬇️ 출력 옵션')
    download_mode = st.radio(
        '다운로드 형태',
        ['📄 전체 병합 PDF (1개 파일)', '🗜️ 그룹별 ZIP (개별 PDF)'],
        key='p_dl_mode'
    )

    st.divider()

    p_file = st.file_uploader('📂 CSV 파일 업로드', type=['csv'], key='p_file')

    if p_file:
        # 파일 분석 미리보기
        raw = p_file.read()
        items = parse_csv_to_items(raw)
        p_file.seek(0)

        if not items:
            st.error('⚠️ CSV에서 데이터를 읽지 못했습니다. 컬럼 형식을 확인해주세요.')
        else:
            if selected_keys:
                grouped = group_items(items, selected_keys)

                # 그룹 미리보기 테이블
                st.markdown(f'**총 {len(grouped)}개 그룹** · {len(items)}개 품목')
                preview_rows = []
                for gk, gitems in grouped.items():
                    preview_rows.append({
                        '그룹 키': gk,
                        '품목 수': len(gitems),
                        '총 수량': sum(i['quantity'] for i in gitems),
                        '입고예정일': gitems[0].get('expectedDate','-'),
                        '입고마감일': calc_deadline(gitems[0].get('expectedDate','')),
                    })
                st.dataframe(preview_rows, use_container_width=True, hide_index=True)

                st.divider()

                if st.button('🚀 PDF 생성 시작', type='primary', key='p_btn'):
                    group_keys = list(grouped.keys())
                    total_g = len(group_keys)
                    progress = st.progress(0)
                    status   = st.empty()

                    pdf_buffers = {}
                    errors = []

                    for i, gk in enumerate(group_keys):
                        try:
                            status.text(f'📝 생성 중: {gk}  ({i+1}/{total_g})')
                            pdf_buf = create_work_order_pdf(gk, grouped[gk])
                            pdf_buffers[gk] = pdf_buf
                        except Exception as e:
                            errors.append(f'[{gk}] 실패: {e}')
                        progress.progress((i+1) / total_g)

                    if errors:
                        for e in errors:
                            st.error(e)

                    if pdf_buffers:
                        status.text(f'✅ {len(pdf_buffers)}개 PDF 생성 완료!')
                        progress.progress(1.0)
                        today = datetime.now().strftime('%Y%m%d_%H%M')

                        if '병합' in download_mode:
                            # 전체 병합 PDF
                            merged = merge_pdfs(list(pdf_buffers.values()))
                            st.download_button(
                                label=f'⬇️ 전체 병합 PDF 다운로드 ({len(pdf_buffers)}페이지)',
                                data=merged,
                                file_name=f'출고지시서_전체_{today}.pdf',
                                mime='application/pdf',
                                key='p_dl_merged'
                            )
                        else:
                            # 개별 ZIP
                            zip_buf = io.BytesIO()
                            with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                                for gk, pdf_b in pdf_buffers.items():
                                    safe_name = gk.replace('/', '_').replace('\\', '_')[:50]
                                    zf.writestr(f'{safe_name}.pdf', pdf_b.read())
                            zip_buf.seek(0)
                            st.download_button(
                                label=f'⬇️ ZIP 다운로드 ({len(pdf_buffers)}개 PDF)',
                                data=zip_buf,
                                file_name=f'출고지시서_{today}.zip',
                                mime='application/zip',
                                key='p_dl_zip'
                            )
            else:
                st.warning('⚠️ 그룹화 기준을 하나 이상 선택해주세요.')

# ══════════════════════════════════════════════════════
# 탭4: PDF 병합
# ══════════════════════════════════════════════════════
with tab4:
    st.header('📎 PDF 병합')
    st.caption('여러 PDF 파일을 하나로 합칩니다. 드래그로 순서 조정 가능')

    uploaded_pdfs = st.file_uploader(
        'PDF 파일 업로드 (여러 개 선택 가능)',
        type=['pdf'],
        accept_multiple_files=True,
        key='merge_pdfs'
    )

    if uploaded_pdfs:
        st.markdown(f'**{len(uploaded_pdfs)}개 파일 선택됨**')
        
        # 파일 순서 표시
        for i, f in enumerate(uploaded_pdfs):
            st.text(f'  {i+1}. {f.name}')

        if st.button('🔗 PDF 병합 시작', type='primary', key='merge_btn'):
            try:
                writer = PdfWriter()
                for pdf_file in uploaded_pdfs:
                    reader = PdfReader(pdf_file)
                    for page in reader.pages:
                        writer.add_page(page)
                
                out_buf = io.BytesIO()
                writer.write(out_buf)
                out_buf.seek(0)
                
                today = datetime.now().strftime('%Y%m%d_%H%M')
                st.success(f'✅ {len(uploaded_pdfs)}개 PDF 병합 완료!')
                st.download_button(
                    label='⬇️ 병합된 PDF 다운로드',
                    data=out_buf,
                    file_name=f'병합_{today}.pdf',
                    mime='application/pdf',
                    key='merge_dl'
                )
            except Exception as e:
                st.error(f'❌ 오류: {e}')

# ══════════════════════════════════════════════════════
# 탭5: PDF → 이미지
# ══════════════════════════════════════════════════════
# PDF→이미지 기능 제거됨
    st.caption('PDF 각 페이지를 PNG 이미지로 변환합니다')

    pdf_to_img_file = st.file_uploader(
        'PDF 파일 업로드',
        type=['pdf'],
        key='pdf_to_img'
    )

    col1, col2 = st.columns(2)
    with col1:
        img_format = st.selectbox('이미지 형식', ['PNG', 'JPEG'], key='img_fmt')
    with col2:
        dpi = st.selectbox('해상도 (DPI)', [72, 150, 200, 300], index=2, key='img_dpi')

    if pdf_to_img_file:
        if st.button('🖼️ 변환 시작', type='primary', key='pdf_img_btn'):
            try:
                import fitz  # PyMuPDF
                
                pdf_bytes = pdf_to_img_file.read()
                doc = fitz.open(stream=pdf_bytes, filetype='pdf')
                total_pages = len(doc)
                
                st.info(f'총 {total_pages}페이지 변환 중...')
                progress = st.progress(0)
                
                img_buffers = []
                for i, page in enumerate(doc):
                    mat = fitz.Matrix(dpi/72, dpi/72)
                    pix = page.get_pixmap(matrix=mat)
                    img_buf = io.BytesIO(pix.tobytes(output=img_format.lower()))
                    img_buffers.append((f'page_{i+1:03d}.{img_format.lower()}', img_buf.getvalue()))
                    progress.progress((i+1)/total_pages)
                
                doc.close()
                
                if len(img_buffers) == 1:
                    st.success('✅ 변환 완료!')
                    st.download_button(
                        label=f'⬇️ 이미지 다운로드',
                        data=img_buffers[0][1],
                        file_name=img_buffers[0][0],
                        mime=f'image/{img_format.lower()}',
                        key='pdf_img_dl'
                    )
                else:
                    # ZIP으로 묶기
                    zip_buf = io.BytesIO()
                    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                        for fname, data in img_buffers:
                            zf.writestr(fname, data)
                    zip_buf.seek(0)
                    today = datetime.now().strftime('%Y%m%d_%H%M')
                    st.success(f'✅ {total_pages}페이지 변환 완료!')
                    st.download_button(
                        label=f'⬇️ 전체 이미지 ZIP 다운로드 ({total_pages}장)',
                        data=zip_buf,
                        file_name=f'pdf_images_{today}.zip',
                        mime='application/zip',
                        key='pdf_img_zip_dl'
                    )
                    
            except ImportError:
                st.error('❌ PyMuPDF 라이브러리가 필요합니다. requirements.txt에 pymupdf를 추가하세요.')
            except Exception as e:
                st.error(f'❌ 오류: {e}')

# ══════════════════════════════════════════════════════
# 탭6: 이미지 → PDF
# ══════════════════════════════════════════════════════
# 이미지→PDF 기능 제거됨
    st.caption('여러 이미지를 하나의 PDF로 변환합니다')

    img_files = st.file_uploader(
        '이미지 파일 업로드 (JPG, PNG, WEBP)',
        type=['jpg', 'jpeg', 'png', 'webp'],
        accept_multiple_files=True,
        key='img_to_pdf'
    )

    col1, col2 = st.columns(2)
    with col1:
        page_size = st.selectbox('페이지 크기', ['A4', 'A3', '원본 비율 유지'], key='page_size')
    with col2:
        margin_mm = st.slider('여백 (mm)', 0, 30, 10, key='img_margin')

    if img_files:
        st.markdown(f'**{len(img_files)}개 이미지 선택됨**')
        
        # 미리보기
        cols = st.columns(min(len(img_files), 4))
        for i, img_f in enumerate(img_files[:4]):
            with cols[i]:
                st.image(img_f, caption=img_f.name, use_container_width=True)
        if len(img_files) > 4:
            st.caption(f'... 외 {len(img_files)-4}개')

        if st.button('📄 PDF 변환 시작', type='primary', key='img_pdf_btn'):
            try:
                from reportlab.lib.pagesizes import A4, A3
                from reportlab.platypus import SimpleDocTemplate, Image as RLImage
                from reportlab.lib.units import mm
                
                out_buf = io.BytesIO()
                
                if page_size == 'A4':
                    ps = A4
                elif page_size == 'A3':
                    ps = A3
                else:
                    ps = None
                
                margin = margin_mm * mm
                
                if ps:
                    doc = SimpleDocTemplate(
                        out_buf,
                        pagesize=ps,
                        leftMargin=margin, rightMargin=margin,
                        topMargin=margin, bottomMargin=margin
                    )
                    w = ps[0] - 2*margin
                    h = ps[1] - 2*margin
                    
                    story = []
                    for i, img_f in enumerate(img_files):
                        img_buf = io.BytesIO(img_f.read())
                        pil_img = Image.open(img_buf)
                        iw, ih = pil_img.size
                        ratio = min(w/iw, h/ih)
                        rw, rh = iw*ratio, ih*ratio
                        img_buf.seek(0)
                        rl_img = RLImage(img_buf, width=rw, height=rh)
                        story.append(rl_img)
                        if i < len(img_files)-1:
                            from reportlab.platypus import PageBreak
                            story.append(PageBreak())
                    
                    doc.build(story)
                else:
                    # 원본 비율 유지 - 첫 이미지 크기로 페이지 설정
                    writer_pdf = PdfWriter()
                    for img_f in img_files:
                        img_buf = io.BytesIO(img_f.read())
                        pil_img = Image.open(img_buf).convert('RGB')
                        iw, ih = pil_img.size
                        single_buf = io.BytesIO()
                        
                        tmp_doc = SimpleDocTemplate(
                            single_buf,
                            pagesize=(iw, ih),
                            leftMargin=margin, rightMargin=margin,
                            topMargin=margin, bottomMargin=margin
                        )
                        pw = iw - 2*margin
                        ph = ih - 2*margin
                        img_buf.seek(0)
                        story = [RLImage(img_buf, width=pw, height=ph)]
                        tmp_doc.build(story)
                        single_buf.seek(0)
                        r = PdfReader(single_buf)
                        for page in r.pages:
                            writer_pdf.add_page(page)
                    
                    writer_pdf.write(out_buf)
                
                out_buf.seek(0)
                today = datetime.now().strftime('%Y%m%d_%H%M')
                st.success(f'✅ {len(img_files)}개 이미지 → PDF 변환 완료!')
                st.download_button(
                    label='⬇️ PDF 다운로드',
                    data=out_buf,
                    file_name=f'images_to_pdf_{today}.pdf',
                    mime='application/pdf',
                    key='img_pdf_dl'
                )
            except Exception as e:
                st.error(f'❌ 오류: {e}')

# ══════════════════════════════════════════════════════
# 탭7: 로켓배송 발주 중단 공문 작성
# ══════════════════════════════════════════════════════
with tab5:
    st.header('📝 로켓배송 발주 중단 공문')
    st.caption('쿠팡 로켓배송 상품 영구적 발주 중단 요청 공문을 자동으로 생성합니다')

    BANNED_KEYWORDS = ['공급가 협의','발주량 협의','가격 인상','단가','시즌 종료','일시적','잠정적']
    REQUIRED_KEYWORDS = ['영구적 생산 중단','영구적 취급 중단','영구적 생산중단','영구적 취급중단']

    col_left, col_right = st.columns([2, 3])

    with col_left:
        st.subheader('📋 공문 정보 입력')

        # 기본 정보
        with st.expander('🏢 업체 정보', expanded=True):
            company_name = st.text_input('업체명 *', placeholder='예: (주)마켓피아', key='gm_company')
            representative = st.text_input('대표이사 성함 *', placeholder='예: 홍길동', key='gm_rep')
            manager_name = st.text_input('담당자명', placeholder='예: 김담당', key='gm_mgr')
            manager_contact = st.text_input('담당자 연락처', placeholder='예: 010-1234-5678', key='gm_contact')

        with st.expander('📄 문서 정보', expanded=True):
            doc_number = st.text_input(
                '문서번호',
                value=f'제 {datetime.now().year}-001호',
                key='gm_docnum'
            )
            doc_date = st.date_input('문서 날짜', value=datetime.now(), key='gm_date')

        with st.expander('✍️ 발주 중단 사유', expanded=True):
            reason_type = st.radio(
                '사유 유형',
                ['생산 중단', '취급 중단', '직접 입력'],
                horizontal=True,
                key='gm_reason_type'
            )

            if reason_type == '생산 중단':
                default_reason = '당사 제조사의 영구적 생산 중단으로 인하여 해당 상품의 지속적인 공급이 불가능하게 되었습니다.'
            elif reason_type == '취급 중단':
                default_reason = '당사의 영구적 취급 중단 결정으로 인하여 해당 상품의 지속적인 공급이 불가능하게 되었습니다.'
            else:
                default_reason = ''

            reason_detail = st.text_area(
                '사유 상세 내용 *',
                value=default_reason,
                height=120,
                placeholder='반드시 "영구적 생산 중단" 또는 "영구적 취급 중단" 문구가 포함되어야 합니다.',
                key='gm_reason'
            )

            # 사유 유효성 검사
            has_banned = any(w in reason_detail for w in BANNED_KEYWORDS)
            has_required = any(w in reason_detail for w in REQUIRED_KEYWORDS)

            if reason_detail:
                if has_banned:
                    banned_found = [w for w in BANNED_KEYWORDS if w in reason_detail]
                    st.error(f'⛔ 금지 키워드 포함: {", ".join(banned_found)}')
                elif not has_required:
                    st.warning('⚠️ "영구적 생산 중단" 또는 "영구적 취급 중단" 문구가 필요합니다')
                else:
                    st.success('✅ 사유 검증 통과')

        with st.expander('🖊️ 직인 이미지 (선택)', expanded=False):
            stamp_file = st.file_uploader('직인 이미지 업로드 (PNG 권장)', type=['png','jpg','jpeg'], key='gm_stamp')
            if stamp_file:
                st.image(stamp_file, width=100)
                stamp_size = st.slider('직인 크기', 50, 200, 80, key='gm_stamp_size')
                stamp_x = st.slider('직인 좌우 위치 (%)', 0, 100, 58, key='gm_stamp_x')
                stamp_y = st.slider('직인 상하 위치 (%)', 0, 100, 50, key='gm_stamp_y')
            else:
                stamp_size = 80
                stamp_x = 58
                stamp_y = 50

        st.subheader('📊 SKU 목록')
        sku_input_type = st.radio('입력 방식', ['엑셀 파일 업로드', '직접 입력'], horizontal=True, key='gm_sku_type')

        sku_list = []

        if sku_input_type == '엑셀 파일 업로드':
            st.caption('A열: SKU ID, B열: 상품명 (1행은 헤더)')
            sku_excel = st.file_uploader('엑셀 파일 업로드', type=['xlsx','xls'], key='gm_excel')
            if sku_excel:
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(sku_excel)
                    ws = wb.active
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[0] and row[1]:
                            sku_list.append({'id': str(row[0]).strip(), 'name': str(row[1]).strip()})
                    if sku_list:
                        st.success(f'✅ {len(sku_list)}개 SKU 로드됨')
                        st.dataframe(sku_list[:5], hide_index=True)
                        if len(sku_list) > 5:
                            st.caption(f'... 외 {len(sku_list)-5}개')
                    else:
                        st.error('SKU 데이터를 읽지 못했습니다')
                except Exception as e:
                    st.error(f'파일 읽기 오류: {e}')
        else:
            st.caption('SKU ID와 상품명을 입력하세요')
            num_skus = st.number_input('SKU 수량', min_value=1, max_value=50, value=3, key='gm_num_sku')
            for i in range(int(num_skus)):
                c1, c2 = st.columns([1, 2])
                with c1:
                    sid = st.text_input(f'SKU ID {i+1}', key=f'gm_sid_{i}')
                with c2:
                    sname = st.text_input(f'상품명 {i+1}', key=f'gm_sname_{i}')
                if sid and sname:
                    sku_list.append({'id': sid, 'name': sname})

    with col_right:
        st.subheader('👁️ 미리보기')

        # 공문 미리보기 HTML 생성
        def make_letter_html(company, rep, mgr, contact, docnum, date, reason, skus, stamp_data=None, stamp_sz=80, stamp_x=58, stamp_y=50):
            date_str = date.strftime('%Y년 %m월 %d일') if hasattr(date, 'strftime') else str(date)
            sku_rows = ''
            for sku in skus:
                sku_rows += f"""
                <tr>
                    <td style="border:1px solid black;padding:6px 8px;text-align:center">{sku['id']}</td>
                    <td style="border:1px solid black;padding:6px 8px">{sku['name']}</td>
                    <td style="border:1px solid black;padding:6px 8px;font-size:7.5pt">{reason}</td>
                </tr>"""
            if not sku_rows:
                sku_rows = '<tr><td colspan="3" style="border:1px solid black;padding:20px;text-align:center;color:#999">엑셀 파일을 업로드해 주세요</td></tr>'

            stamp_html = ''
            if stamp_data:
                import base64
                b64 = base64.b64encode(stamp_data).decode()
                stamp_html = f'<img src="data:image/png;base64,{b64}" style="position:absolute;left:{stamp_x}%;top:{stamp_y}%;transform:translate(-50%,-50%);width:{stamp_sz}px;height:{stamp_sz}px;object-fit:contain;mix-blend-mode:multiply;pointer-events:none"/>'

            return f"""
            <div style="background:white;padding:15px;width:100%;box-sizing:border-box;font-family:serif;font-size:9pt;line-height:1.6;color:black">
                <table style="width:100%;border-collapse:collapse;margin-bottom:8px">
                    <tr><td style="width:110px;font-weight:bold;padding:3px 0">문서번호 :</td><td style="padding:3px 0">{docnum}</td></tr>
                    <tr><td style="font-weight:bold;padding:3px 0">수&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;신 :</td><td style="font-weight:bold;padding:3px 0">쿠팡 주식회사 귀중</td></tr>
                    <tr><td style="font-weight:bold;padding:3px 0">발&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;신 :</td><td style="padding:3px 0"><b>{company or '[업체명]'}</b>&nbsp;/&nbsp;{mgr or '[담당자명]'}&nbsp;{contact or '[연락처]'}</td></tr>
                    <tr><td style="font-weight:bold;padding:3px 0">제&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;목 :</td><td style="font-weight:bold;text-decoration:underline;padding:3px 0">로켓배송 상품 영구적 발주 중단 요청의 건</td></tr>
                </table>
                <hr style="border:1.5px solid black;margin:20px 0"/>
                <p>1. 귀사의 무궁한 발전을 기원합니다.</p>
                <p>2. 당사는 아래와 같은 불가피한 사유(영구적 생산 및 취급 중단)로 인해 해당 상품들의 공급을 지속할 수 없게 되었습니다. 이에 따라 로켓배송 서비스의 안정적인 운영을 위해 해당 제품들의 발주 중단을 정중히 요청드리는 바입니다.</p>
                <div style="text-align:center;font-weight:bold;font-size:13pt;margin:20px 0;letter-spacing:4px">- 아 래 -</div>
                <table style="width:100%;border-collapse:collapse;font-size:10pt;margin-bottom:30px">
                    <thead>
                        <tr style="background:#f3f4f6;text-align:center;font-weight:bold">
                            <th style="border:1px solid black;padding:8px;width:20%">SKU ID</th>
                            <th style="border:1px solid black;padding:8px;width:45%">SKU 명칭</th>
                            <th style="border:1px solid black;padding:8px;width:35%">발주 중단 사유</th>
                        </tr>
                    </thead>
                    <tbody>{sku_rows}</tbody>
                </table>
                <div style="position:relative;text-align:center;margin-top:40px;padding-top:20px;border-top:1px solid #e5e7eb">
                    <div style="font-size:13pt;font-weight:bold;margin-bottom:30px;letter-spacing:2px">{date_str}</div>
                    <h1 style="font-size:20pt;font-weight:bold;letter-spacing:6px;margin-bottom:30px">{company or '[업체명]'}</h1>
                    <div style="display:inline-flex;align-items:center;gap:10px">
                        <span style="font-size:14pt;font-weight:bold;letter-spacing:4px">대표이사&nbsp;&nbsp;{rep or '[성함]'}</span>
                        <span style="font-size:13pt;font-weight:bold">(인)</span>
                    </div>
                    {stamp_html}
                </div>
            </div>"""

        stamp_data = stamp_file.read() if stamp_file else None
        if stamp_data:
            stamp_file.seek(0)

        html = make_letter_html(
            company_name, representative, manager_name, manager_contact,
            doc_number, doc_date, reason_detail, sku_list,
            stamp_data, stamp_size, stamp_x, stamp_y
        )
        components_v1.html(html, height=1000, scrolling=True)

        st.divider()

        # PDF 생성 & 다운로드
        is_valid = (
            company_name and representative and sku_list
            and not has_banned and has_required
        ) if reason_detail else False

        if not is_valid:
            if not company_name:
                st.warning('업체명을 입력해주세요')
            elif not representative:
                st.warning('대표이사 성함을 입력해주세요')
            elif not sku_list:
                st.warning('SKU를 1개 이상 입력해주세요')
            elif not reason_detail:
                st.warning('발주 중단 사유를 입력해주세요')

        if st.button('📄 공문 PDF 생성', type='primary', key='gm_pdf_btn', disabled=not is_valid):
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.lib.units import mm
                from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
                from reportlab.lib.styles import ParagraphStyle
                from reportlab.lib import colors

                buf = io.BytesIO()
                doc = SimpleDocTemplate(
                    buf, pagesize=A4,
                    leftMargin=25*mm, rightMargin=25*mm,
                    topMargin=20*mm, bottomMargin=20*mm
                )

                styles_normal = ParagraphStyle('normal', fontName='NanumReg', fontSize=10, leading=16)
                styles_bold   = ParagraphStyle('bold',   fontName='NanumBold', fontSize=10, leading=16)
                styles_title  = ParagraphStyle('title',  fontName='NanumBold', fontSize=14, leading=20, alignment=1)
                styles_center = ParagraphStyle('center', fontName='NanumBold', fontSize=10, leading=16, alignment=1)
                styles_big    = ParagraphStyle('big',    fontName='NanumBold', fontSize=18, leading=26, alignment=1, spaceAfter=10)
                styles_small  = ParagraphStyle('small',  fontName='NanumReg', fontSize=8, leading=11)

                date_str = doc_date.strftime('%Y년 %m월 %d일')
                story = []

                # 헤더 테이블
                header_data = [
                    [Paragraph('문서번호 :', styles_bold), Paragraph(doc_number, styles_normal)],
                    [Paragraph('수      신 :', styles_bold), Paragraph('쿠팡 주식회사 귀중', styles_bold)],
                    [Paragraph('발      신 :', styles_bold), Paragraph(f'{company_name}  /  {manager_name}  {manager_contact}', styles_normal)],
                    [Paragraph('제      목 :', styles_bold), Paragraph('<u>로켓배송 상품 영구적 발주 중단 요청의 건</u>', styles_bold)],
                ]
                header_table = Table(header_data, colWidths=[35*mm, None])
                header_table.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP'), ('BOTTOMPADDING', (0,0), (-1,-1), 4)]))
                story.append(header_table)
                story.append(HRFlowable(width='100%', thickness=1.5, color=colors.black, spaceAfter=12))

                story.append(Paragraph('1. 귀사의 무궁한 발전을 기원합니다.', styles_normal))
                story.append(Spacer(1, 8))
                story.append(Paragraph('2. 당사는 아래와 같은 불가피한 사유(영구적 생산 및 취급 중단)로 인해 해당 상품들의 공급을 지속할 수 없게 되었습니다. 이에 따라 로켓배송 서비스의 안정적인 운영을 위해 해당 제품들의 발주 중단을 정중히 요청드리는 바입니다.', styles_normal))
                story.append(Spacer(1, 16))
                story.append(Paragraph('- 아 래 -', styles_center))
                story.append(Spacer(1, 12))

                # SKU 테이블
                sku_table_data = [
                    [Paragraph('SKU ID', styles_bold), Paragraph('SKU 명칭', styles_bold), Paragraph('발주 중단 사유', styles_bold)]
                ]
                for sku in sku_list:
                    sku_table_data.append([
                        Paragraph(sku['id'], styles_normal),
                        Paragraph(sku['name'], styles_normal),
                        Paragraph(reason_detail, styles_small)
                    ])

                sku_table = Table(sku_table_data, colWidths=[35*mm, 75*mm, 50*mm])
                sku_table.setStyle(TableStyle([
                    ('GRID', (0,0), (-1,-1), 1, colors.black),
                    ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                    ('ALIGN', (0,0), (0,-1), 'CENTER'),
                    ('ALIGN', (2,0), (2,-1), 'CENTER'),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                    ('TOPPADDING', (0,0), (-1,-1), 6),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 6),
                ]))
                story.append(sku_table)
                story.append(Spacer(1, 30))

                # 서명
                story.append(Paragraph(date_str, styles_center))
                story.append(Spacer(1, 20))
                story.append(Paragraph(company_name, styles_big))
                story.append(Spacer(1, 10))

                # 직인 + 대표이사 서명
                if stamp_data:
                    from reportlab.platypus import Image as RLImage, Flowable
                    sz = stamp_size * 0.8

                    class StampWithText(Flowable):
                        def __init__(self, stamp_bytes, stamp_sz, rep_name):
                            Flowable.__init__(self)
                            self.stamp_bytes = stamp_bytes
                            self.stamp_sz = stamp_sz
                            self.rep_name = rep_name
                            self.width = 160*mm
                            self.height = stamp_sz + 5

                        def draw(self):
                            canvas = self.canv
                            # 텍스트를 중앙에 배치
                            text = f'대표이사   {self.rep_name}   (인)'
                            canvas.setFont('NanumBold', 10)
                            text_w = canvas.stringWidth(text, 'NanumBold', 10)
                            text_x = self.width / 2 - text_w / 2
                            text_y = 5
                            canvas.drawString(text_x, text_y, text)
                            # 도장을 이름 왼쪽에 겹치도록 배치
                            from reportlab.lib.utils import ImageReader
                            stamp_reader = ImageReader(io.BytesIO(self.stamp_bytes))
                            stamp_x = text_x - self.stamp_sz * 0.15
                            stamp_y = text_y - self.stamp_sz * 0.35
                            canvas.drawImage(stamp_reader, stamp_x, stamp_y,
                                           width=self.stamp_sz, height=self.stamp_sz,
                                           preserveAspectRatio=True, mask='auto')

                    stamp_flow = StampWithText(stamp_data, sz, representative)
                    stamp_flow.hAlign = 'CENTER'
                    story.append(stamp_flow)
                else:
                    story.append(Paragraph(f'대표이사&nbsp;&nbsp;&nbsp;{representative}&nbsp;&nbsp;&nbsp;(인)', styles_center))

                doc.build(story)
                buf.seek(0)

                today = datetime.now().strftime('%Y%m%d')
                st.success('✅ 공문 PDF 생성 완료!')
                st.download_button(
                    label='⬇️ 공문 PDF 다운로드',
                    data=buf,
                    file_name=f'발주중단공문_{company_name}_{today}.pdf',
                    mime='application/pdf',
                    key='gm_dl'
                )
            except Exception as e:
                st.error(f'❌ PDF 생성 오류: {e}')
                import traceback
                st.code(traceback.format_exc())

# ══════════════════════════════════════════════════════
# 탭6: 쉽먼트 통합 관리
# ══════════════════════════════════════════════════════

# ── 쉽먼트 분석 헬퍼 함수들 ────────────────────────────

def _extract_manifest_info(pdf_bytes):
    """매니페스트 PDF 바이트에서 박스/송장/쉽먼트번호 정보 추출"""
    pdf = pdfplumber.open(io.BytesIO(pdf_bytes))
    pages_info = []
    for i, page in enumerate(pdf.pages):
        text = page.extract_text() or ''
        box_match = re.search(r'박스\s*(\d+-\d+)', text)
        invoice_match = re.search(r'송장번호\s*\n?\s*(\d{12,})', text)
        if not invoice_match:
            invoice_match = re.search(r'(4\d{11})', text)
        # 쉽먼트번호 (7~10자리, 보통 8자리) — \b로 12자리 운송장번호와 혼동 방지
        shipment_id_match = re.search(r'쉽먼트\s*번호\s*\n\s*(\d{7,10})\b', text)
        if not shipment_id_match:
            shipment_id_match = re.search(r'쉽먼트\s*번호[\s:]+(\d{7,10})\b', text)
        if not shipment_id_match:
            shipment_id_match = re.search(r'쉽먼트\s*번호.{0,50}?\b(\d{7,10})\b', text, re.DOTALL)
        pages_info.append({
            'page_idx': i,
            'box_number': box_match.group(1) if box_match else None,
            'invoice_number': invoice_match.group(1) if invoice_match else None,
            'shipment_id': shipment_id_match.group(1) if shipment_id_match else None,
            'is_main_page': box_match is not None
        })
    pdf.close()
    return pages_info


def _extract_manifest_products(pdf_bytes):
    """매니페스트 PDF에서 박스별 상품 목록(상품번호/바코드) 추출.
    상품 테이블 행은 'R<12~13자리 바코드> <6~10자리 SKU ID>' 패턴.
    반환: [{'box_number', 'invoice_number', 'shipment_id',
            'products': [{'barcode', 'sku_id'}, ...]}]
    """
    pdf = pdfplumber.open(io.BytesIO(pdf_bytes))
    boxes = []
    current = None
    for _, page in enumerate(pdf.pages):
        text = page.extract_text() or ''
        box_match = re.search(r'박스\s*(\d+-\d+)', text)
        if box_match:
            if current:
                boxes.append(current)
            invoice_match = re.search(r'송장번호\s*\n?\s*(\d{12,})', text)
            if not invoice_match:
                invoice_match = re.search(r'(4\d{11})', text)
            shipment_id_match = re.search(r'쉽먼트\s*번호\s*\n\s*(\d{7,10})\b', text)
            if not shipment_id_match:
                shipment_id_match = re.search(r'쉽먼트\s*번호[\s:]+(\d{7,10})\b', text)
            if not shipment_id_match:
                shipment_id_match = re.search(r'쉽먼트\s*번호.{0,50}?\b(\d{7,10})\b', text, re.DOTALL)
            current = {
                'box_number': box_match.group(1),
                'invoice_number': invoice_match.group(1) if invoice_match else None,
                'shipment_id': shipment_id_match.group(1) if shipment_id_match else None,
                'products': [],
            }
        if current is None:
            continue
        # 상품 테이블 행 (제품 바코드 + 상품번호)
        for m in re.finditer(r'(R\d{12,13})\s+(\d{6,10})\b', text):
            current['products'].append({
                'barcode': m.group(1),
                'sku_id': m.group(2),
            })
    if current:
        boxes.append(current)
    pdf.close()
    return boxes


def _extract_label_info(pdf_bytes):
    """라벨(동봉문서) PDF 바이트에서 박스/송장/쉽먼트번호 정보 추출"""
    pdf = pdfplumber.open(io.BytesIO(pdf_bytes))
    pages_info = []
    for i, page in enumerate(pdf.pages):
        text = page.extract_text() or ''
        box_match = re.search(r'박스\s*(\d+-\d+)', text)
        invoice_match = re.search(r'(4\d{11})', text)
        # 동봉문서 '쉽먼트번호\n45232459' 추출 (레이아웃: 라벨 다음 줄에 값)
        shipment_id_match = re.search(r'쉽먼트\s*번호\s*\n\s*(\d{7,10})\b', text)
        if not shipment_id_match:
            shipment_id_match = re.search(r'쉽먼트\s*번호[\s:]+(\d{7,10})\b', text)
        if not shipment_id_match:
            # 유연 패턴: 쉽먼트번호 ~ 50자 안에 7~10자리 독립 숫자
            shipment_id_match = re.search(r'쉽먼트\s*번호.{0,50}?\b(\d{7,10})\b', text, re.DOTALL)
        pages_info.append({
            'page_idx': i,
            'box_number': box_match.group(1) if box_match else None,
            'invoice_number': invoice_match.group(1) if invoice_match else None,
            'shipment_id': shipment_id_match.group(1) if shipment_id_match else None,
        })
    pdf.close()
    return pages_info


def _group_manifest_pages(pages_info):
    """매니페스트 박스별 그룹핑"""
    groups, current = [], None
    for info in pages_info:
        if info['is_main_page']:
            if current:
                groups.append(current)
            current = {
                'box_number': info['box_number'],
                'invoice_number': info['invoice_number'],
                'shipment_id': info.get('shipment_id'),
                'page_indices': [info['page_idx']],
            }
        else:
            if current:
                current['page_indices'].append(info['page_idx'])
                # 서브페이지에만 쉽먼트번호가 있는 경우 대비 — 첫 번째 발견 값 유지
                if not current.get('shipment_id') and info.get('shipment_id'):
                    current['shipment_id'] = info['shipment_id']
    if current:
        groups.append(current)
    return groups


def _group_label_pages(pages_info):
    """라벨 박스별 그룹핑"""
    box_groups = OrderedDict()
    for info in pages_info:
        box = info['box_number']
        if box not in box_groups:
            box_groups[box] = {
                'box_number': box,
                'invoice_number': info['invoice_number'],
                'shipment_id': info.get('shipment_id'),
                'page_indices': [],
            }
        box_groups[box]['page_indices'].append(info['page_idx'])
        if not box_groups[box]['invoice_number'] and info['invoice_number']:
            box_groups[box]['invoice_number'] = info['invoice_number']
        if not box_groups[box].get('shipment_id') and info.get('shipment_id'):
            box_groups[box]['shipment_id'] = info['shipment_id']
    return list(box_groups.values())


def _render_labels_4up(pdf_bytes, sorted_groups, dpi=300):
    """라벨 PDF → 이미지 렌더링 → 4분할 A4"""
    pdf_doc = pdfium.PdfDocument(pdf_bytes)
    all_indices = []
    for g in sorted_groups:
        all_indices.extend(g['page_indices'])

    images = []
    for idx in all_indices:
        page = pdf_doc[idx]
        bitmap = page.render(scale=dpi / 72)
        images.append(bitmap.to_pil())
    pdf_doc.close()

    # 4-up
    a4_w, a4_h = int(8.27 * dpi), int(11.69 * dpi)
    slot_w, slot_h = a4_w // 2, a4_h // 2
    result = []

    for start in range(0, len(images), 4):
        chunk = images[start:start + 4]
        canvas = Image.new('RGB', (a4_w, a4_h), 'white')
        positions = [(0, 0), (slot_w, 0), (0, slot_h), (slot_w, slot_h)]
        for i, img in enumerate(chunk):
            ratio = img.width / img.height
            s_ratio = slot_w / slot_h
            if ratio > s_ratio:
                nw, nh = slot_w, int(slot_w / ratio)
            else:
                nh, nw = slot_h, int(slot_h * ratio)
            resized = img.resize((nw, nh), Image.LANCZOS)
            x = positions[i][0] + (slot_w - nw) // 2
            y = positions[i][1] + (slot_h - nh) // 2
            canvas.paste(resized, (x, y))
        result.append(canvas)
    return result


def _parse_csv_bytes(csv_bytes):
    """CSV 바이트 → 아이템 리스트"""
    text = csv_bytes.decode('utf-8-sig', errors='replace')
    reader = csv.reader(text.splitlines())
    rows = list(reader)
    if not rows:
        return []
    items = []
    for row in rows[1:]:
        if len(row) < 11 or not (row[1] if len(row) > 1 else '').strip():
            continue
        def safe(idx, default=''):
            return row[idx].strip() if idx < len(row) else default
        try:
            qty = int(safe(7, '0') or '0')
        except ValueError:
            qty = 0
        items.append({
            'logisticsCenter': safe(1),
            'expectedDate': safe(3),
            'productBarcode': safe(5),
            'productName': safe(6),
            'quantity': qty,
            'shipmentNumber': safe(8),
            'orderDate': safe(9),
            'boxNumber': safe(10),
            'location': safe(12),
        })
    return items


def _parse_xlsx_bytes(xlsx_bytes):
    """엑셀 바이트 → 아이템 리스트 (헤더 기반 자동 매핑)"""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    # 헤더 행 찾기
    headers = None
    for row in rows_iter:
        cells = [str(c).strip() if c is not None else '' for c in row]
        if '바코드' in cells and '상품명' in cells:
            headers = cells
            break
    if not headers:
        return []

    FIELD_MAP = {
        'logisticsCenter': ['물류센터', 'FC'],
        'expectedDate':    ['입고예정일'],
        'productBarcode':  ['바코드', '상품바코드'],
        'productName':     ['상품명', '품명'],
        'quantity':        ['수량'],
        'shipmentNumber':  ['쉽먼트운송장', '송장번호', '운송장번호'],
        'orderDate':       ['발주일'],
        'boxNumber':       ['박스번호'],
        'location':        ['위치', '적재위치'],
    }

    col_map = {}
    for field, keywords in FIELD_MAP.items():
        for kw in keywords:
            for i, h in enumerate(headers):
                if h == kw or h.startswith(kw) or kw in h:
                    col_map[field] = i
                    break
            if field in col_map:
                break

    if 'productBarcode' not in col_map:
        return []

    items = []
    for row in rows_iter:
        cells = [str(c).strip() if c is not None else '' for c in row]
        def safe(field, default=''):
            idx = col_map.get(field)
            if idx is not None and idx < len(cells):
                return cells[idx]
            return default
        barcode = safe('productBarcode')
        if not barcode:
            continue
        try:
            qty = int(float(safe('quantity', '0') or '0'))
        except (ValueError, TypeError):
            qty = 0
        # 쉽먼트운송장번호가 과학표기법(4.62E+11)인 경우 정수로 변환
        shipment = safe('shipmentNumber')
        try:
            if 'E' in shipment or 'e' in shipment:
                shipment = str(int(float(shipment)))
        except (ValueError, TypeError):
            pass
        items.append({
            'logisticsCenter': safe('logisticsCenter'),
            'expectedDate': safe('expectedDate'),
            'productBarcode': barcode,
            'productName': safe('productName'),
            'quantity': qty,
            'shipmentNumber': shipment,
            'orderDate': safe('orderDate'),
            'boxNumber': safe('boxNumber'),
            'location': safe('location'),
        })
    wb.close()
    return items
    return items


with tab6:
    st.header('🚛 쉽먼트 통합 관리')
    st.caption('CSV + 매니페스트 + 라벨을 한번에 업로드하면 출고지시서 생성 + 송장번호순 정렬 + 라벨 4분할 + 전체 통합 PDF를 만듭니다')

    st.divider()

    # ── 파일 업로드 ──────────────────────────────────
    st.subheader('📂 파일 업로드')
    st.caption('CSV 1개 + 매니페스트/라벨 PDF 여러 개를 한꺼번에 드래그 & 드롭하세요')

    uploaded_files = st.file_uploader(
        '파일 선택 (CSV + PDF)',
        type=['csv', 'xlsx', 'pdf'],
        accept_multiple_files=True,
        key='shipment_files'
    )

    if uploaded_files:
        # 파일 분류
        csv_file = None
        manifest_files = {}
        label_files = {}

        for f in uploaded_files:
            fname = f.name.lower()
            if fname.endswith('.csv') or fname.endswith('.xlsx'):
                csv_file = f
            elif 'manifest' in fname:
                sid = re.search(r'\((\d+)\)', f.name)
                if sid:
                    manifest_files[sid.group(1)] = f
            elif 'label' in fname:
                sid = re.search(r'\((\d+)\)', f.name)
                if sid:
                    label_files[sid.group(1)] = f

        # 짝 매칭
        pairs = []
        for sid in sorted(manifest_files.keys()):
            if sid in label_files:
                pairs.append((sid, manifest_files[sid], label_files[sid]))

        # 미리보기 표시
        st.markdown(f'**분류 결과:**')
        if csv_file:
            st.markdown(f'- CSV: `{csv_file.name}`')
        else:
            st.warning('CSV 파일이 없습니다. 출고지시서 없이 쉽먼트만 처리합니다.')

        st.markdown(f'- 쉽먼트 세트: **{len(pairs)}개**')
        for sid, m, l in pairs:
            st.markdown(f'  - `[{sid}]` {m.name} + {l.name}')

        # 매칭 안 된 파일 경고
        for sid in manifest_files:
            if sid not in label_files:
                st.warning(f'쉽먼트 {sid}: 매니페스트만 있고 라벨 없음')
        for sid in label_files:
            if sid not in manifest_files:
                st.warning(f'쉽먼트 {sid}: 라벨만 있고 매니페스트 없음')

        if not pairs:
            st.error('매칭되는 매니페스트/라벨 세트가 없습니다.')
        else:
            st.divider()

            if st.button('🚀 통합 PDF 생성 시작', type='primary', key='shipment_btn'):
                progress = st.progress(0)
                status = st.empty()
                total_steps = len(pairs) + (1 if csv_file else 0) + 1
                step = 0

                try:
                    # ===== 1. 매니페스트 분석 → 송장→(쉽먼트ID, 박스) 매핑 =====
                    status.text('📊 매니페스트 분석 중...')
                    invoice_mapping = {}
                    manifest_data = {}  # sid → (bytes, sorted_groups)

                    inv_to_ship_id_map = {}  # 송장번호 → 쉽먼트번호
                    for sid, mf, lf in pairs:
                        m_bytes = mf.read()
                        mf.seek(0)
                        m_info = _extract_manifest_info(m_bytes)
                        m_groups = _group_manifest_pages(m_info)
                        sorted_m = sorted(m_groups, key=lambda g: g['invoice_number'] or '')
                        manifest_data[sid] = (m_bytes, sorted_m)

                        for g in sorted_m:
                            if g['invoice_number']:
                                invoice_mapping[g['invoice_number']] = (sid, g['box_number'])
                                if g.get('shipment_id'):
                                    inv_to_ship_id_map[g['invoice_number']] = g['shipment_id']

                        # 라벨(동봉문서) PDF에서도 쉽먼트번호 추출 (매니페스트에 없을 수도 있음)
                        try:
                            l_bytes = lf.read()
                            lf.seek(0)
                            l_info = _extract_label_info(l_bytes)
                            l_groups = _group_label_pages(l_info)
                            for g in l_groups:
                                if g.get('invoice_number') and g.get('shipment_id'):
                                    # 라벨에서 추출한 것을 우선 (매니페스트 값 덮어쓰기)
                                    inv_to_ship_id_map[g['invoice_number']] = g['shipment_id']
                        except Exception:
                            pass

                    # ===== 2. CSV → 출고지시서 PDF 생성 (송장번호별) =====
                    so_by_invoice = {}  # 송장번호 → PDF BytesIO
                    so_pdf_buf = None
                    so_pages = 0
                    if csv_file:
                        status.text('📄 출고지시서 생성 중...')
                        file_bytes = csv_file.read()
                        if csv_file.name.lower().endswith('.xlsx'):
                            items = _parse_xlsx_bytes(file_bytes)
                        else:
                            items = _parse_csv_bytes(file_bytes)

                        if items:
                            grouped = OrderedDict()
                            for item in items:
                                key = item.get('shipmentNumber', '')
                                grouped.setdefault(key, []).append(item)

                            for key in grouped:
                                grouped[key].sort(key=lambda x: (
                                    [int(c) if c.isdigit() else c.lower()
                                     for c in re.split(r'(\d+)', x.get('boxNumber', ''))],
                                    x.get('productName', '')
                                ))

                            # 송장별 박스번호 자동 부여 (입고예정일>물류센터>송장번호 정렬)
                            all_items_for_box = []
                            for inv_num, inv_items in grouped.items():
                                all_items_for_box.extend(inv_items)
                            ship_to_box_num = assign_box_numbers(all_items_for_box)

                            all_so_bufs = []
                            for inv_num in sorted(grouped.keys()):
                                inv_items = grouped[inv_num]
                                real_ship_id = inv_to_ship_id_map.get(inv_num, inv_num)
                                auto_box_num = ship_to_box_num.get(inv_num)
                                center = inv_items[0].get('logisticsCenter', '')
                                gk = f"{center}_{inv_num}" if center else inv_num
                                pdf_buf = create_work_order_pdf(gk, inv_items, real_ship_id, auto_box_num)
                                so_by_invoice[inv_num] = pdf_buf
                                all_so_bufs.append(pdf_buf)

                            # 출고지시서 전체 병합본
                            so_writer = PdfWriter()
                            for buf in all_so_bufs:
                                buf.seek(0)
                                reader = PdfReader(buf)
                                for page in reader.pages:
                                    so_writer.add_page(page)
                            so_pdf_buf = io.BytesIO()
                            so_writer.write(so_pdf_buf)
                            so_pdf_buf.seek(0)
                            so_pages = len(PdfReader(io.BytesIO(so_pdf_buf.getvalue())).pages)

                            st.success(f'출고지시서 {len(all_so_bufs)}건 생성 완료')

                        step += 1
                        progress.progress(step / total_steps)

                    # ===== 3. 각 세트별 처리 =====
                    label_data = {}  # sid → (l_bytes, sorted_l)
                    shipment_only_writer = PdfWriter()  # 쉽먼트만 (매니페스트+라벨)

                    for sid, mf, lf in pairs:
                        status.text(f'📦 쉽먼트 {sid} 처리 중...')

                        m_bytes, sorted_m = manifest_data[sid]
                        m_reader = PdfReader(io.BytesIO(m_bytes))

                        # 매니페스트 정렬 → 쉽먼트 전용 PDF에 추가
                        for g in sorted_m:
                            for pidx in g['page_indices']:
                                shipment_only_writer.add_page(m_reader.pages[pidx])

                        # 라벨 분석
                        l_bytes = lf.read()
                        lf.seek(0)
                        l_info = _extract_label_info(l_bytes)
                        l_groups = _group_label_pages(l_info)
                        sorted_l = sorted(l_groups, key=lambda g: g['invoice_number'] or '')
                        label_data[sid] = (l_bytes, sorted_l)

                        step += 1
                        progress.progress(step / total_steps)

                    # 라벨 4분할 → 쉽먼트 전용에 추가
                    for sid, mf, lf in pairs:
                        l_bytes, sorted_l = label_data[sid]
                        four_up = _render_labels_4up(l_bytes, sorted_l)
                        for img in four_up:
                            img_buf = io.BytesIO()
                            img.save(img_buf, format='PDF', resolution=300)
                            img_buf.seek(0)
                            lp = PdfReader(img_buf)
                            shipment_only_writer.add_page(lp.pages[0])

                    shipment_only_buf = io.BytesIO()
                    shipment_only_writer.write(shipment_only_buf)
                    shipment_only_buf.seek(0)
                    shipment_only_pages = len(PdfReader(io.BytesIO(shipment_only_buf.getvalue())).pages)

                    # ===== 4. 전체 통합 =====
                    # 순서: [출고지시서→매니페스트→라벨] 송장번호순 통합
                    status.text('📎 전체 통합 PDF 생성 중...')
                    final_writer = PdfWriter()
                    total_pages = 0
                    label_total = 0

                    # 송장번호별 라벨 그룹 매핑 (sid → {invoice → [groups]})
                    label_by_invoice = {}
                    for sid, mf, lf in pairs:
                        l_bytes, sorted_l = label_data[sid]
                        inv_map = {}
                        for g in sorted_l:
                            inv = g['invoice_number'] or ''
                            inv_map.setdefault(inv, []).append(g)
                        label_by_invoice[sid] = inv_map

                    # 송장번호순 출고지시서→매니페스트→라벨 통합 배치
                    for sid, mf, lf in pairs:
                        m_bytes, sorted_m = manifest_data[sid]
                        m_reader = PdfReader(io.BytesIO(m_bytes))
                        l_bytes, sorted_l = label_data[sid]

                        for g in sorted_m:
                            inv = g['invoice_number']
                            # 출고지시서 먼저
                            if inv and inv in so_by_invoice:
                                so_buf = so_by_invoice[inv]
                                so_buf.seek(0)
                                so_reader = PdfReader(so_buf)
                                for page in so_reader.pages:
                                    final_writer.add_page(page)
                                total_pages += len(so_reader.pages)
                            # 매니페스트
                            for pidx in g['page_indices']:
                                final_writer.add_page(m_reader.pages[pidx])
                                total_pages += 1
                            # 해당 송장의 라벨 바로 뒤에 배치
                            inv_key = inv or ''
                            if inv_key in label_by_invoice.get(sid, {}):
                                inv_label_groups = label_by_invoice[sid].pop(inv_key)
                                four_up = _render_labels_4up(l_bytes, inv_label_groups)
                                for img in four_up:
                                    img_buf = io.BytesIO()
                                    img.save(img_buf, format='PDF', resolution=300)
                                    img_buf.seek(0)
                                    lp = PdfReader(img_buf)
                                    final_writer.add_page(lp.pages[0])
                                    total_pages += 1
                                    label_total += 1

                        # 매칭 안 된 나머지 라벨 처리
                        for inv_key, groups in label_by_invoice.get(sid, {}).items():
                            four_up = _render_labels_4up(l_bytes, groups)
                            for img in four_up:
                                img_buf = io.BytesIO()
                                img.save(img_buf, format='PDF', resolution=300)
                                img_buf.seek(0)
                                lp = PdfReader(img_buf)
                                final_writer.add_page(lp.pages[0])
                                total_pages += 1
                                label_total += 1

                    final_buf = io.BytesIO()
                    final_writer.write(final_buf)
                    final_buf.seek(0)

                    step += 1
                    progress.progress(1.0)
                    status.text('✅ 완료!')

                    # 결과를 session_state에 저장 (다운로드 버튼 클릭 후에도 유지)
                    st.session_state.shipment_result = {
                        'final_bytes': final_buf.getvalue(),
                        'shipment_only_bytes': shipment_only_buf.getvalue(),
                        'so_bytes': so_pdf_buf.getvalue() if so_pdf_buf else None,
                        'total_pages': total_pages,
                        'label_total': label_total,
                        'shipment_only_pages': shipment_only_pages,
                        'so_pages': so_pages,
                        'timestamp': datetime.now().strftime('%Y%m%d_%H%M'),
                    }

                except Exception as e:
                    st.error(f'❌ 오류 발생: {e}')
                    import traceback
                    st.code(traceback.format_exc())

        # ===== 결과 표시 (session_state에서 읽어 버튼 유지) =====
        if 'shipment_result' in st.session_state:
            sres = st.session_state.shipment_result
            st.divider()
            st.subheader('📋 처리 결과')
            st.markdown(f"""
| 구분 | 페이지 |
|------|--------|
| 출고지시서 + 매니페스트 (교차) | {sres['total_pages'] - sres['label_total']}p |
| 라벨 4분할 | {sres['label_total']}p |
| **전체 합계** | **{sres['total_pages']}p** |
""")
            st.caption('순서: [출고지시서→매니페스트→라벨] 송장번호순 통합 배치')

            st.divider()

            col_a, col_b, col_c = st.columns(3)
            with col_a:
                st.download_button(
                    label=f"⬇️ 전체 통합 PDF ({sres['total_pages']}p)",
                    data=sres['final_bytes'],
                    file_name=f"shipment_ALL_merged_{sres['timestamp']}.pdf",
                    mime='application/pdf',
                    key='ship_dl_all',
                    type='primary',
                    use_container_width=True,
                )

            with col_b:
                st.download_button(
                    label=f"⬇️ 쉽먼트만 ({sres['shipment_only_pages']}p)",
                    data=sres['shipment_only_bytes'],
                    file_name=f"shipment_only_{sres['timestamp']}.pdf",
                    mime='application/pdf',
                    key='ship_dl_shipment',
                    use_container_width=True,
                )

            with col_c:
                if sres['so_bytes']:
                    st.download_button(
                        label=f"⬇️ 출고지시서만 ({sres['so_pages']}p)",
                        data=sres['so_bytes'],
                        file_name=f"출고지시서_{sres['timestamp']}.pdf",
                        mime='application/pdf',
                        key='ship_dl_so',
                        use_container_width=True,
                    )

# ── 쉽먼트 재출력 탭 ──────────────────────────────────────
def _pick_df_to_items(df):
    """피킹&분류 탭의 df_출고 → create_work_order_pdf 용 items 리스트 변환"""
    items = []
    for _, row in df.iterrows():
        try:
            qty = int(row.get('수량', 0) or 0)
        except (ValueError, TypeError):
            qty = 0
        items.append({
            'logisticsCenter': str(row.get('물류센터(FC)', '') or '').strip(),
            'expectedDate': str(row.get('입고예정일(EDD)', '') or row.get('입고예정일', '') or '').strip(),
            'productBarcode': str(row.get('바코드', '') or '').strip(),
            'productName': str(row.get('상품명', '') or '').strip(),
            'quantity': qty,
            'shipmentNumber': str(row.get('쉽먼트운송장번호', '') or '').strip(),
            'orderDate': str(row.get('발주일', '') or '').strip(),
            'boxNumber': str(row.get('박스번호', '') or '').strip(),
            'location': str(row.get('위치', '') or row.get('적재위치', '') or '').strip(),
        })
    return items


def _run_reprint_pipeline(rp_items, rp_manifest_files, rp_label_files,
                          existing_box_map=None, write_new_to_sheet=False,
                          sheet_client=None, sheet_url=None, sheet_tab=None,
                          on_progress=None, on_status=None):
    """재출력 공통 처리: 매니페스트/라벨 PDF 파싱 → 송장 매칭 → 박스번호 부여
    → 출고지시서 PDF 생성 → 통합 PDF 병합.

    existing_box_map: 기존 {송장: 박스번호}. None이면 fresh 부여(assign_box_numbers).
        dict면 assign_box_numbers_with_existing로 기존값 보존.
    write_new_to_sheet: True면 신규 부여분을 시트 M열에 기록 (기존값 보존).
    on_progress: 0~1 float 콜백 (진행률 UI용)
    on_status: str 콜백 (상태 메시지용)

    반환: 성공시 dict, 실패시 {'error': str}
    """
    def _p(v):
        if on_progress:
            try:
                on_progress(v)
            except Exception:
                pass

    def _s(msg):
        if on_status:
            try:
                on_status(msg)
            except Exception:
                pass

    _s('📄 데이터 분석 중...')
    if not rp_items:
        return {'error': '데이터에서 항목을 찾을 수 없습니다.'}

    csv_invoices = set()
    rp_grouped = OrderedDict()
    for item in rp_items:
        inv = item.get('shipmentNumber', '')
        if inv:
            csv_invoices.add(inv)
            rp_grouped.setdefault(inv, []).append(item)
    for key in rp_grouped:
        rp_grouped[key].sort(key=lambda x: (
            [int(c) if c.isdigit() else c.lower()
             for c in re.split(r'(\d+)', x.get('boxNumber', ''))],
            x.get('productName', '')
        ))
    _p(0.2)

    _s('📊 매니페스트/라벨 분석 중...')
    rp_manifest_data = []
    for fname, mf in rp_manifest_files:
        m_bytes = mf.read(); mf.seek(0)
        m_info = _extract_manifest_info(m_bytes)
        m_groups = _group_manifest_pages(m_info)
        sorted_m = sorted(m_groups, key=lambda g: g['invoice_number'] or '')
        rp_manifest_data.append((m_bytes, sorted_m))
    rp_label_data = []
    for fname, lf in rp_label_files:
        l_bytes = lf.read(); lf.seek(0)
        l_info = _extract_label_info(l_bytes)
        l_groups = _group_label_pages(l_info)
        sorted_l = sorted(l_groups, key=lambda g: g['invoice_number'] or '')
        rp_label_data.append((l_bytes, sorted_l))
    _p(0.4)

    all_manifest_invoices = set()
    invoice_to_shipment_id = {}
    for _, sorted_m in rp_manifest_data:
        for g in sorted_m:
            if g['invoice_number']:
                all_manifest_invoices.add(g['invoice_number'])
                if g.get('shipment_id'):
                    invoice_to_shipment_id[g['invoice_number']] = g['shipment_id']
    all_label_invoices = set()
    for _, sorted_l in rp_label_data:
        for g in sorted_l:
            if g['invoice_number']:
                all_label_invoices.add(g['invoice_number'])
                if g.get('shipment_id'):
                    invoice_to_shipment_id[g['invoice_number']] = g['shipment_id']

    all_pdf_invoices = all_manifest_invoices | all_label_invoices
    matched = csv_invoices & all_pdf_invoices
    not_in_manifest = csv_invoices - all_pdf_invoices
    not_in_csv = all_pdf_invoices - csv_invoices

    if not matched:
        return {'error': '데이터와 매니페스트/라벨 간 매칭되는 송장번호가 없습니다.'}

    _s('📄 출고지시서 생성 중...')
    rp_all_items = []
    for inv_num in sorted(matched):
        rp_all_items.extend(rp_grouped[inv_num])
    if existing_box_map is not None:
        rp_ship_to_box_num = assign_box_numbers_with_existing(rp_all_items, existing_box_map)
    else:
        rp_ship_to_box_num = assign_box_numbers(rp_all_items)

    # 신규 부여된 박스번호를 시트 M열에 기록
    new_box_only = {}
    sheet_write_result = None  # None: 시도 안 함 / >=0: 기록된 셀 수 / -1: API 실패
    if existing_box_map is not None:
        new_box_only = {s: n for s, n in rp_ship_to_box_num.items()
                        if s not in existing_box_map}
        if write_new_to_sheet and new_box_only and sheet_client and sheet_url and sheet_tab:
            sheet_write_result = pick_write_box_numbers(
                sheet_client, sheet_url, sheet_tab,
                new_box_only, only_empty=True,
            )

    rp_so_by_invoice = {}
    for inv_num in sorted(matched):
        inv_items = rp_grouped[inv_num]
        auto_box_num = rp_ship_to_box_num.get(inv_num)
        center = inv_items[0].get('logisticsCenter', '')
        gk = f"{center}_{inv_num}" if center else inv_num
        real_shipment_id = invoice_to_shipment_id.get(inv_num, inv_num)
        pdf_buf = create_work_order_pdf(gk, inv_items, real_shipment_id, auto_box_num)
        rp_so_by_invoice[inv_num] = pdf_buf
    _p(0.6)

    _s('📎 통합 PDF 생성 중...')
    rp_final_writer = PdfWriter()
    rp_shipment_only_writer = PdfWriter()
    rp_so_only_writer = PdfWriter()
    rp_total = rp_label_total = rp_shipment_total = rp_so_total = 0

    manifest_pages_by_inv = {}
    for m_bytes, sorted_m in rp_manifest_data:
        for g in sorted_m:
            inv = g['invoice_number']
            if not inv or inv not in matched:
                continue
            manifest_pages_by_inv.setdefault(inv, []).append((m_bytes, g['page_indices']))

    label_groups_by_inv = {}
    for l_bytes, sorted_l in rp_label_data:
        for g in sorted_l:
            inv = g['invoice_number']
            if not inv or inv not in matched:
                continue
            label_groups_by_inv.setdefault(inv, []).append((l_bytes, g))

    for inv_num in sorted(matched):
        if inv_num in rp_so_by_invoice:
            so_buf = rp_so_by_invoice[inv_num]
            so_buf.seek(0)
            so_reader = PdfReader(so_buf)
            for page in so_reader.pages:
                rp_final_writer.add_page(page)
                rp_so_only_writer.add_page(page)
            rp_total += len(so_reader.pages)
            rp_so_total += len(so_reader.pages)
        for m_bytes, page_indices in manifest_pages_by_inv.get(inv_num, []):
            m_reader = PdfReader(io.BytesIO(m_bytes))
            for pidx in page_indices:
                rp_final_writer.add_page(m_reader.pages[pidx])
                rp_shipment_only_writer.add_page(m_reader.pages[pidx])
                rp_total += 1
                rp_shipment_total += 1
        inv_label_list = label_groups_by_inv.get(inv_num, [])
        if inv_label_list:
            grouped_by_lbytes = {}
            for l_bytes, grp in inv_label_list:
                grouped_by_lbytes.setdefault(id(l_bytes), [l_bytes, []])[1].append(grp)
            for _, (l_bytes, groups) in grouped_by_lbytes.items():
                four_up = _render_labels_4up(l_bytes, groups)
                for img in four_up:
                    img_buf = io.BytesIO()
                    img.save(img_buf, format='PDF', resolution=300)
                    img_buf.seek(0)
                    lp = PdfReader(img_buf)
                    rp_final_writer.add_page(lp.pages[0])
                    rp_shipment_only_writer.add_page(lp.pages[0])
                    rp_total += 1
                    rp_label_total += 1
                    rp_shipment_total += 1

    _p(0.9)
    rp_final_buf = io.BytesIO()
    rp_final_writer.write(rp_final_buf)
    rp_final_buf.seek(0)
    rp_shipment_only_buf = io.BytesIO()
    rp_shipment_only_writer.write(rp_shipment_only_buf)
    rp_shipment_only_buf.seek(0)
    rp_so_only_buf = io.BytesIO()
    rp_so_only_writer.write(rp_so_only_buf)
    rp_so_only_buf.seek(0)
    _p(1.0)
    _s('✅ 완료!')

    return {
        'final_bytes': rp_final_buf.getvalue(),
        'shipment_only_bytes': rp_shipment_only_buf.getvalue(),
        'so_only_bytes': rp_so_only_buf.getvalue(),
        'total': rp_total,
        'shipment_total': rp_shipment_total,
        'so_total': rp_so_total,
        'matched': len(matched),
        'not_in_manifest': sorted(not_in_manifest),
        'not_in_csv': sorted(not_in_csv),
        'ship_to_box_num': rp_ship_to_box_num,
        'new_box_only': new_box_only,
        'sheet_write_result': sheet_write_result,
        'timestamp': datetime.now().strftime('%Y%m%d_%H%M'),
    }


with tab7:
    st.header('🔄 쉽먼트 재출력')
    st.caption('피킹&분류에 로드된 시트 데이터(또는 CSV)의 송장번호와 매니페스트/라벨을 매칭하여 재출력합니다')

    # ── 데이터 소스 모드 선택 ──
    _has_sheet_data = st.session_state.get('pick_df_출고') is not None
    _use_sheet_source = False
    if _has_sheet_data:
        _use_sheet_source = st.toggle(
            '📋 피킹&분류 시트 데이터 사용 (박스번호는 시트 M열에서 보존)',
            value=True,
            key='reprint_use_sheet',
            help='시트에 이미 박스번호가 저장되어 있어서 CSV 업로드 없이도 쉽먼트/라벨만 올리면 됩니다. 발주 취소돼도 박스번호 안 틀어짐.'
        )

    st.divider()

    st.subheader('📂 파일 업로드')
    if _use_sheet_source:
        st.caption('매니페스트/라벨 PDF만 업로드하세요 (CSV 불필요 — 시트 데이터 사용)')
    else:
        st.caption('CSV 1개(필수) + 매니페스트/라벨 PDF를 업로드하세요')

    reprint_files = st.file_uploader(
        '파일 선택 (CSV + PDF)' if not _use_sheet_source else '파일 선택 (PDF)',
        type=(['pdf'] if _use_sheet_source else ['csv', 'xlsx', 'pdf']),
        accept_multiple_files=True,
        key='reprint_files'
    )

    if reprint_files:
        rp_csv = None
        rp_manifest_files = []  # [(파일명, 파일)]
        rp_label_files = []     # [(파일명, 파일)]
        rp_unknown_files = []   # 이름 패턴 불일치로 무시된 파일

        for f in reprint_files:
            fname = f.name.lower()
            if fname.endswith('.csv') or fname.endswith('.xlsx'):
                rp_csv = f
            elif 'manifest' in fname:
                rp_manifest_files.append((f.name, f))
            elif 'label' in fname:
                rp_label_files.append((f.name, f))
            else:
                rp_unknown_files.append(f.name)

        if rp_unknown_files:
            st.warning(
                f'⚠️ 파일명에 `manifest`/`label`이 없어 무시된 파일 {len(rp_unknown_files)}개: '
                + ', '.join(f'`{n}`' for n in rp_unknown_files)
            )

        st.markdown(f'**분류 결과:**')
        if _use_sheet_source:
            st.markdown(f'- 📋 시트 데이터: `피킹&분류 로드됨 ({len(st.session_state.pick_df_출고)}행)`')
        elif rp_csv:
            st.markdown(f'- CSV: `{rp_csv.name}`')
        else:
            st.error('⚠️ CSV 파일은 필수입니다. 송장번호 매칭에 사용됩니다.')

        st.markdown(f'- 매니페스트 PDF: **{len(rp_manifest_files)}개**')
        st.markdown(f'- 라벨 PDF: **{len(rp_label_files)}개**')

        if not _use_sheet_source and not rp_csv:
            st.stop()

        if not rp_manifest_files and not rp_label_files:
            st.error('매니페스트 또는 라벨 PDF가 필요합니다.')
        else:
            st.divider()

            if st.button('🔄 쉽먼트 재출력 시작', type='primary', key='reprint_btn'):
                rp_progress = st.progress(0)
                rp_status = st.empty()

                try:
                    # 데이터 소스 준비
                    if _use_sheet_source:
                        rp_items = _pick_df_to_items(st.session_state.pick_df_출고)
                    else:
                        file_bytes = rp_csv.read()
                        if rp_csv.name.lower().endswith('.xlsx'):
                            rp_items = _parse_xlsx_bytes(file_bytes)
                        else:
                            rp_items = _parse_csv_bytes(file_bytes)

                    # 시트 모드면 M열(기존값) 읽어 보존. 읽기 실패 시 경고 후 중단.
                    _existing_map = None
                    if _use_sheet_source:
                        _existing_map = dict(st.session_state.get('pick_ship_to_box') or {})
                        if not _existing_map and st.session_state.get('pick_gsheet_client'):
                            _read = pick_read_box_numbers(
                                st.session_state.pick_gsheet_client,
                                st.session_state.get('pick_sheet_url_출고', ''),
                                st.session_state.get('pick_sheet_tab_출고', ''),
                            )
                            if _read is None:
                                st.error('❌ 시트 M열 읽기 실패. 새로고침 후 다시 시도하세요.')
                                st.stop()
                            _existing_map = _read

                    result = _run_reprint_pipeline(
                        rp_items, rp_manifest_files, rp_label_files,
                        existing_box_map=_existing_map,
                        write_new_to_sheet=_use_sheet_source,
                        sheet_client=st.session_state.get('pick_gsheet_client'),
                        sheet_url=st.session_state.get('pick_sheet_url_출고', ''),
                        sheet_tab=st.session_state.get('pick_sheet_tab_출고', ''),
                        on_progress=rp_progress.progress,
                        on_status=rp_status.text,
                    )

                    if 'error' in result:
                        st.error(f'❌ {result["error"]}')
                        st.stop()

                    # 시트 세션 캐시 업데이트 — 매칭 안 된 기존 송장 번호를 지우지 않도록 MERGE
                    if _use_sheet_source and result.get('ship_to_box_num'):
                        _new_map = result['ship_to_box_num']
                        _merged_ship = dict(st.session_state.get('pick_ship_to_box') or {})
                        _merged_ship.update(_new_map)
                        st.session_state['pick_ship_to_box'] = _merged_ship
                        _sync_url = st.session_state.get('pick_sheet_url_출고', '')
                        _sync_tab = st.session_state.get('pick_sheet_tab_출고', '')
                        if _sync_url and _sync_tab:
                            _cache_key_sync = f"_pick_existing_box_{_sync_url}_{_sync_tab}"
                            _merged_cache = dict(st.session_state.get(_cache_key_sync) or {})
                            _merged_cache.update(_new_map)
                            st.session_state[_cache_key_sync] = _merged_cache
                    if result.get('sheet_write_result') == -1:
                        st.warning('⚠️ 시트 M열 쓰기 실패 — 박스번호가 시트에 저장되지 않았습니다. 다시 시도하세요.')

                    col_m1, col_m2, col_m3 = st.columns(3)
                    with col_m1:
                        st.metric('매칭됨', f'{result["matched"]}건')
                    with col_m2:
                        st.metric('데이터에만 존재', f'{len(result["not_in_manifest"])}건')
                    with col_m3:
                        st.metric('쉽먼트에만 존재', f'{len(result["not_in_csv"])}건')

                    if result['not_in_manifest']:
                        with st.expander(f'데이터에만 있는 송장 ({len(result["not_in_manifest"])}건) - 매니페스트 없음'):
                            st.code('\n'.join(result['not_in_manifest']))
                    if result['not_in_csv']:
                        with st.expander(f'쉽먼트에만 있는 송장 ({len(result["not_in_csv"])}건) - 이번에 미출력'):
                            st.code('\n'.join(result['not_in_csv']))

                    st.session_state.reprint_result = {
                        'final_bytes': result['final_bytes'],
                        'shipment_only_bytes': result['shipment_only_bytes'],
                        'so_only_bytes': result['so_only_bytes'],
                        'total': result['total'],
                        'shipment_total': result['shipment_total'],
                        'so_total': result['so_total'],
                        'matched': result['matched'],
                        'timestamp': result['timestamp'],
                    }

                except Exception as e:
                    st.error(f'❌ 오류 발생: {e}')
                    import traceback
                    st.code(traceback.format_exc())

        # ===== 결과 표시 (session_state에서 읽어 버튼 유지) =====
        if 'reprint_result' in st.session_state:
            res = st.session_state.reprint_result
            st.divider()
            st.subheader('📋 재출력 결과')
            st.markdown(f"""
| 구분 | 수량 |
|------|------|
| 매칭된 송장 | {res['matched']}건 |
| 출고지시서 | {res['so_total']}p |
| 쉽먼트 (매니페스트+라벨) | {res['shipment_total']}p |
| **전체 합계** | **{res['total']}p** |
""")
            st.caption('순서: [출고지시서→매니페스트→라벨] 매칭 송장번호순 통합 배치')

            st.divider()

            rp_col_a, rp_col_b, rp_col_c = st.columns(3)
            with rp_col_a:
                st.download_button(
                    label=f"⬇️ 전체 통합 PDF ({res['total']}p)",
                    data=res['final_bytes'],
                    file_name=f"shipment_reprint_ALL_{res['timestamp']}.pdf",
                    mime='application/pdf',
                    key='reprint_dl',
                    type='primary',
                    use_container_width=True,
                )
            with rp_col_b:
                st.download_button(
                    label=f"⬇️ 쉽먼트만 ({res['shipment_total']}p)",
                    data=res['shipment_only_bytes'],
                    file_name=f"shipment_reprint_shipment_{res['timestamp']}.pdf",
                    mime='application/pdf',
                    key='reprint_dl_shipment',
                    use_container_width=True,
                )
            with rp_col_c:
                st.download_button(
                    label=f"⬇️ 출고지시서만 ({res['so_total']}p)",
                    data=res['so_only_bytes'],
                    file_name=f"shipment_reprint_so_{res['timestamp']}.pdf",
                    mime='application/pdf',
                    key='reprint_dl_so',
                    use_container_width=True,
                )

# ══════════════════════════════════════════════════════
# 탭8: 피킹 검증 시스템
# ══════════════════════════════════════════════════════
with tab8:
    import pandas as _pd

    st.header('📦 피킹 & 분류')
    st.caption('하나의 시트로 피킹검증 또는 입고분류를 모드 전환하며 사용')

    # ── 데이터 소스 선택 ──
    pick_mode = st.radio(
        "데이터 소스",
        ["📊 구글 시트 (실시간)", "📂 CSV 파일 업로드"],
        index=0, key="pick_mode", horizontal=True,
    )

    if pick_mode == "📊 구글 시트 (실시간)":
        # 기본 고정 시트 (운영용)
        _DEFAULT_SHEET_URL = "https://docs.google.com/spreadsheets/d/1M-r5BfuVRh2dunBsR_6lZZ7f4sH7NSI0B9rbBMuMdTc/edit?gid=224790693#gid=224790693"
        _DEFAULT_DAPAE_URL = "https://docs.google.com/spreadsheets/d/1M-r5BfuVRh2dunBsR_6lZZ7f4sH7NSI0B9rbBMuMdTc/edit?gid=980080486#gid=980080486"
        # 페이지 새로고침 시에도 유지되도록 query_params에서 복원
        _qp = st.query_params
        if 'pick_url_출고' not in st.session_state:
            st.session_state['pick_url_출고'] = _qp.get('pu') or _DEFAULT_SHEET_URL
        if 'pick_tab_출고' not in st.session_state:
            st.session_state['pick_tab_출고'] = _qp.get('pt') or '출고확인'
        if 'pick_url_배대지' not in st.session_state:
            st.session_state['pick_url_배대지'] = _qp.get('bu') or _DEFAULT_DAPAE_URL
        if 'pick_tab_배대지' not in st.session_state:
            st.session_state['pick_tab_배대지'] = _qp.get('bt') or '배대지입고리스트'

        st.markdown("##### 쉽먼트 시트 (출고지시서)")
        gs_col1, gs_col2 = st.columns([3, 1])
        with gs_col1:
            url_출고 = st.text_input("구글 시트 URL", placeholder="https://docs.google.com/spreadsheets/d/...", key="pick_url_출고")
        with gs_col2:
            tab_출고 = st.text_input("탭 이름", value="출고확인", key="pick_tab_출고")

        st.markdown("##### 배대지 입고 시트 (선택)")
        gs_col3, gs_col4 = st.columns([3, 1])
        with gs_col3:
            url_배대지 = st.text_input("구글 시트 URL", placeholder="비워두면 같은 시트 사용", key="pick_url_배대지")
        with gs_col4:
            tab_배대지 = st.text_input("탭 이름", value="배대지입고리스트", key="pick_tab_배대지")

        # 입력값을 query_params에 저장 (새로고침 후 유지)
        if url_출고:
            st.query_params['pu'] = url_출고
            st.query_params['pt'] = tab_출고
        if url_배대지:
            st.query_params['bu'] = url_배대지
            st.query_params['bt'] = tab_배대지

        # 배대지 URL 비어있으면 출고지시서 URL 사용
        if not url_배대지.strip() and url_출고.strip():
            url_배대지 = url_출고

        if st.button("🔄 구글 시트 연결", use_container_width=True, key="pick_gsheet_btn"):
            if not url_출고.strip():
                st.error("쉽먼트 시트 URL을 입력해주세요")
            else:
                with st.spinner("구글 시트 연결 중..."):
                    success = pick_load_all_data(url_출고, tab_출고, url_배대지, tab_배대지)
                    if success:
                        pick_init_inventory()
                        st.success("✅ 구글 시트 연결 완료!")
                        st.rerun()
                    else:
                        st.error("연결 실패 — URL과 탭 이름을 확인하세요")
    else:
        pc1, pc2 = st.columns(2)
        with pc1:
            pick_csv_출고 = st.file_uploader("출고지시서 CSV", type=["csv"], key="pick_csv_출고")
        with pc2:
            pick_csv_배대지 = st.file_uploader("배대지 입고 CSV (선택)", type=["csv"], key="pick_csv_배대지")
        if pick_csv_출고:
            df = _pd.read_csv(pick_csv_출고, encoding="utf-8-sig")
            st.session_state.pick_df_출고 = pick_clean_출고(df)
            if st.session_state.pick_df_출고 is not None:
                st.session_state.pick_data_loaded = True
        if pick_csv_배대지:
            df = _pd.read_csv(pick_csv_배대지, encoding="utf-8-sig")
            st.session_state.pick_df_배대지 = pick_clean_배대지(df)
            if st.session_state.pick_data_loaded and st.session_state.pick_df_배대지 is not None:
                pick_init_inventory()

    # ── 데이터 상태 표시 ──
    if st.session_state.pick_df_출고 is not None:
        n_rows = len(st.session_state.pick_df_출고)
        n_ship = st.session_state.pick_df_출고["쉽먼트운송장번호"].nunique()
        st.success(f"출고지시서: {n_rows}행 / {n_ship}개 쉽먼트")
    if st.session_state.pick_df_배대지 is not None:
        st.success(f"배대지 입고: {len(st.session_state.pick_df_배대지)}행 로드됨")

    # ── 데이터 없으면 가이드 ──
    if not st.session_state.pick_data_loaded:
        st.info("위에서 데이터를 연결하세요 (구글 시트 또는 CSV)")
        st.markdown("""
**구글 시트 모드:**
1. 구글 시트 URL을 위 입력칸에 붙여넣기
2. 탭 이름을 정확히 입력 (예: 출고확인, 배대지입고리스트)
3. '구글 시트 연결' 클릭

**CSV 모드:**
1. 'CSV 파일 업로드' 선택
2. 출고지시서 CSV 업로드 (필수)
3. 배대지 입고 CSV 업로드 (선택)
        """)
        st.stop()

    # ── 출고지시서 재출력 (피킹 시작 전에 박스번호 부여 + M열 기록) ──
    st.divider()
    with st.expander("📄 출고지시서 재출력 (쉽먼트/라벨 PDF 업로드 → 박스번호 부여)", expanded=False):
        st.caption("피킹 시작 전에 쉽먼트/라벨 PDF를 업로드하면, 현재 시트 송장과 매칭해 출고지시서 PDF를 만들고 박스번호를 시트 M열에 저장합니다. 기존에 M열에 값이 있으면 그대로 유지(발주 취소 내성). ⚠️ 동일 시트를 여러 사용자가 동시에 재출력하지 마세요 — 박스번호 충돌 가능.")

        pick_reprint_files = st.file_uploader(
            '매니페스트/라벨 PDF (파일명에 manifest/label 포함)',
            type=['pdf'],
            accept_multiple_files=True,
            key='pick_reprint_files'
        )

        if pick_reprint_files:
            pk_manifest_files = []
            pk_label_files = []
            pk_unknown_files = []
            for f in pick_reprint_files:
                fname = f.name.lower()
                if 'manifest' in fname:
                    pk_manifest_files.append((f.name, f))
                elif 'label' in fname:
                    pk_label_files.append((f.name, f))
                else:
                    pk_unknown_files.append(f.name)

            if pk_unknown_files:
                st.warning(
                    f'⚠️ 파일명에 `manifest`/`label`이 없어 무시된 파일 {len(pk_unknown_files)}개: '
                    + ', '.join(f'`{n}`' for n in pk_unknown_files)
                )

            st.markdown(f'- 매니페스트 PDF: **{len(pk_manifest_files)}개** / 라벨 PDF: **{len(pk_label_files)}개**')

            if not pk_manifest_files and not pk_label_files:
                st.warning('매니페스트 또는 라벨 PDF가 필요합니다. 파일명에 `manifest` 또는 `label`을 포함시켜 주세요.')
            elif st.button('🔄 출고지시서 재출력 시작', type='primary', key='pick_reprint_btn'):
                _pk_progress = st.progress(0)
                _pk_status = st.empty()
                try:
                    _pk_items = _pick_df_to_items(st.session_state.pick_df_출고)
                    _pk_existing = dict(st.session_state.get('pick_ship_to_box') or {})
                    if (not _pk_existing
                            and st.session_state.get('pick_gsheet_client')
                            and st.session_state.get('pick_sheet_url_출고')):
                        _pk_read = pick_read_box_numbers(
                            st.session_state.pick_gsheet_client,
                            st.session_state.pick_sheet_url_출고,
                            st.session_state.pick_sheet_tab_출고,
                        )
                        if _pk_read is None:
                            st.error('❌ 시트 M열 읽기 실패. 새로고침 후 다시 시도하세요.')
                            st.stop()
                        _pk_existing = _pk_read

                    _pk_result = _run_reprint_pipeline(
                        _pk_items, pk_manifest_files, pk_label_files,
                        existing_box_map=_pk_existing,
                        write_new_to_sheet=bool(st.session_state.get('pick_use_gsheet')),
                        sheet_client=st.session_state.get('pick_gsheet_client'),
                        sheet_url=st.session_state.get('pick_sheet_url_출고', ''),
                        sheet_tab=st.session_state.get('pick_sheet_tab_출고', ''),
                        on_progress=_pk_progress.progress,
                        on_status=_pk_status.text,
                    )

                    if 'error' in _pk_result:
                        st.error(f'❌ {_pk_result["error"]}')
                    else:
                        if _pk_result.get('ship_to_box_num'):
                            # 재출력 결과는 매칭된 송장만 포함하므로 기존 값에 MERGE (덮어쓰기 X)
                            _new_map = _pk_result['ship_to_box_num']
                            _merged_ship = dict(st.session_state.get('pick_ship_to_box') or {})
                            _merged_ship.update(_new_map)
                            st.session_state['pick_ship_to_box'] = _merged_ship
                            # 입고분류 모드 캐시도 merge 방식으로 동기화
                            _sync_url = st.session_state.get('pick_sheet_url_출고', '')
                            _sync_tab = st.session_state.get('pick_sheet_tab_출고', '')
                            if _sync_url and _sync_tab:
                                _cache_key_sync = f"_pick_existing_box_{_sync_url}_{_sync_tab}"
                                _merged_cache = dict(st.session_state.get(_cache_key_sync) or {})
                                _merged_cache.update(_new_map)
                                st.session_state[_cache_key_sync] = _merged_cache
                        if _pk_result.get('sheet_write_result') == -1:
                            st.warning('⚠️ 시트 M열 쓰기 실패 — 박스번호가 시트에 저장되지 않았습니다.')
                        _n_new = len(_pk_result.get('new_box_only') or {})
                        _n_reuse = _pk_result['matched'] - _n_new
                        st.success(
                            f'✅ 재출력 완료 — 매칭 {_pk_result["matched"]}건 '
                            f'(신규 박스번호 {_n_new}건 기록 / 기존 {_n_reuse}건 재사용)'
                        )

                        _c1, _c2, _c3 = st.columns(3)
                        with _c1:
                            st.metric('매칭됨', f'{_pk_result["matched"]}건')
                        with _c2:
                            st.metric('데이터에만', f'{len(_pk_result["not_in_manifest"])}건')
                        with _c3:
                            st.metric('쉽먼트에만', f'{len(_pk_result["not_in_csv"])}건')

                        if _pk_result['not_in_manifest']:
                            with st.expander(f'데이터에만 있는 송장 ({len(_pk_result["not_in_manifest"])}건)'):
                                st.code('\n'.join(_pk_result['not_in_manifest']))
                        if _pk_result['not_in_csv']:
                            with st.expander(f'쉽먼트에만 있는 송장 ({len(_pk_result["not_in_csv"])}건)'):
                                st.code('\n'.join(_pk_result['not_in_csv']))

                        st.session_state['pick_reprint_result'] = {
                            'final_bytes': _pk_result['final_bytes'],
                            'shipment_only_bytes': _pk_result['shipment_only_bytes'],
                            'so_only_bytes': _pk_result['so_only_bytes'],
                            'total': _pk_result['total'],
                            'shipment_total': _pk_result['shipment_total'],
                            'so_total': _pk_result['so_total'],
                            'matched': _pk_result['matched'],
                            'timestamp': _pk_result['timestamp'],
                        }
                except Exception as e:
                    st.error(f'❌ 오류: {e}')
                    import traceback
                    st.code(traceback.format_exc())

        # 결과 다운로드 버튼 (rerun 후에도 유지)
        if 'pick_reprint_result' in st.session_state:
            _pres = st.session_state['pick_reprint_result']
            st.divider()
            st.markdown(
                f"**결과**: 출고지시서 {_pres['so_total']}p / "
                f"쉽먼트 {_pres['shipment_total']}p / 전체 {_pres['total']}p"
            )
            _dc1, _dc2, _dc3 = st.columns(3)
            with _dc1:
                st.download_button(
                    f"⬇️ 전체 통합 PDF ({_pres['total']}p)",
                    data=_pres['final_bytes'],
                    file_name=f"shipment_reprint_ALL_{_pres['timestamp']}.pdf",
                    mime='application/pdf',
                    key='pick_reprint_dl_all',
                    type='primary',
                    use_container_width=True,
                )
            with _dc2:
                st.download_button(
                    f"⬇️ 쉽먼트만 ({_pres['shipment_total']}p)",
                    data=_pres['shipment_only_bytes'],
                    file_name=f"shipment_reprint_shipment_{_pres['timestamp']}.pdf",
                    mime='application/pdf',
                    key='pick_reprint_dl_ship',
                    use_container_width=True,
                )
            with _dc3:
                st.download_button(
                    f"⬇️ 출고지시서만 ({_pres['so_total']}p)",
                    data=_pres['so_only_bytes'],
                    file_name=f"출고지시서_{_pres['timestamp']}.pdf",
                    mime='application/pdf',
                    key='pick_reprint_dl_so',
                    use_container_width=True,
                )

    # ── 시트 송장-상품 ↔ 동봉문서(매니페스트) 일치 검증 ──
    # 출고확인 시트는 사용자가 직접 송장을 분류해서 만든 것이라 송장에 엉뚱한 상품이
    # 들어갈 수 있음. 동봉문서(=쉽먼트의 정답)와 대조해서 잘못 분류된 상품을 찾음.
    st.divider()
    with st.expander("🔍 시트 송장-상품 ↔ 동봉문서 일치 검증", expanded=False):
        st.caption(
            "출고확인 시트의 송장-상품 매핑이 실제 매니페스트(동봉문서)와 일치하는지 검증합니다. "
            "**시트 상품 중 매니페스트에 없는 것은 🚨 오분류(엉뚱한 송장에 배정)**. "
            "매니페스트에 있는데 시트에 없는 것은 아직 도착 안 한 상품일 수 있어 참고용."
        )

        _df_pick = st.session_state.pick_df_출고
        _sku_col = None
        for _c in _df_pick.columns:
            _cn = str(_c).strip().upper().replace(' ', '')
            if _cn in ('SKUID', 'SKU_ID', 'SKU'):
                _sku_col = _c
                break
        if _sku_col is None:
            st.warning('⚠️ 시트에 `SKU ID` 컬럼(E열)이 없습니다. 시트 헤더를 확인하세요.')
        else:
            verify_files = st.file_uploader(
                '매니페스트 PDF (여러 개 가능)',
                type=['pdf'],
                accept_multiple_files=True,
                key='verify_manifest_files',
            )

            if verify_files and st.button('🔍 검증 시작', type='primary', key='verify_btn'):
                with st.spinner(f'매니페스트 {len(verify_files)}개 파싱 중...'):
                    _all_boxes = []
                    _parse_errors = []
                    for f in verify_files:
                        try:
                            _b = f.read(); f.seek(0)
                            _all_boxes.extend(_extract_manifest_products(_b))
                        except Exception as e:
                            _parse_errors.append((f.name, str(e)))

                if _parse_errors:
                    for _n, _e in _parse_errors:
                        st.error(f'❌ `{_n}` 파싱 실패: {_e}')

                # 송장별 SKU 집계 (매니페스트 = 정답)
                _manifest_skus = {}   # {invoice: {sku_id}}
                _manifest_box_count = {}
                for _box in _all_boxes:
                    _inv = _box.get('invoice_number')
                    if not _inv:
                        continue
                    _manifest_skus.setdefault(_inv, set()).update(
                        p['sku_id'] for p in _box['products']
                    )
                    _manifest_box_count[_inv] = _manifest_box_count.get(_inv, 0) + 1

                # 송장별 SKU 집계 (시트 = 검증 대상)
                def _norm_sku(v):
                    s = str(v or '').strip()
                    if not s:
                        return ''
                    try:
                        return str(int(float(s)))  # "71572440.0" → "71572440"
                    except (ValueError, TypeError):
                        return s

                _sheet_skus = {}
                for _, _row in _df_pick.iterrows():
                    _inv = str(_row.get('쉽먼트운송장번호', '') or '').strip()
                    _sku = _norm_sku(_row.get(_sku_col))
                    if not _inv or not _sku:
                        continue
                    _sheet_skus.setdefault(_inv, set()).add(_sku)

                # 비교 — 시트 기준으로 검증 (시트의 SKU가 매니페스트에 있는가?)
                _issues = []            # 오분류 있는 송장
                _sheet_only_inv = set(_sheet_skus.keys()) - set(_manifest_skus.keys())
                _clean_count = 0
                for _inv, _ssk in _sheet_skus.items():
                    _msk = _manifest_skus.get(_inv, set())
                    if not _msk:  # 매니페스트 자체가 없는 송장 → 별도 처리
                        continue
                    _wrong = _ssk - _msk   # 🚨 시트에만 있음 = 오분류
                    _missing = _msk - _ssk  # ℹ️ 매니페스트에만 있음 = 미도착 또는 참고
                    if not _wrong:
                        _clean_count += 1
                    else:
                        _issues.append({
                            'invoice': _inv,
                            'n_sheet': len(_ssk),
                            'n_manifest': len(_msk),
                            'wrong_in_sheet': sorted(_wrong),
                            'missing_vs_manifest': sorted(_missing),
                            'n_boxes': _manifest_box_count.get(_inv, 0),
                        })

                # 전체 결과
                _total_sheet_inv = len(_sheet_skus)
                _matched_inv = len(set(_sheet_skus.keys()) & set(_manifest_skus.keys()))

                st.markdown('### 검증 결과')
                _c1, _c2, _c3, _c4 = st.columns(4)
                with _c1:
                    st.metric('시트 송장', f'{_total_sheet_inv}건')
                with _c2:
                    st.metric('송장 매칭', f'{_matched_inv}건',
                              help='시트 송장 중 매니페스트에도 존재하는 송장 수')
                with _c3:
                    st.metric('🚨 오분류 송장', f'{len(_issues)}건',
                              help='시트에 매니페스트에 없는 상품이 배정된 송장')
                with _c4:
                    st.metric('매니페스트 없음', f'{len(_sheet_only_inv)}건',
                              help='시트엔 있지만 업로드한 매니페스트에 송장 자체가 없음')

                if len(_issues) == 0 and _sheet_only_inv == set():
                    st.success(f'✅ 완전 일치 — 시트 송장 {_total_sheet_inv}건 모두 매니페스트와 SKU 일치')
                else:
                    if _sheet_only_inv:
                        with st.expander(f'ℹ️ 매니페스트에 없는 송장 ({len(_sheet_only_inv)}건) — 매니페스트 PDF 추가 업로드 필요할 수도'):
                            st.code('\n'.join(sorted(_sheet_only_inv)))

                    if _issues:
                        st.error(f'🚨 오분류 발견 — {len(_issues)}개 송장에 매니페스트에 없는 상품이 배정되어 있음. 시트에서 해당 행의 송장번호를 올바르게 수정하세요.')
                        _rows = []
                        for _it in _issues:
                            _rows.append({
                                '송장번호': _it['invoice'],
                                '🚨 오분류 SKU': len(_it['wrong_in_sheet']),
                                '시트 SKU 총': _it['n_sheet'],
                                '매니페스트 SKU 총': _it['n_manifest'],
                                '미도착 가능': len(_it['missing_vs_manifest']),
                                '박스 수': _it['n_boxes'],
                            })
                        import pandas as _pdv
                        st.dataframe(_pdv.DataFrame(_rows),
                                     use_container_width=True, hide_index=True)

                        for _it in _issues:
                            with st.expander(f"🚨 송장 {_it['invoice']} — 오분류 {len(_it['wrong_in_sheet'])}개"):
                                st.error(
                                    f"**시트에 잘못 배정된 SKU ({len(_it['wrong_in_sheet'])}개)** — "
                                    f"이 송장의 매니페스트에 없는 상품임:\n\n"
                                    + ', '.join(f'`{s}`' for s in _it['wrong_in_sheet'])
                                )
                                if _it['missing_vs_manifest']:
                                    st.info(
                                        f"참고: 매니페스트엔 있는데 시트에 없는 SKU ({len(_it['missing_vs_manifest'])}개) — "
                                        f"아직 도착 안 했을 수 있음:\n\n"
                                        + ', '.join(f'`{s}`' for s in _it['missing_vs_manifest'])
                                    )

    # ── 모드 토글 (피킹 검증 ↔ 입고 분류) ──
    st.divider()
    work_mode = st.radio(
        "🎯 작업 모드 선택",
        options=["📥 입고 분류", "📤 피킹 검증"],
        index=0,
        horizontal=True,
        key="pick_work_mode",
        help="입고 분류 = 배대지 박스 열고 박스별로 분류 / 피킹 검증 = 송장별 출고 박스 채우기",
    )
    st.divider()

    if work_mode == "📤 피킹 검증":
        if not st.session_state.pick_selected_shipment:
            # ── 송장번호 선택 ──
            st.markdown('<div class="shipment-input">', unsafe_allow_html=True)
            st.markdown("### 📋 쉽먼트 선택")
            st.caption("송장번호 입력(또는 바코드 스캔) 후 Enter → 자동으로 피킹 시작. 여러 개면 쉼표로 구분 후 🚀 피킹 시작 클릭.")

            pick_df = st.session_state.pick_df_출고

            # 스캔 후 입력창 초기화용 카운터
            if 'pick_ship_input_counter' not in st.session_state:
                st.session_state.pick_ship_input_counter = 0

            p_col1, p_col2 = st.columns([2, 1])
            with p_col1:
                _ship_input_key = f"pick_shipment_input_{st.session_state.pick_ship_input_counter}"
                input_shipment = st.text_input(
                    "송장번호 입력 (Enter로 자동 시작)",
                    placeholder="예: 461938764685 (스캐너로 스캔 후 Enter 또는 자동 입력)",
                    key=_ship_input_key,
                )
            with p_col2:
                centers = ["전체"] + sorted(pick_df["물류센터(FC)"].unique().tolist()) if "물류센터(FC)" in pick_df.columns else ["전체"]
                center = st.selectbox("물류센터", centers, key="pick_center_filter")

            # ── 자동 시작 로직 (단일 유효 송장 입력 시) ──
            if input_shipment and input_shipment.strip():
                # 쉼표/줄바꿈/공백으로 토큰 분리
                _toks = [t.strip() for t in re.split(r'[,\s\n]+', input_shipment.strip()) if t.strip()]
                if len(_toks) == 1:
                    _stgt = _toks[0]
                    _valid_ids = list(pick_df["쉽먼트운송장번호"].unique())
                    _resolved = None
                    if _stgt in _valid_ids:
                        _resolved = _stgt
                    else:
                        _mm = [s for s in _valid_ids if s.endswith(_stgt)]
                        if len(_mm) == 1:
                            _resolved = _mm[0]
                    if _resolved:
                        pick_init_picking([_resolved])
                        st.session_state.pick_start_audio_pending = True
                        st.session_state.pick_ship_input_counter += 1
                        st.rerun()

            pick_df = st.session_state.pick_df_출고
            if center != "전체" and "물류센터(FC)" in pick_df.columns:
                filtered = pick_df[pick_df["물류센터(FC)"] == center]
            else:
                filtered = pick_df

            summary = filtered.groupby("쉽먼트운송장번호").agg(
                SKU수=("바코드", "nunique"), 총수량=("수량", "sum"),
            ).reset_index().sort_values("총수량", ascending=False)

            selected_shipment = st.selectbox(
                "또는 목록에서 선택",
                options=summary["쉽먼트운송장번호"].tolist(),
                format_func=lambda x: (
                    f"{'✅ ' if x in st.session_state.pick_completed_shipments else ''}"
                    f"{x[-6:]} | "
                    f"{summary[summary['쉽먼트운송장번호']==x]['SKU수'].values[0]}종 "
                    f"{summary[summary['쉽먼트운송장번호']==x]['총수량'].values[0]}개"
                ),
                key="pick_shipment_select",
            )

            # 다중 송장 파싱: 쉼표/공백/줄바꿈으로 구분
            input_targets = []
            if input_shipment and input_shipment.strip():
                for token in re.split(r'[,\s\n]+', input_shipment.strip()):
                    token = token.strip()
                    if token:
                        input_targets.append(token)

            if st.button("🚀 피킹 시작", type="primary", use_container_width=True, key="pick_start_btn"):
                valid_ids = list(pick_df["쉽먼트운송장번호"].unique())
                resolved = []
                errors = []
                if input_targets:
                    for tgt in input_targets:
                        if tgt in valid_ids:
                            resolved.append(tgt)
                        else:
                            matches = [s for s in valid_ids if s.endswith(tgt)]
                            if len(matches) == 1:
                                resolved.append(matches[0])
                            elif len(matches) > 1:
                                errors.append(f"'{tgt}'에 매칭되는 쉽먼트가 {len(matches)}개입니다.")
                            else:
                                errors.append(f"'{tgt}'에 해당하는 쉽먼트를 찾을 수 없습니다.")
                elif selected_shipment:
                    resolved.append(selected_shipment)

                if errors:
                    for err in errors:
                        st.error(err)
                elif resolved:
                    pick_init_picking(resolved)
                    st.rerun()
                else:
                    st.warning("송장번호를 입력하거나 목록에서 선택해주세요.")
            st.markdown('</div>', unsafe_allow_html=True)
        else:
            # ── 피킹 진행 화면 ──
            shipment_id = st.session_state.pick_selected_shipment

            # 신규 진입 시 "확인을 시작하세요" 음성 안내 (1회)
            if st.session_state.get('pick_start_audio_pending'):
                from streamlit.components.v1 import html as _st_start_html
                _st_start_html("""<script>
                try{
                    window.speechSynthesis.cancel();
                    setTimeout(function(){
                        var u = new SpeechSynthesisUtterance('확인을 시작하세요');
                        u.lang = 'ko-KR';
                        u.rate = 1.15;
                        u.volume = 1.0;
                        var voices = window.speechSynthesis.getVoices();
                        var koVoice = voices.find(v => v.lang && v.lang.startsWith('ko'));
                        if (koVoice) u.voice = koVoice;
                        window.speechSynthesis.speak(u);
                    }, 120);
                }catch(e){}
                </script>""", height=0)
                st.session_state.pick_start_audio_pending = False

            hcol1, hcol2, hcol3 = st.columns([3, 1, 1])
            with hcol1:
                item0 = list(st.session_state.pick_picking_state.values())[0] if st.session_state.pick_picking_state else {}
                ships = st.session_state.get('pick_selected_shipments', [shipment_id])
                if len(ships) > 1:
                    ship_lines = " | ".join([f"**{i+1}번박스:** `{s[-6:]}`" for i, s in enumerate(ships)])
                    st.markdown(f"{ship_lines} | **센터:** {item0.get('물류센터','')}")
                else:
                    st.markdown(f"**쉽먼트:** `{shipment_id}` | **센터:** {item0.get('물류센터','')} | **회차:** {item0.get('회차기호','')}")
            with hcol2:
                if st.button("➕ 쉽먼트 추가", use_container_width=True, key="pick_add_btn"):
                    st.session_state.pick_show_add_input = True
                    st.rerun()
            with hcol3:
                if st.button("🔄 다른 쉽먼트", use_container_width=True, key="pick_change_btn"):
                    # 쉽먼트 관련 상태 완전 초기화 (로그/재고/완료목록은 유지)
                    st.session_state.pick_selected_shipment = None
                    st.session_state.pick_selected_shipments = []
                    st.session_state.pick_picking_state = {}
                    st.session_state.pick_shortage_items = []
                    st.session_state.pick_last_scan_result = None
                    st.session_state.pick_scan_counter = 0
                    st.session_state.pick_show_add_input = False
                    # 다량 모드 상태 초기화
                    st.session_state.pick_next_qty = 1
                    st.session_state.pick_qty_input_mode = False
                    st.rerun()

            # 쉽먼트 추가 입력 영역
            if st.session_state.get('pick_show_add_input'):
                with st.container():
                    ac1, ac2, ac3 = st.columns([3, 1, 1])
                    with ac1:
                        add_input = st.text_input("추가할 송장번호", key="pick_add_shipment_input",
                                                  placeholder="송장번호 입력 후 추가 클릭")
                    with ac2:
                        if st.button("✅ 추가", use_container_width=True, key="pick_add_confirm"):
                            target = (add_input or '').strip()
                            if target:
                                valid_ids = list(st.session_state.pick_df_출고["쉽먼트운송장번호"].unique())
                                resolved = None
                                if target in valid_ids:
                                    resolved = target
                                else:
                                    matches = [s for s in valid_ids if s.endswith(target)]
                                    if len(matches) == 1:
                                        resolved = matches[0]
                                if resolved and resolved not in ships:
                                    new_ships = list(ships) + [resolved]
                                    pick_init_picking(new_ships)
                                    st.session_state.pick_show_add_input = False
                                    if 'pick_add_shipment_input' in st.session_state:
                                        del st.session_state['pick_add_shipment_input']
                                    st.rerun()
                                elif resolved in ships:
                                    st.warning("이미 추가된 송장입니다")
                                else:
                                    st.error(f"'{target}' 송장을 찾을 수 없습니다")
                    with ac3:
                        if st.button("❌ 취소", use_container_width=True, key="pick_add_cancel"):
                            st.session_state.pick_show_add_input = False
                            st.rerun()

            st.markdown("---")

            # ── 바코드 스캔 (fragment으로 감싸서 전체 앱 리런 없이 조각만 재실행) ──
            def _pick_scan_rerun():
                """Fragment 내부면 조각 리런, 아니면 전체 리런 (구버전 Streamlit fallback)."""
                try:
                    st.rerun(scope='fragment')
                except TypeError:
                    st.rerun()
                except Exception:
                    st.rerun()

            _pick_use_fragment = getattr(st, 'fragment', lambda f: f)

            @_pick_use_fragment
            def _pick_scan_fragment():
                # ── 진행률 (매 스캔마다 갱신되도록 fragment 안에서 계산) ──
                _prog = pick_get_progress()
                _sid = st.session_state.pick_selected_shipment or ''
                pc1, pc2, pc3, pc4, pc5 = st.columns(5)
                pc1.metric("스캔", f"{_prog['scanned']}/{_prog['total']}")
                pc2.metric("SKU 완료", f"{_prog['done_skus']}/{_prog['skus']}")
                pc3.metric("진행률", f"{_prog['pct']:.0%}")
                pc4.metric("초과 스캔", f"{_prog['over']}건",
                           delta=f"+{_prog['over']}" if _prog['over'] > 0 else None, delta_color="inverse")
                pc5.metric("재고 부족", f"{_prog['shortage']}건",
                           delta=f"{_prog['shortage']}" if _prog['shortage'] > 0 else None, delta_color="inverse")
                st.progress(_prog["pct"])

                if _prog["is_complete"]:
                    st.markdown(
                        f'<div class="scan-complete">'
                        f'<strong style="font-size:1.4rem;">🎉 검증확인이 완료되었습니다. 출고하세요!</strong><br>'
                        f'<span style="font-size:1.05rem;">쉽먼트 {_sid[-6:]} — {_prog["total"]}개 전부 검증 완료</span>'
                        f'</div>',
                        unsafe_allow_html=True)
                    _newly_done = _sid not in st.session_state.pick_completed_shipments
                    st.session_state.pick_completed_shipments.add(_sid)
                    if _newly_done:
                        from streamlit.components.v1 import html as _st_html_done
                        _st_html_done("""<script>
                        try{
                            window.speechSynthesis.cancel();
                            setTimeout(function(){
                                var u = new SpeechSynthesisUtterance('검증확인이 완료되었습니다. 출고하세요');
                                u.lang = 'ko-KR';
                                u.rate = 1.15;
                                u.volume = 1.0;
                                var voices = window.speechSynthesis.getVoices();
                                var koVoice = voices.find(v => v.lang && v.lang.startsWith('ko'));
                                if (koVoice) u.voice = koVoice;
                                window.speechSynthesis.speak(u);
                            }, 120);
                        }catch(e){}
                        </script>""", height=0)

                st.markdown("---")

                # 다량 모드 상태 초기화
                if 'pick_next_qty' not in st.session_state:
                    st.session_state.pick_next_qty = 1
                if 'pick_qty_input_mode' not in st.session_state:
                    st.session_state.pick_qty_input_mode = False

                # 수량 입력 모드: 숫자 입력 후 Enter
                if st.session_state.pick_qty_input_mode:
                    st.warning('🔢 **수량을 입력하세요** — 숫자 입력 후 Enter')
                    pick_qty_text_key = f'pick_qty_text_{st.session_state.pick_scan_counter}'
                    pick_qty_text = st.text_input(
                        '다량 수량',
                        key=pick_qty_text_key,
                        placeholder='숫자 입력 후 Enter (예: 50)',
                        label_visibility='collapsed',
                    )
                    if pick_qty_text:
                        try:
                            _qv = int(pick_qty_text.strip())
                            if _qv >= 1:
                                st.session_state.pick_next_qty = _qv
                                st.session_state.pick_qty_input_mode = False
                                st.session_state.pick_scan_counter += 1
                                _pick_scan_rerun()
                        except ValueError:
                            st.error('숫자만 입력 가능합니다')

                # 수량 표시 + 1개 모드 리셋 버튼
                pqcol1, pqcol2 = st.columns([1, 1])
                with pqcol1:
                    if st.session_state.pick_next_qty > 1:
                        st.markdown(
                            f'<div style="background:#f59e0b;color:white;padding:0.5rem;border-radius:6px;text-align:center;font-weight:bold;font-size:1.1rem">'
                            f'📦 다음 스캔: {st.session_state.pick_next_qty}개'
                            f'</div>', unsafe_allow_html=True)
                    else:
                        st.markdown(
                            '<div style="background:#e5e7eb;padding:0.5rem;border-radius:6px;text-align:center">'
                            '1개 모드'
                            '</div>', unsafe_allow_html=True)
                with pqcol2:
                    if st.session_state.pick_next_qty > 1 and not st.session_state.pick_qty_input_mode:
                        if st.button('🔄 1개 모드로 복귀', key='pick_qty_reset', use_container_width=True):
                            st.session_state.pick_next_qty = 1
                            _pick_scan_rerun()

                scan_key = f"pick_scan_{st.session_state.pick_scan_counter}"
                scanned = st.text_input("🔫 바코드 스캔 (스캐너 또는 직접 입력)", key=scan_key,
                                        placeholder="스캐너 대기 중... 바코드 (여러 개면 #MULTI 먼저)")
                if scanned:
                    _pick_qty = int(st.session_state.get('pick_next_qty', 1) or 1)
                    pick_process_scan(scanned, qty=_pick_qty)
                    _pick_scan_rerun()

                # 바코드 입력창에 자동 포커스 유지
                from streamlit.components.v1 import html as _st_html
                _st_html("""<script>
                (function(){
                    const doc = window.parent.document;
                    function findScanInput() {
                        const inputs = doc.querySelectorAll('input[type="text"]');
                        for (const inp of inputs) {
                            if (inp.placeholder && inp.placeholder.includes('스캐너')) {
                                return inp;
                            }
                        }
                        return null;
                    }
                    function isInteractingOther(e) {
                        // 드래그 중(스크롤바/리사이즈 등)이거나 dataframe 내부면 포커스 복귀 스킵
                        const t = e && e.target;
                        if (!t) return false;
                        if (t.closest && (
                            t.closest('[data-testid="stDataFrame"]') ||
                            t.closest('canvas') ||
                            t.closest('.glideDataEditor') ||
                            t.closest('[role="grid"]') ||
                            t.closest('[data-testid="stExpander"]')
                        )) return true;
                        return false;
                    }
                    function focusScan(e) {
                        if (e && isInteractingOther(e)) return;
                        const inp = findScanInput();
                        if (inp && doc.activeElement !== inp) {
                            // preventScroll: 포커스 이동 시 브라우저 자동 스크롤 방지
                            inp.focus({preventScroll: true});
                        }
                    }
                    // 즉시 포커스
                    focusScan();
                    // 짧은 간격으로 반복 (0.5초) — 스크롤 조작 방해 최소화
                    if (window._scanFocusInterval) clearInterval(window._scanFocusInterval);
                    window._scanFocusInterval = setInterval(function(){
                        // 사용자가 다른 곳과 상호작용 중(selection, 드래그)이면 스킵
                        const sel = doc.getSelection && doc.getSelection();
                        if (sel && sel.toString().length > 0) return;
                        focusScan();
                    }, 500);
                    // DOM 변경 감지는 제거 (dataframe 스크롤이 DOM 변경 유발 → 루프)
                    if (window._scanObserver) { window._scanObserver.disconnect(); window._scanObserver = null; }
                    // 다른 곳 클릭해도 입력창으로 복귀 (버튼/링크/테이블/expander 제외)
                    doc.addEventListener('click', function(e){
                        const tag = (e.target.tagName||'').toLowerCase();
                        if (tag === 'button' || tag === 'a' || tag === 'input' || tag === 'select' || tag === 'textarea') return;
                        if (isInteractingOther(e)) return;
                        setTimeout(function(){ focusScan(e); }, 50);
                    }, true);
                })();
                </script>""", height=0)

                r = st.session_state.pick_last_scan_result
                if r:
                    css_class = {"ok":"scan-ok","over":"scan-warning","error":"scan-error","shortage":"scan-shortage"}.get(r["status"],"scan-ok")
                    st.markdown(
                        f'<div class="{css_class}"><strong style="font-size:1.1rem;">{r["message"]}</strong><br>{r["detail"]}</div>',
                        unsafe_allow_html=True)
                    # 스캔 결과 소리 (비프음)
                    sound_js = {
                        "ok": "o.frequency.value=880;g.gain.value=0.3;o.start();setTimeout(()=>g.gain.value=0,150);setTimeout(()=>o.stop(),200);",
                        "error": "o.type='square';o.frequency.value=200;g.gain.value=0.5;o.start();setTimeout(()=>{o.frequency.value=150},150);setTimeout(()=>g.gain.value=0,500);setTimeout(()=>o.stop(),600);",
                        "over": "o.type='sawtooth';o.frequency.value=400;g.gain.value=0.4;o.start();setTimeout(()=>{o.frequency.value=300},100);setTimeout(()=>g.gain.value=0,300);setTimeout(()=>o.stop(),400);",
                        "shortage": "o.frequency.value=600;g.gain.value=0.3;o.start();setTimeout(()=>{o.frequency.value=400},100);setTimeout(()=>g.gain.value=0,250);setTimeout(()=>o.stop(),300);",
                    }
                    js_code = sound_js.get(r["status"], sound_js["ok"])

                    # 음성 안내 (한국어 TTS)
                    box_label = r.get("박스", "")
                    # 1번박스 → 일번박스 형태로 변환
                    _KOR_NUMS = {'1':'일','2':'이','3':'삼','4':'사','5':'오','6':'육','7':'칠','8':'팔','9':'구','10':'십'}
                    def _kor_box(label):
                        import re as _re
                        m = _re.match(r'(\d+)번박스', label or '')
                        if not m:
                            return label
                        n = m.group(1)
                        kor = _KOR_NUMS.get(n, n)
                        return f"{kor}번박스"
                    box_kor = _kor_box(box_label)

                    ships_count = len(st.session_state.get('pick_selected_shipments', []))
                    if r["status"] == "error":
                        speak_text = "없는 상품 입니다"
                    elif r["status"] == "over":
                        speak_text = "수량 초과"
                    elif r["status"] == "shortage":
                        if ships_count <= 1:
                            speak_text = "입고완료 재고 부족"
                        else:
                            speak_text = f"{box_kor} 재고 부족" if box_kor else "재고 부족"
                    elif ships_count <= 1:
                        speak_text = "입고완료"
                    elif box_kor:
                        speak_text = f"{box_kor}"
                    else:
                        speak_text = "확인"

                    # JS 문자열 안전 이스케이프
                    speak_text_js = speak_text.replace("'", "\\'").replace('"', '\\"')
                    # 매 스캔마다 새 컴포넌트로 강제 재실행 (같은 박스도 소리 나도록)
                    scan_id = st.session_state.pick_scan_counter

                    from streamlit.components.v1 import html as st_html
                    st_html(f"""<script>
                    // scan_id={scan_id} (강제 재실행용)
                    try{{var a=new(window.AudioContext||window.webkitAudioContext)();var o=a.createOscillator();var g=a.createGain();o.connect(g);g.connect(a.destination);{js_code}}}catch(e){{}}
                    try{{
                        window.speechSynthesis.cancel();
                        setTimeout(function(){{
                            var u = new SpeechSynthesisUtterance('{speak_text_js}');
                            u.lang = 'ko-KR';
                            u.rate = 1.3;
                            u.volume = 1.0;
                            var voices = window.speechSynthesis.getVoices();
                            var koVoice = voices.find(v => v.lang && v.lang.startsWith('ko'));
                            if (koVoice) u.voice = koVoice;
                            window.speechSynthesis.speak(u);
                        }}, 100);
                    }}catch(e){{}}
                    </script>""", height=0)

                # ── 피킹 현황 (매 스캔마다 갱신되도록 fragment 안에서 렌더링) ──
                st.markdown("---")
                st.subheader("📋 피킹 현황")
                rows = []
                for bc, info in st.session_state.pick_picking_state.items():
                    s, n = info["스캔수량"], info["필요수량"]
                    if s > n: status_txt = f"⚠️ 초과 ({s}/{n})"
                    elif s >= n: status_txt = "✅ 완료"
                    elif s > 0: status_txt = f"🔄 {s}/{n}"
                    else: status_txt = "⬜ 대기"
                    inv = info.get("배대지잔여")
                    ship_boxes = info.get("쉽먼트박스목록", [])
                    ship_box_str = ",".join(ship_boxes) if ship_boxes else ""
                    rows.append({
                        "상태": status_txt, "바코드": bc,
                        "상품명": info["상품명"],
                        "쉽먼트박스": ship_box_str,
                        "필요": n, "스캔": s, "남은": max(0, n - s),
                        "회차": info.get("회차기호",""), "박스": info.get("박스번호",""),
                        "배대지재고": f"{inv}" if inv is not None else "-",
                    })
                pick_order = {"🔄":0,"⬜":1,"✅":2,"⚠️":3}
                rows.sort(key=lambda x: pick_order.get(x["상태"][0], 9))
                st.dataframe(_pd.DataFrame(rows), use_container_width=True, hide_index=True,
                             height=min(500, len(rows) * 38 + 40))

                shortage = st.session_state.get("pick_shortage_items", [])
                if shortage:
                    with st.expander(f"⛔ 부족분 — 피킹 불가 ({len(shortage)}건)", expanded=False):
                        st.caption("출고지시서에 '부족'으로 표시된 항목입니다.")
                        st.dataframe(_pd.DataFrame(shortage), use_container_width=True, hide_index=True)

                if st.session_state.pick_scan_log:
                    with st.expander(f"📜 스캔 로그 ({len(st.session_state.pick_scan_log)}건)"):
                        log_display = []
                        for entry in reversed(st.session_state.pick_scan_log[-50:]):
                            icon = {"ok":"✅","over":"⚠️","error":"🚨","shortage":"📦"}.get(entry["status"],"?")
                            log_display.append({"시간":entry["시간"],"결과":icon,"바코드":entry["barcode"],"내용":entry["message"]})
                        st.dataframe(_pd.DataFrame(log_display), use_container_width=True, hide_index=True)

                    st.download_button(
                        "📥 스캔 로그 CSV",
                        data=_pd.DataFrame(st.session_state.pick_scan_log).to_csv(index=False, encoding="utf-8-sig"),
                        file_name=f"picking_log_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                        mime="text/csv", use_container_width=True, key="pick_log_dl")

            _pick_scan_fragment()

    elif work_mode == "📥 입고 분류":
        import pandas as _pd2
        df_sort = st.session_state.pick_df_출고

        # ── 송장별 박스번호: 시트 M열(기존값) 우선 보존 + 신규만 부여 ──
        # 박스번호가 시트 M열에 영구 저장되므로 발주 취소돼도 재정렬 안 됨.
        sort_items_for_box = []
        for _, row in df_sort.iterrows():
            sort_items_for_box.append({
                'shipmentNumber': str(row.get('쉽먼트운송장번호', '') or '').strip(),
                'logisticsCenter': str(row.get('물류센터(FC)', '') or '').strip(),
                'expectedDate': str(row.get('입고예정일(EDD)', '') or row.get('입고예정일', '') or '').strip(),
                'boxNumber': str(row.get('박스번호', '') or '').strip(),
            })

        # 1) 시트 M열에서 기존 박스번호 읽기 (gsheet 모드 + 세션 캐시)
        #    실패(None) 시 캐시 저장 안 함 → 다음 rerun에서 재시도
        _pick_url = st.session_state.get('pick_sheet_url_출고', '')
        _pick_tab = st.session_state.get('pick_sheet_tab_출고', '')
        _existing_box_map = {}
        _read_failed = False
        _gsheet_ready = (st.session_state.get('pick_use_gsheet')
                         and st.session_state.get('pick_gsheet_client')
                         and _pick_url and _pick_tab)
        if _gsheet_ready:
            _cache_key = f"_pick_existing_box_{_pick_url}_{_pick_tab}"
            if _cache_key in st.session_state:
                _existing_box_map = st.session_state[_cache_key]
            else:
                _read = pick_read_box_numbers(
                    st.session_state.pick_gsheet_client, _pick_url, _pick_tab
                )
                if _read is None:
                    _read_failed = True
                    st.warning('⚠️ 시트 M열 읽기 실패 — 이번 세션에서 박스번호 쓰기는 건너뜁니다. 페이지를 새로고침하세요.')
                else:
                    _existing_box_map = _read
                    st.session_state[_cache_key] = _read

        # 2) 기존 유지 + 신규 송장만 순차 부여
        sort_ship_to_box = assign_box_numbers_with_existing(
            sort_items_for_box, _existing_box_map
        )

        # 3) 신규 부여된 것만 시트 M열에 동기 기록 (기존값 보존, 실패 시 캐시 갱신 안 함)
        if _gsheet_ready and not _read_failed:
            _new_only = {s: n for s, n in sort_ship_to_box.items()
                         if s not in _existing_box_map}
            _written_key = f"_pick_box_written_{_pick_url}_{_pick_tab}"
            _last_written = st.session_state.get(_written_key, {})
            _to_write = {s: n for s, n in _new_only.items()
                         if _last_written.get(s) != n}
            if _to_write:
                with st.spinner(f'📝 시트 M열에 박스번호 {len(_to_write)}건 기록 중...'):
                    _result = pick_write_box_numbers(
                        st.session_state.pick_gsheet_client,
                        _pick_url, _pick_tab, _to_write, only_empty=True,
                    )
                if _result >= 0:
                    # 성공 시에만 캐시 갱신
                    _merged = dict(_existing_box_map)
                    _merged.update(_new_only)
                    st.session_state[_cache_key] = _merged
                    _last_written.update(_to_write)
                    st.session_state[_written_key] = _last_written
                else:
                    st.warning('⚠️ 시트 M열 쓰기 실패 — 박스번호가 시트에 저장되지 않았습니다. 다음 rerun에서 재시도됩니다.')

        # 쉽먼트 재출력 탭에서 재사용하도록 세션에 저장
        st.session_state['pick_ship_to_box'] = sort_ship_to_box

        # ── 바코드 → (배대지박스, 송장박스, 수량, 상품명, 송장) 매핑 ──
        def _build_sort_state():
            state = {}  # barcode → {상품명, items: [{dapae_box, out_box, ship, needed, scanned}]}
            for _, row in df_sort.iterrows():
                bc = str(row.get('바코드', '')).strip()
                ship = str(row.get('쉽먼트운송장번호', '')).strip()
                qty = int(row.get('수량', 0) or 0)
                name = str(row.get('상품명', '')).strip()
                if not bc or not ship or qty <= 0:
                    continue
                # 출고 박스번호 (송장별 자동 부여, 국내재고 전용 송장은 None)
                out_box = sort_ship_to_box.get(ship)
                if out_box is None:
                    continue  # 국내재고/부족 전용 송장은 박스 분류 제외
                # 배대지 박스번호 (K열에서 파싱한 값: M1, W3 등)
                dapae_raw = str(row.get('박스넘버') or '').strip().upper()
                if not dapae_raw or dapae_raw == 'NAN':
                    dapae_raw = ''
                # 시트 L열(확인수량)에서 이전 진행 상태 복원
                try:
                    prev_scanned = int(row.get('확인수량', 0) or 0)
                except (ValueError, TypeError):
                    prev_scanned = 0
                prev_scanned = max(0, min(prev_scanned, qty))
                if bc not in state:
                    state[bc] = {'상품명': name, 'items': []}
                state[bc]['items'].append({
                    'out_box': str(out_box),      # 출고 박스번호 (1~58번)
                    'dapae_box': dapae_raw,       # 배대지 박스번호 (M1, W3 등)
                    'box_num': str(out_box),      # 기존 코드 호환: 음성/화면 안내용
                    'box_key': str(out_box),      # 기존 코드 호환
                    'sym': '',
                    'ship': ship,
                    'needed': qty,
                    'scanned': prev_scanned,      # L열에서 복원한 값
                })
            return state

        # 초기화
        if 'sort_state' not in st.session_state or st.session_state.get('sort_data_ver') != id(df_sort):
            st.session_state.sort_state = _build_sort_state()
            st.session_state.sort_data_ver = id(df_sort)
            st.session_state.sort_scan_counter = 0
            st.session_state.sort_last_result = None

        sort_state = st.session_state.sort_state

        # ── 배대지 박스별 집계 (작업 순서 추천용) ──
        # 배대지 박스 = K열에서 파싱된 M1, W3 등 (작업자가 물리적으로 열 박스)
        import re as _re_bq
        box_qty_map = {}  # dapae_box → {box_num, total_qty, ships, out_boxes}
        for v in sort_state.values():
            for it in v['items']:
                dp = str(it.get('dapae_box', '')).strip().upper()
                if not dp or dp == 'NAN':
                    continue
                if not _re_bq.match(r'^[A-Z]*\d+$', dp):
                    continue
                ent = box_qty_map.setdefault(dp, {
                    'box_num': dp, 'sym': '',
                    'total_qty': 0, 'ships': set(), 'out_boxes': set(),
                })
                ent['total_qty'] += it['needed']
                if it.get('ship'):
                    ent['ships'].add(it['ship'])
                if it.get('out_box'):
                    ent['out_boxes'].add(it['out_box'])

        # ── 출고(송장) 박스별 집계 (라벨 PDF / 크기 분류용) ──
        # 송장박스 = 1~58번 (우리가 준비하는 박스, 사용자 라벨 부착용)
        out_box_map = {}  # out_box(str) → {box_num, total_qty, ship}
        for v in sort_state.values():
            for it in v['items']:
                ob = str(it.get('out_box', '')).strip()
                if not ob or not ob.isdigit():
                    continue
                ent = out_box_map.setdefault(ob, {
                    'box_num': ob,
                    'total_qty': 0,
                    'ship': it.get('ship', ''),
                })
                ent['total_qty'] += it['needed']

        def _box_size(qty):
            """수량 기준으로 박스 크기 분류"""
            if qty >= 50:
                return ('대', '🟢')
            elif qty >= 30:
                return ('중', '🟡')
            else:
                return ('소', '🔵')

        def _box_sort_key(box_str):
            """박스 번호 정렬 키: (알파벳 부분, 숫자 부분) 튜플
            'W1' → ('W', 1), 'M3' → ('M', 3), '1' → ('', 1)
            """
            import re as _re_sk
            m = _re_sk.match(r'^([A-Z]*)(\d+)$', str(box_str).upper())
            if m:
                return (m.group(1), int(m.group(2)))
            return (str(box_str), 0)

        # 송장(출고) 박스 크기별 그룹 (우리가 준비할 박스 기준)
        boxes_large = []
        boxes_med = []
        boxes_small = []
        for key, info in out_box_map.items():
            size_label, _ = _box_size(info['total_qty'])
            entry = (_box_sort_key(key), info['box_num'], info['total_qty'])
            if size_label == '대':
                boxes_large.append(entry)
            elif size_label == '중':
                boxes_med.append(entry)
            else:
                boxes_small.append(entry)
        boxes_large.sort()
        boxes_med.sort()
        boxes_small.sort()

        # 송장 박스 크기 매핑 (스캔 결과 표시용)
        box_size_lookup = {}  # out_box → (size_label, emoji)
        for key, info in out_box_map.items():
            box_size_lookup[key] = _box_size(info['total_qty'])

        # ── 상단 요약 ──
        total_qty = sum(it['needed'] for v in sort_state.values() for it in v['items'])
        total_scanned = sum(it['scanned'] for v in sort_state.values() for it in v['items'])

        hh1, hh2, hh3, hh4 = st.columns(4)
        hh1.metric('출고박스', f'{len(out_box_map)}개')
        hh2.metric('배대지박스', f'{len(box_qty_map)}개')
        hh3.metric('스캔 진행', f'{total_scanned}/{total_qty}')
        hh4.metric('진행률', f'{(total_scanned/total_qty*100 if total_qty else 0):.0f}%')

        # ── 박스 크기별 준비 안내 ──
        with st.expander('📦 박스 준비 안내 (크기별)', expanded=True):
            size_rows = [
                {
                    '크기': '🟢 대형',
                    '기준': '50개 이상',
                    '개수': f'{len(boxes_large)}개',
                    '박스 번호': ', '.join(f'{b[1]}번({b[2]})' for b in boxes_large) if boxes_large else '-',
                },
                {
                    '크기': '🟡 중형',
                    '기준': '30~49개',
                    '개수': f'{len(boxes_med)}개',
                    '박스 번호': ', '.join(f'{b[1]}번({b[2]})' for b in boxes_med) if boxes_med else '-',
                },
                {
                    '크기': '🔵 소형',
                    '기준': '30개 미만',
                    '개수': f'{len(boxes_small)}개',
                    '박스 번호': ', '.join(f'{b[1]}번({b[2]})' for b in boxes_small) if boxes_small else '-',
                },
            ]
            st.dataframe(_pd2.DataFrame(size_rows), use_container_width=True, hide_index=True)
            st.caption(f'💡 총 {len(out_box_map)}개 박스 준비 ・ 박스 옆 괄호는 들어갈 총 수량')

            # ── 폼텍 3100 라벨 PDF 다운로드 (송장박스 1~N번) ──
            label_entries = []
            for key, info in sorted(out_box_map.items(),
                                    key=lambda x: _box_sort_key(x[1]['box_num'])):
                size_lbl, size_emoji = _box_size(info['total_qty'])
                label_entries.append((info['box_num'], info['total_qty'], f'{size_emoji}{size_lbl}'))
            try:
                label_pdf_buf = create_box_labels_pdf(label_entries)
                st.download_button(
                    label=f'🏷️ 박스 라벨 PDF 다운로드 ({len(label_entries)}장 / 폼텍 3100)',
                    data=label_pdf_buf,
                    file_name=f'box_labels_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf',
                    mime='application/pdf',
                    key='sort_label_dl',
                    use_container_width=True,
                )
                st.caption('📄 폼텍 3100 (38.1×21.2mm, A4 65칸) 라벨지에 출력하세요')
            except Exception as _e:
                st.caption(f'라벨 생성 오류: {_e}')

            # #MULTI 트리거 라벨 PDF
            try:
                multi_pdf_buf = create_multi_trigger_label_pdf()
                st.download_button(
                    label='🔢 다량 입력 트리거 바코드 PDF (1회만 출력)',
                    data=multi_pdf_buf,
                    file_name='multi_trigger_label.pdf',
                    mime='application/pdf',
                    key='sort_multi_trigger_dl',
                    use_container_width=True,
                )
                st.caption('📄 A4 한 장에 큰 바코드. 인쇄해서 잘 보이는 곳에 부착하세요')
            except Exception as _e:
                st.caption(f'트리거 라벨 생성 오류: {_e}')

        # ── 송장별 필요 배대지 박스 집계 (집합 커버 계산용) ──
        # 각 송장이 어느 배대지 박스에 있는지: {송장: {배대지박스 set}}
        import re as _re_ship
        ship_need_boxes = {}  # ship_id → set of dapae_box
        for v in sort_state.values():
            for it in v['items']:
                ship = it.get('ship')
                if not ship:
                    continue
                dp = str(it.get('dapae_box', '')).strip().upper()
                if not dp or dp == 'NAN':
                    continue
                if not _re_ship.match(r'^[A-Z]*\d+$', dp):
                    continue
                ship_need_boxes.setdefault(ship, set()).add(dp)

        # ── 배대지 박스별 완료 상태 계산 (sort_state 기준) ──
        # box_done: dapae_box → True(모두 스캔됨) / False
        box_done = {}
        box_progress = {}  # dapae_box → (scanned, needed)
        for _bc_bd, _v_bd in sort_state.items():
            for _it_bd in _v_bd['items']:
                _dp_bd = str(_it_bd.get('dapae_box', '')).strip().upper()
                if not _dp_bd or _dp_bd == 'NAN':
                    continue
                s, n = box_progress.get(_dp_bd, (0, 0))
                box_progress[_dp_bd] = (s + _it_bd['scanned'], n + _it_bd['needed'])
        for _dp_bd, (s, n) in box_progress.items():
            box_done[_dp_bd] = (n > 0 and s >= n)

        # ── 시작 박스 지정 ──
        st.markdown('### 🎯 작업할 배대지 박스')

        # 1순위 추천: 미완료 박스 중 수량이 가장 많은 배대지 박스
        top_box = None
        _incomplete_boxes = {k: v for k, v in box_qty_map.items() if not box_done.get(k, False)}
        if _incomplete_boxes:
            top_box_key = max(_incomplete_boxes.keys(), key=lambda k: _incomplete_boxes[k]['total_qty'])
            top_info = _incomplete_boxes[top_box_key]
            top_out_boxes = sorted(top_info['out_boxes'], key=_box_sort_key)
            out_box_str = ', '.join(f'{b}번' for b in top_out_boxes)
            st.success(
                f'🏆 **우선 작업 추천**: **{top_box_key} 박스부터 열어주세요**\n\n'
                f'→ 수량 {top_info["total_qty"]}개\n\n'
                f'📦 **준비할 출고박스 ({len(top_out_boxes)}개)**: {out_box_str}'
            )
            top_box = top_box_key
        elif box_qty_map:
            st.success('🎉 **모든 배대지 박스 완료!**')

        # 드롭다운 정렬: 미완료 우선(수량 많은 순) → 완료는 맨 아래(알파벳+번호 순)
        def _dropdown_sort_key(item):
            k, v = item
            done = box_done.get(k, False)
            # (완료 여부, -수량, box_sort_key) — 미완료(False=0)가 완료(True=1)보다 앞
            return (1 if done else 0, -v['total_qty'], _box_sort_key(k))

        recommended_boxes = sorted(box_qty_map.items(), key=_dropdown_sort_key)
        all_box_nums_sorted = [k for k, _ in recommended_boxes]

        def _fmt_box_num(x):
            info = box_qty_map[x]
            size_lbl, size_emo = _box_size(info['total_qty'])
            ship_cnt = len(info['ships'])
            done = box_done.get(x, False)
            prefix = '✅ ' if done else ''
            suffix = ' — 완료' if done else ''
            return f"{prefix}{x}번 ({info['total_qty']}개 / {size_emo}{size_lbl} / 송장 {ship_cnt}개){suffix}"

        # 창고 수용력 + 자동 추천 박스 세트
        sac0a, sac0b = st.columns([1, 3])
        with sac0a:
            capacity = st.number_input(
                '창고 수용력',
                min_value=1, max_value=10, value=1,
                key='sort_capacity',
                help='창고에 동시에 펼쳐놓을 수 있는 배대지 박스 개수 (1 = 한 박스씩 차례대로)',
            )

        # 집합 커버 추천: 수량 절대 우선 (수량 많은 박스부터 → 빠르게 출고박스 채움)
        def _recommend_box_set(target_count, already_committed=None):
            committed = set(already_committed or [])
            picks = []
            # 완료된 박스는 추천 대상에서 제외
            available = set(k for k in box_qty_map.keys() if not box_done.get(k, False)) - committed
            while len(picks) < target_count and available:
                best = None
                best_score = (-1, -1)
                for box in available:
                    qty = box_qty_map[box]['total_qty']
                    ship_cnt = len(box_qty_map[box]['ships'])
                    # 점수: (수량, -송장수) — 수량 우선 + 동률이면 송장 적은 것(단순한 박스)
                    # 송장 적을수록 다른 배대지 박스 의존성↓ → 다른 박스 동시에 안 깔아도 됨
                    score = (qty, -ship_cnt)
                    if score > best_score:
                        best_score = score
                        best = box
                if best is None:
                    break
                picks.append(best)
                available.discard(best)
            return picks

        rec_set = _recommend_box_set(int(capacity))
        if rec_set and int(capacity) > 1:
            # 수용력이 2 이상일 때만 세트 안내 (1이면 위의 1순위 추천과 동일하므로 중복 표시 안 함)
            rec_set_sorted = sorted(rec_set, key=_box_sort_key)
            completed_with_set = sum(
                1 for need in ship_need_boxes.values() if need.issubset(set(rec_set))
            )
            with sac0b:
                st.success(
                    f'💡 **지금 열 박스 {int(capacity)}개**: '
                    f'**{", ".join(str(b) + "번" for b in rec_set_sorted)}**  '
                    f'→ 이것만 열면 **{completed_with_set}개 송장** 완성'
                )
        elif int(capacity) == 1:
            with sac0b:
                st.caption('💡 **1박스씩 처리** — 수량 많은 박스부터 차례로 끝내면 출고박스가 빠르게 채워져요. 동시 작업 원하면 수용력↑')

        # ── 활성 박스 (멀티 선택) ──
        if 'sort_active_boxes' not in st.session_state:
            st.session_state.sort_active_boxes = []

        sac1, sac2 = st.columns([4, 1])
        with sac1:
            active_boxes = st.multiselect(
                '🎯 지금 열어놓은 배대지 박스 (바코드 #1, #2... 찍으면 자동 추가)',
                options=all_box_nums_sorted,
                format_func=_fmt_box_num,
                default=st.session_state.sort_active_boxes,
                key='sort_active_boxes_ms',
                help='여러 박스를 동시에 선택 가능. 라벨의 #N 바코드를 찍으면 자동 토글',
            )
            # multiselect 변경 반영
            st.session_state.sort_active_boxes = active_boxes

        with sac2:
            if st.button('🔄 초기화', key='sort_reset', use_container_width=True):
                st.session_state.sort_state = _build_sort_state()
                st.session_state.sort_scan_counter = 0
                st.session_state.sort_last_result = None
                st.session_state.sort_active_boxes = []
                st.rerun()

        if active_boxes:
            # 활성 박스별로 필요한 출고박스 번호 상세 표시 (출고박스 크기별 그룹핑)
            active_info_lines = []
            all_needed_out_boxes = set()
            for b in sorted(active_boxes, key=_box_sort_key):
                info = box_qty_map[b]
                size_lbl, size_emo = _box_size(info['total_qty'])
                out_boxes = sorted(info['out_boxes'], key=_box_sort_key)
                all_needed_out_boxes.update(info['out_boxes'])
                # 출고박스를 크기별로 그룹화 (box_size_lookup 사용)
                _out_L, _out_M, _out_S = [], [], []
                for _o in out_boxes:
                    _osl, _ = box_size_lookup.get(_o, ('소', '🔵'))
                    if _osl == '대':
                        _out_L.append(_o)
                    elif _osl == '중':
                        _out_M.append(_o)
                    else:
                        _out_S.append(_o)
                parts = []
                if _out_L:
                    parts.append(f"🟢대형 {len(_out_L)}개: " + ', '.join(f'{o}번' for o in _out_L))
                if _out_M:
                    parts.append(f"🟡중형 {len(_out_M)}개: " + ', '.join(f'{o}번' for o in _out_M))
                if _out_S:
                    parts.append(f"🔵소형 {len(_out_S)}개: " + ', '.join(f'{o}번' for o in _out_S))
                active_info_lines.append(
                    f"**{b} 박스** ({size_emo}{size_lbl}, {info['total_qty']}개)\n"
                    + '\n'.join(f'　→ {p}' for p in parts)
                )
            total_out_count = len(all_needed_out_boxes)
            header = f'📦 **활성 배대지 박스 {len(active_boxes)}개 — 준비할 출고박스 총 {total_out_count}개**'
            st.info(header + '\n\n' + '\n\n'.join(active_info_lines))

            # ── 활성 박스별 빠른 완료 처리 버튼 (expander 밖으로 노출) ──
            def _mark_box_done_bulk(target_box):
                """배대지 박스 안의 모든 항목을 완료 처리 + 시트 L열 일괄 업데이트"""
                updates_for_sheet = []
                for _bc2, _v2 in sort_state.items():
                    for _it2 in _v2['items']:
                        if str(_it2.get('dapae_box', '')).strip().upper() == target_box:
                            _it2['scanned'] = _it2['needed']
                for _bc2, _v2 in sort_state.items():
                    _ship_cum = {}
                    touched_in_box = False
                    for _it2 in _v2['items']:
                        _dp2 = str(_it2.get('dapae_box', '')).strip().upper()
                        _ship2 = _it2.get('ship', '')
                        if _dp2 == target_box:
                            touched_in_box = True
                        if _ship2:
                            _ship_cum[_ship2] = _ship_cum.get(_ship2, 0) + _it2['scanned']
                    if touched_in_box:
                        for _s2, _c2 in _ship_cum.items():
                            updates_for_sheet.append((_bc2, _s2, _c2))
                if (st.session_state.get('pick_use_gsheet')
                        and st.session_state.get('pick_gsheet_client')
                        and st.session_state.get('pick_sheet_url_출고')
                        and st.session_state.get('pick_sheet_tab_출고')
                        and updates_for_sheet):
                    import threading
                    _client = st.session_state.pick_gsheet_client
                    _url = st.session_state.pick_sheet_url_출고
                    _tab = st.session_state.pick_sheet_tab_출고
                    def _bg_bulk():
                        for _ub, _us, _uq in updates_for_sheet:
                            try:
                                pick_update_check_qty(_client, _url, _tab, _ub, _us, _uq)
                            except Exception:
                                pass
                    threading.Thread(target=_bg_bulk, daemon=True).start()
                    return True, len(updates_for_sheet)
                return False, 0

            st.markdown('##### ✅ 박스 완료 처리 (이미 끝낸 배대지 박스)')
            _done_cols = st.columns(min(4, len(active_boxes)) or 1)
            for _i, _b in enumerate(sorted(active_boxes, key=_box_sort_key)):
                _info = box_qty_map[_b]
                _cur_s = 0
                _cur_n = 0
                for _bc_x, _v_x in sort_state.items():
                    for _it_x in _v_x['items']:
                        if str(_it_x.get('dapae_box', '')).strip().upper() == _b:
                            _cur_s += _it_x['scanned']
                            _cur_n += _it_x['needed']
                _is_done = _cur_s >= _cur_n and _cur_n > 0
                _label = f'✅ {_b} 완료 처리' if not _is_done else f'✔ {_b} 이미 완료'
                with _done_cols[_i % len(_done_cols)]:
                    if st.button(
                        _label,
                        key=f'sort_quick_done_{_b}',
                        use_container_width=True,
                        type='primary' if not _is_done else 'secondary',
                        disabled=_is_done,
                    ):
                        ok, n = _mark_box_done_bulk(_b)
                        if ok:
                            st.success(f'✅ {_b} 완료 처리 — 시트 L열 업데이트 중 ({n}건)')
                        else:
                            st.success(f'✅ {_b} 완료 처리 (세션만, 시트 미연결)')
                        st.rerun()

            # ── 활성 배대지 박스별 → 각 출고박스에 들어갈 내용물 리스트 ──
            st.markdown('#### 📋 출고박스에 담을 내용물 (배대지 박스 → 출고박스별)')
            active_set_upper = set(str(b).strip().upper() for b in active_boxes)

            # 각 출고박스별 배대지 구성 (out_box → {dapae_box → needed_qty}) — 한 번만 계산
            ob_composition = {}
            for _bc_c, _v_c in sort_state.items():
                for _it_c in _v_c['items']:
                    _ob_c = str(_it_c.get('out_box', '')).strip()
                    if not _ob_c.isdigit():
                        continue
                    _dp_c = str(_it_c.get('dapae_box', '')).strip().upper()
                    if not _dp_c or _dp_c == 'NAN':
                        continue
                    ob_composition.setdefault(_ob_c, {})
                    ob_composition[_ob_c][_dp_c] = ob_composition[_ob_c].get(_dp_c, 0) + _it_c['needed']

            for b in sorted(active_boxes, key=_box_sort_key):
                info = box_qty_map[b]
                size_lbl, size_emo = _box_size(info['total_qty'])
                # 이 배대지 박스에서 나가는 출고박스별 항목 집계
                ob_items = {}  # out_box → list of {바코드, 상품명, 필요, 스캔, 남음}
                for _bc, _v in sort_state.items():
                    for _it in _v['items']:
                        _dp = str(_it.get('dapae_box', '')).strip().upper()
                        if _dp != b:
                            continue
                        _ob = str(_it.get('out_box', '')).strip()
                        if not _ob.isdigit():
                            continue
                        ob_items.setdefault(_ob, []).append({
                            '바코드': _bc,
                            '상품명': _v['상품명'][:35],
                            '필요': _it['needed'],
                            '스캔': _it['scanned'],
                            '남음': max(0, _it['needed'] - _it['scanned']),
                        })
                # 합계 표시
                total_ob_n = sum(it['필요'] for items in ob_items.values() for it in items)
                total_ob_s = sum(it['스캔'] for items in ob_items.values() for it in items)
                head_pct = (total_ob_s / total_ob_n * 100) if total_ob_n else 0
                head_status = '✅' if total_ob_s >= total_ob_n and total_ob_n > 0 else ('🔄' if total_ob_s > 0 else '⬜')
                exp_label = (
                    f'{head_status} {b} 배대지박스 ({size_emo}{size_lbl}, {total_ob_s}/{total_ob_n}개, '
                    f'{head_pct:.0f}%) → 출고박스 {len(ob_items)}개'
                )
                with st.expander(exp_label, expanded=False):
                    # ── 박스 일괄 완료 처리 (이미 물리적으로 끝난 박스용) ──
                    btn_col1, btn_col2 = st.columns([2, 1])
                    with btn_col1:
                        st.caption('💡 이 배대지 박스가 이미 물리적으로 완료된 경우 → 우측 버튼으로 시트 L열 일괄 채움')
                    with btn_col2:
                        if st.button(
                            f'✅ {b} 박스 완료 처리',
                            key=f'sort_box_done_{b}',
                            use_container_width=True,
                            type='primary' if total_ob_s < total_ob_n else 'secondary',
                        ):
                            # 1) session_state 즉시 채움
                            updates_for_sheet = []  # [(bc, ship, cum_scanned), ...]
                            ship_cum_per_bc = {}    # bc → {ship: cum}
                            for _bc2, _v2 in sort_state.items():
                                for _it2 in _v2['items']:
                                    _dp2 = str(_it2.get('dapae_box', '')).strip().upper()
                                    if _dp2 != b:
                                        continue
                                    _it2['scanned'] = _it2['needed']
                            # 2) 송장별 누적 스캔수량 계산 (L열 덮어쓰기용)
                            for _bc2, _v2 in sort_state.items():
                                _ship_cum = {}
                                touched_in_box = False
                                for _it2 in _v2['items']:
                                    _dp2 = str(_it2.get('dapae_box', '')).strip().upper()
                                    _ship2 = _it2.get('ship', '')
                                    if _dp2 == b:
                                        touched_in_box = True
                                    if _ship2:
                                        _ship_cum[_ship2] = _ship_cum.get(_ship2, 0) + _it2['scanned']
                                if touched_in_box:
                                    for _s2, _c2 in _ship_cum.items():
                                        updates_for_sheet.append((_bc2, _s2, _c2))
                            # 3) 시트 L열 백그라운드 일괄 업데이트
                            if (st.session_state.get('pick_use_gsheet')
                                    and st.session_state.get('pick_gsheet_client')
                                    and st.session_state.get('pick_sheet_url_출고')
                                    and st.session_state.get('pick_sheet_tab_출고')
                                    and updates_for_sheet):
                                import threading
                                _client = st.session_state.pick_gsheet_client
                                _url = st.session_state.pick_sheet_url_출고
                                _tab = st.session_state.pick_sheet_tab_출고
                                def _bg_bulk():
                                    for _ub, _us, _uq in updates_for_sheet:
                                        try:
                                            pick_update_check_qty(_client, _url, _tab, _ub, _us, _uq)
                                        except Exception:
                                            pass
                                threading.Thread(target=_bg_bulk, daemon=True).start()
                                st.success(f'✅ {b} 배대지박스 완료 처리 — 시트 L열 업데이트 중 ({len(updates_for_sheet)}건)')
                            else:
                                st.success(f'✅ {b} 배대지박스 완료 처리 (세션만, 시트 미연결)')
                            st.rerun()

                    # ── 🎯 상품 단위 리스트 (수량 많은 순) — 제일 많은 것부터 준비 ──
                    sku_rows = []
                    for _bc_s, _v_s in sort_state.items():
                        total_in_box = 0
                        total_scanned_in_box = 0
                        out_box_break = {}  # out_box → qty
                        for _it_s in _v_s['items']:
                            _dp_s = str(_it_s.get('dapae_box', '')).strip().upper()
                            if _dp_s != b:
                                continue
                            total_in_box += _it_s['needed']
                            total_scanned_in_box += _it_s['scanned']
                            _ob_s = str(_it_s.get('out_box', '')).strip()
                            if _ob_s:
                                out_box_break[_ob_s] = out_box_break.get(_ob_s, 0) + _it_s['needed']
                        if total_in_box <= 0:
                            continue
                        # 출고박스 브레이크다운 (박스 번호 순)
                        ob_parts = sorted(out_box_break.items(), key=lambda x: _box_sort_key(x[0]))
                        ob_str = ', '.join(f'{k}번({v})' for k, v in ob_parts)
                        _sku_status = '✅' if total_scanned_in_box >= total_in_box else (
                            '🔄' if total_scanned_in_box > 0 else '⬜'
                        )
                        sku_rows.append({
                            '상태': _sku_status,
                            '이 박스 수량': total_in_box,
                            '스캔': total_scanned_in_box,
                            '남음': max(0, total_in_box - total_scanned_in_box),
                            '바코드': _bc_s,
                            '상품명': _v_s['상품명'],
                            '출고박스 배분': ob_str,
                        })
                    # 수량 많은 순 정렬 (수량 동률이면 남은 수량 많은 순)
                    sku_rows.sort(key=lambda r: (-r['이 박스 수량'], -r['남음']))
                    if sku_rows:
                        st.markdown('##### 🎯 이 박스 상품 (수량 많은 순) — 제일 많이 담긴 것부터 준비')
                        st.dataframe(
                            _pd2.DataFrame(sku_rows),
                            use_container_width=True, hide_index=True,
                            height=min(500, len(sku_rows) * 38 + 40),
                        )

                    st.markdown('---')
                    st.markdown('##### 📦 출고박스별 상세 (출고박스마다 담을 상품)')
                    for _ob in sorted(ob_items.keys(), key=_box_sort_key):
                        _items = ob_items[_ob]
                        _n = sum(x['필요'] for x in _items)
                        _s = sum(x['스캔'] for x in _items)
                        _ob_status = '✅' if _s >= _n and _n > 0 else ('🔄' if _s > 0 else '⬜')
                        _size_lbl2, _size_emo2 = box_size_lookup.get(_ob, ('', ''))
                        # 이 출고박스에 들어가는 모든 배대지 박스 구성 (현재 b는 굵게 표시)
                        _comp = ob_composition.get(_ob, {})
                        _all_parts = sorted(_comp.items(), key=lambda x: _box_sort_key(x[0]))
                        _total_qty = sum(v for _, v in _all_parts)
                        if _all_parts:
                            _parts_str = ', '.join(
                                (f'**{k}({v})**' if k == b else f'{k}({v})')
                                for k, v in _all_parts
                            )
                            _tail = f"  ·  📦 **구성 {_total_qty}개**: {_parts_str}"
                        else:
                            _tail = ''
                        st.markdown(
                            f"**{_ob_status} {_ob}번 출고박스** "
                            f"{_size_emo2}{_size_lbl2} — {_s}/{_n}개 ({len(_items)} SKU)"
                            f"{_tail}"
                        )
                        st.dataframe(
                            _pd2.DataFrame(_items),
                            use_container_width=True, hide_index=True,
                            height=min(250, len(_items) * 38 + 40),
                        )

        st.markdown('---')

        # ── 바코드 스캔 (fragment으로 감싸서 전체 앱 리런 없이 조각만 재실행) ──
        def _scan_rerun():
            """Fragment 내부면 조각 리런, 아니면 전체 리런 (구버전 Streamlit fallback)."""
            try:
                st.rerun(scope='fragment')
            except TypeError:
                st.rerun()
            except Exception:
                st.rerun()

        _use_fragment = getattr(st, 'fragment', lambda f: f)

        @_use_fragment
        def _scan_fragment():
            # ── 바코드 스캔 ──
            # 다량 모드 상태
            if 'sort_next_qty' not in st.session_state:
                st.session_state.sort_next_qty = 1
            if 'sort_qty_input_mode' not in st.session_state:
                st.session_state.sort_qty_input_mode = False

            # 수량 입력 모드: 큰 알림 + 전체 너비 입력창
            if st.session_state.sort_qty_input_mode:
                st.warning('🔢 **수량을 입력하세요** — 숫자 입력 후 Enter')
                qty_text_key = f'sort_qty_text_{st.session_state.sort_scan_counter}'
                qty_text = st.text_input(
                    '다량 수량',
                    key=qty_text_key,
                    placeholder='숫자 입력 후 Enter (예: 50)',
                    label_visibility='collapsed',
                )
                if qty_text:
                    try:
                        qty_val = int(qty_text.strip())
                        if qty_val >= 1:
                            st.session_state.sort_next_qty = qty_val
                            st.session_state.sort_qty_input_mode = False
                            st.session_state.sort_scan_counter += 1
                            _scan_rerun()
                    except ValueError:
                        st.error('숫자만 입력 가능합니다')

            # 수량 표시 + 1개 모드 리셋 버튼
            qcol1, qcol2 = st.columns([1, 1])
            with qcol1:
                if st.session_state.sort_next_qty > 1:
                    st.markdown(
                        f'<div style="background:#f59e0b;color:white;padding:0.5rem;border-radius:6px;text-align:center;font-weight:bold;font-size:1.1rem">'
                        f'📦 다음 스캔: {st.session_state.sort_next_qty}개'
                        f'</div>', unsafe_allow_html=True)
                else:
                    st.markdown(
                        '<div style="background:#e5e7eb;padding:0.5rem;border-radius:6px;text-align:center">'
                        '1개 모드'
                        '</div>', unsafe_allow_html=True)
            with qcol2:
                if st.session_state.sort_next_qty > 1 and not st.session_state.sort_qty_input_mode:
                    if st.button('🔄 1개 모드로 복귀', key='sort_qty_reset', use_container_width=True):
                        st.session_state.sort_next_qty = 1
                        _scan_rerun()

            sort_scan_key = f"sort_scan_{st.session_state.sort_scan_counter}"
            sort_scanned = st.text_input(
                '🔫 바코드 스캔',
                key=sort_scan_key,
                placeholder='박스에서 꺼낸 상품의 바코드를 스캔하세요 (여러 개면 #MULTI 바코드 먼저)',
            )

            def _process_sort_scan(bc):
                bc = bc.strip()

                # #MULTI 트리거 → 수량 입력 모드 진입
                if bc.upper() == '#MULTI':
                    st.session_state.sort_qty_input_mode = True
                    st.session_state.sort_next_qty = 1
                    return {
                        'status': 'multi_trigger',
                        'barcode': bc,
                        'message': '🔢 다량 입력 모드',
                        'detail': '수량을 입력한 후 상품 바코드를 스캔하세요',
                    }

                # #W1, #M2, #1 바코드 → 박스 토글
                import re as _re_bc
                box_label_match = _re_bc.match(r'^#([A-Za-z]*\d+)$', bc)
                if box_label_match:
                    bn = box_label_match.group(1).upper()
                    if bn in box_qty_map:
                        cur = list(st.session_state.get('sort_active_boxes', []))
                        if bn in cur:
                            cur.remove(bn)
                            msg_suffix = '제외'
                        else:
                            cur.append(bn)
                            msg_suffix = '선택'
                        st.session_state.sort_active_boxes = cur
                        info = box_qty_map[bn]
                        size_lbl, size_emo = _box_size(info['total_qty'])
                        return {
                            'status': 'box_toggle',
                            'barcode': bc,
                            'box_num': bn,
                            'message': f'📦 {bn}번 박스 {msg_suffix}',
                            'detail': f'{size_emo}{size_lbl}형 · {info["total_qty"]}개 · 송장 {len(info["ships"])}개',
                        }
                    else:
                        return {'status': 'error', 'barcode': bc,
                                'message': f'🚨 {bn}번 박스 없음', 'detail': bc}

                if bc not in sort_state:
                    return {'status': 'error', 'barcode': bc,
                            'message': '🚨 출고지시서에 없는 바코드',
                            'detail': bc}
                item_data = sort_state[bc]
                # 아직 채워야 할 박스 중 후보 선택
                candidates = [it for it in item_data['items'] if it['scanned'] < it['needed']]
                if not candidates:
                    return {'status': 'over', 'barcode': bc,
                            '상품명': item_data['상품명'],
                            'message': '⚠️ 이 상품은 모두 분류 완료',
                            'detail': item_data['상품명'][:35]}

                # 활성 배대지 박스 집합 필터: 선택된 배대지 박스들의 상품만 유효
                _active_boxes_set = set(
                    str(b).strip().upper() for b in st.session_state.get('sort_active_boxes', [])
                )
                if _active_boxes_set:
                    box_candidates = []
                    for it in candidates:
                        it_dp = str(it.get('dapae_box', '')).strip().upper()
                        if it_dp and it_dp in _active_boxes_set:
                            box_candidates.append(it)
                    if not box_candidates:
                        # 이 상품의 배대지 박스가 활성 목록에 없음
                        other_boxes = sorted(set(
                            str(it.get('dapae_box', ''))
                            for it in item_data['items']
                            if it.get('dapae_box')
                        ))
                        return {'status': 'wrong_box', 'barcode': bc,
                                '상품명': item_data['상품명'],
                                'message': f'🚨 활성 배대지 박스에 없는 상품',
                                'detail': f'이 상품은 {", ".join(other_boxes)} 배대지 박스에 있음'}
                    candidates = box_candidates

                # 다량 모드: next_qty 만큼 차감 (여러 박스에 걸쳐 자동 분배)
                requested_qty = int(st.session_state.get('sort_next_qty', 1))
                requested_qty = max(1, requested_qty)
                processed = 0
                last_target = candidates[0]
                touched_ships = set()  # 이번 스캔으로 영향받은 송장(시트 L열 업데이트용)
                # 후보들을 순회하며 각 박스 채워가기
                remaining_to_scan = requested_qty
                idx = 0
                while remaining_to_scan > 0 and idx < len(candidates):
                    it = candidates[idx]
                    space = it['needed'] - it['scanned']
                    if space <= 0:
                        idx += 1
                        continue
                    take = min(space, remaining_to_scan)
                    it['scanned'] += take
                    processed += take
                    remaining_to_scan -= take
                    last_target = it
                    if it.get('ship'):
                        touched_ships.add(it['ship'])
                    if it['scanned'] >= it['needed']:
                        idx += 1

                # 영향받은 송장별 누적 스캔수량 (L열 덮어쓰기용)
                touched_updates = []
                if touched_ships:
                    ship_cum = {}
                    for _it in item_data['items']:
                        _s = _it.get('ship')
                        if _s in touched_ships:
                            ship_cum[_s] = ship_cum.get(_s, 0) + _it['scanned']
                    for _s, _c in ship_cum.items():
                        touched_updates.append((bc, _s, _c))

                # 모두 차감 후 남은 수량 (수량 초과)
                over_qty = requested_qty - processed

                # 다량 모드는 1회용: 원래대로 복귀
                st.session_state.sort_next_qty = 1
                st.session_state.sort_qty_input_mode = False

                # 이 박스(box_key)가 다 채워졌는지 확인
                target_box_key = last_target['box_key']
                box_complete = True
                for _bc, _v in sort_state.items():
                    for _it in _v['items']:
                        if _it['box_key'] == target_box_key and _it['scanned'] < _it['needed']:
                            box_complete = False
                            break
                    if not box_complete:
                        break

                return {
                    'status': 'ok',
                    'barcode': bc,
                    '상품명': item_data['상품명'],
                    'box_key': last_target['box_key'],
                    'box_num': last_target['box_num'],
                    'sym': last_target['sym'],
                    'ship': last_target['ship'],
                    'remaining': last_target['needed'] - last_target['scanned'],
                    'box_complete': box_complete,
                    'processed_qty': processed,
                    'over_qty': over_qty,
                    'touched_updates': touched_updates,  # [(bc, ship, cum_scanned), ...]
                }

            if sort_scanned:
                scan_result = _process_sort_scan(sort_scanned)
                st.session_state.sort_last_result = scan_result
                st.session_state.sort_scan_counter += 1

                # 구글 시트 L열(확인 수량) 업데이트 (백그라운드)
                # touched_updates: 이번 스캔으로 영향받은 (bc, ship, 누적스캔수량) 목록
                # 누적값으로 덮어쓰므로 시트를 다시 로드해도 진행 상태가 보존됨
                if (scan_result.get('status') == 'ok'
                        and st.session_state.get('pick_use_gsheet')
                        and st.session_state.get('pick_gsheet_client')
                        and st.session_state.get('pick_sheet_url_출고')
                        and st.session_state.get('pick_sheet_tab_출고')):
                    import threading
                    _updates = scan_result.get('touched_updates') or []
                    if not _updates:
                        # fallback: 단일 스캔 시 last_target 정보로
                        _bc = scan_result.get('barcode', '')
                        _ship = scan_result.get('ship', '')
                        _qty = scan_result.get('processed_qty', 1)
                        _updates = [(_bc, _ship, _qty)]
                    _client = st.session_state.pick_gsheet_client
                    _url = st.session_state.pick_sheet_url_출고
                    _tab = st.session_state.pick_sheet_tab_출고
                    def _bg_update_check():
                        for _ub, _us, _uq in _updates:
                            try:
                                pick_update_check_qty(_client, _url, _tab, _ub, _us, _uq)
                            except Exception:
                                pass
                    threading.Thread(target=_bg_update_check, daemon=True).start()

                _scan_rerun()

            # 자동 포커스 (multiselect/number input 상호작용 중에는 포커스 안 가로챔)
            from streamlit.components.v1 import html as _sort_html
            _sort_html("""<script>
            (function(){
                const doc = window.parent.document;
                function findScan(){
                    const inputs = doc.querySelectorAll('input[type="text"]');
                    for (const inp of inputs){
                        if (inp.placeholder && inp.placeholder.includes('박스에서 꺼낸')) return inp;
                    }
                    return null;
                }
                function findQtyInput(){
                    const inputs = doc.querySelectorAll('input[type="text"]');
                    for (const inp of inputs){
                        if (inp.placeholder && inp.placeholder.includes('숫자 입력')) return inp;
                    }
                    return null;
                }
                function isInteractingOther(){
                    const active = doc.activeElement;
                    if (!active) return false;
                    const tag = (active.tagName || '').toLowerCase();
                    // number input (수량), textarea, button
                    if (tag === 'button' || tag === 'textarea') return true;
                    if (tag === 'input' && active.type !== 'text') return true;
                    // Streamlit BaseWeb select (multiselect) 내부
                    if (active.closest) {
                        if (active.closest('[data-baseweb="select"]')) return true;
                        if (active.closest('[data-baseweb="popover"]')) return true;
                        if (active.closest('[role="listbox"]')) return true;
                        if (active.closest('[role="combobox"]')) return true;
                    }
                    return false;
                }
                function focusScan(){
                    // 수량 입력창이 표시되어 있으면 우선 그 창으로 포커스
                    const qty = findQtyInput();
                    if (qty) {
                        if (doc.activeElement !== qty) qty.focus({preventScroll: true});
                        return;
                    }
                    const inp = findScan();
                    if (!inp) return;
                    if (doc.activeElement === inp) return;
                    if (isInteractingOther()) return;
                    // 팝오버/드롭다운이 열려있으면 포커스 안 함
                    if (doc.querySelector('[data-baseweb="popover"]')) return;
                    // preventScroll: 포커스 이동 시 브라우저 자동 스크롤 방지
                    inp.focus({preventScroll: true});
                }
                focusScan();
                if (window._sortFocusInterval) clearInterval(window._sortFocusInterval);
                window._sortFocusInterval = setInterval(focusScan, 500);
            })();
            </script>""", height=0)

            # ── 결과 표시 + 음성 ──
            r = st.session_state.get('sort_last_result')
            if r:
                _KOR_NUMS_SORT = {
                    1:'일',2:'이',3:'삼',4:'사',5:'오',6:'육',7:'칠',8:'팔',9:'구',10:'십',
                    11:'십일',12:'십이',13:'십삼',14:'십사',15:'십오',16:'십육',17:'십칠',
                    18:'십팔',19:'십구',20:'이십',21:'이십일',22:'이십이',23:'이십삼',24:'이십사',
                    25:'이십오',26:'이십육',27:'이십칠',28:'이십팔',29:'이십구',30:'삼십',
                }
                _KOR_ALPHA = {
                    'W': '더블유', 'M': '엠', 'L': '엘', 'S': '에스',
                    'A': '에이', 'B': '비', 'C': '씨', 'D': '디', 'E': '이', 'F': '에프',
                    'G': '지', 'H': '에이치', 'I': '아이', 'J': '제이', 'K': '케이',
                    'N': '엔', 'O': '오', 'P': '피', 'Q': '큐', 'R': '알', 'T': '티',
                    'U': '유', 'V': '브이', 'X': '엑스', 'Y': '와이', 'Z': '지',
                }
                def _num_to_kor(n):
                    if n in _KOR_NUMS_SORT:
                        return _KOR_NUMS_SORT[n]
                    if n <= 99:
                        tens = n // 10
                        ones = n % 10
                        t_str = ('이삼사오육칠팔구'[tens - 2] if tens >= 2 else '') + '십'
                        return t_str + (_KOR_NUMS_SORT.get(ones, '') if ones else '')
                    return str(n)
                def _box_to_kor(n_str):
                    """W1 → '더블유 일', M3 → '엠 삼', 1 → '일'"""
                    import re as _re_k
                    s = str(n_str).strip().upper()
                    m = _re_k.match(r'^([A-Z]*)(\d+)$', s)
                    if not m:
                        return s
                    alpha_part = m.group(1)
                    num_part = int(m.group(2))
                    alpha_kor = ' '.join(_KOR_ALPHA.get(ch, ch) for ch in alpha_part)
                    num_kor = _num_to_kor(num_part)
                    if alpha_kor:
                        return f'{alpha_kor} {num_kor}'
                    return num_kor

                if r['status'] == 'multi_trigger':
                    st.markdown(
                        f'<div class="scan-complete" style="background:#f59e0b;color:white;padding:2rem;border-radius:12px;text-align:center;border-left:8px solid #d97706">'
                        f'<div style="font-size:2.2rem;font-weight:bold;">🔢 수량을 입력하세요</div>'
                        f'<div style="font-size:1.1rem;margin-top:0.8rem;">수량 입력 후 "수량 확정" 또는 바로 상품 바코드 스캔</div>'
                        f'</div>',
                        unsafe_allow_html=True)
                    speak = '수량을 입력하세요'
                elif r['status'] == 'box_toggle':
                    st.markdown(
                        f'<div class="scan-complete" style="background:#3b82f6;color:white;padding:1.5rem;border-radius:10px;text-align:center;border-left:8px solid #1e40af">'
                        f'<div style="font-size:1.8rem;font-weight:bold;">{r["message"]}</div>'
                        f'<div style="font-size:1rem;margin-top:0.5rem;opacity:0.95;">{r["detail"]}</div>'
                        f'</div>',
                        unsafe_allow_html=True)
                    _kor_bt = _box_to_kor(str(r['box_num']))
                    speak = f'{_kor_bt}번'
                elif r['status'] == 'error':
                    st.markdown(
                        f'<div class="scan-error"><strong style="font-size:1.4rem;">📥 보류</strong><br>'
                        f'쉽먼트 정보 없음 - 따로 보관<br>{r["detail"]}</div>',
                        unsafe_allow_html=True)
                    speak = '보류'
                elif r['status'] == 'wrong_box':
                    st.markdown(
                        f'<div class="scan-error"><strong style="font-size:1.3rem;">{r["message"]}</strong><br>{r["detail"]}</div>',
                        unsafe_allow_html=True)
                    speak = '다른 박스 상품'
                elif r['status'] == 'over':
                    st.markdown(f'<div class="scan-warning"><strong style="font-size:1.2rem;">{r["message"]}</strong><br>{r["detail"]}</div>', unsafe_allow_html=True)
                    speak = '분류 완료'
                else:
                    box_num_str = str(r['box_num']).strip().upper()
                    kor_n = _box_to_kor(box_num_str)
                    size_label, size_emoji = box_size_lookup.get(box_num_str, ('', ''))
                    size_str = f' ({size_emoji}{size_label}형)' if size_label else ''
                    if r.get('box_complete'):
                        # 박스 완료! 큰 알림 + 포장 안내
                        st.markdown(
                            f'<div class="scan-complete" style="background:#10b981;color:white;padding:2rem;border-radius:12px;text-align:center;border-left:8px solid #059669">'
                            f'<div style="font-size:2.5rem;font-weight:bold;">🎉 {box_num_str}번 {size_str} 완료!</div>'
                            f'<div style="font-size:1.3rem;margin-top:0.8rem;">📦 포장하고 출고지시서 종이를 끼워주세요</div>'
                            f'<div style="font-size:1rem;margin-top:0.5rem;opacity:0.9;word-break:break-all;">마지막 상품: {r["상품명"]}</div>'
                            f'</div>',
                            unsafe_allow_html=True)
                        speak = f'{kor_n}번 완료. 포장하세요'
                    else:
                        processed_qty = r.get('processed_qty', 1)
                        over_qty = r.get('over_qty', 0)
                        qty_str = f' × {processed_qty}개' if processed_qty > 1 else ''
                        over_str = f' ⚠️ {over_qty}개 초과' if over_qty > 0 else ''
                        st.markdown(
                            f'<div class="scan-ok" style="word-break:break-all;"><strong style="font-size:1.5rem;">✅ {box_num_str}번{size_str}{qty_str} → {r["상품명"]}</strong><br>'
                            f'송장 {r["ship"][-6:]} | 남은 수량: {r["remaining"]}개{over_str}</div>',
                            unsafe_allow_html=True)
                        if processed_qty > 1:
                            speak = f'{kor_n}번 {processed_qty}개'
                        else:
                            speak = f'{kor_n}번'

                # 소리 + 음성
                speak_js = speak.replace("'", "\\'").replace('"', '\\"')
                scan_id_s = st.session_state.sort_scan_counter
                beep_js = "o.frequency.value=880;g.gain.value=0.3;o.start();setTimeout(()=>g.gain.value=0,150);setTimeout(()=>o.stop(),200);"
                if r['status'] in ('error', 'wrong_box'):
                    beep_js = "o.type='square';o.frequency.value=200;g.gain.value=0.5;o.start();setTimeout(()=>{o.frequency.value=150},150);setTimeout(()=>g.gain.value=0,500);setTimeout(()=>o.stop(),600);"
                elif r['status'] == 'over':
                    beep_js = "o.type='sawtooth';o.frequency.value=400;g.gain.value=0.4;o.start();setTimeout(()=>g.gain.value=0,300);setTimeout(()=>o.stop(),400);"
                elif r.get('box_complete'):
                    # 박스 완료 - 축하 멜로디 (3음)
                    beep_js = ("o.frequency.value=523;g.gain.value=0.4;o.start();"
                               "setTimeout(()=>{o.frequency.value=659},120);"
                               "setTimeout(()=>{o.frequency.value=784},240);"
                               "setTimeout(()=>g.gain.value=0,400);"
                               "setTimeout(()=>o.stop(),500);")
                _sort_html(f"""<script>
                // sort_id={scan_id_s}
                try{{var a=new(window.AudioContext||window.webkitAudioContext)();var o=a.createOscillator();var g=a.createGain();o.connect(g);g.connect(a.destination);{beep_js}}}catch(e){{}}
                try{{
                    window.speechSynthesis.cancel();
                    setTimeout(function(){{
                        var u = new SpeechSynthesisUtterance('{speak_js}');
                        u.lang='ko-KR'; u.rate=1.3; u.volume=1.0;
                        var v = window.speechSynthesis.getVoices().find(x => x.lang && x.lang.startsWith('ko'));
                        if(v) u.voice=v;
                        window.speechSynthesis.speak(u);
                    }}, 100);
                }}catch(e){{}}
                </script>""", height=0)

        _scan_fragment()

        # ── 박스별 진행 현황 ──
        st.markdown('---')
        st.subheader('📋 박스별 진행 현황')
        box_summary = {}
        for bc, v in sort_state.items():
            for it in v['items']:
                key = it['box_key']
                ent = box_summary.setdefault(key, {
                    'box_num': it['box_num'], 'sym': it['sym'],
                    'needed': 0, 'scanned': 0, 'sku_total': 0, 'sku_done': 0,
                    'ships': set(),
                })
                ent['needed'] += it['needed']
                ent['scanned'] += it['scanned']
                ent['sku_total'] += 1
                if it.get('ship'):
                    ent['ships'].add(it['ship'])
                if it['scanned'] >= it['needed']:
                    ent['sku_done'] += 1

        # 정렬: 미완료(대기/진행) 먼저 → 완료 맨 아래 / 박스 번호 순
        def _box_prog_sort_key(kv):
            key, ent = kv
            done = (ent['needed'] > 0 and ent['scanned'] >= ent['needed'])
            return (1 if done else 0, ent['sym'], _box_sort_key(ent['box_num']))

        st.caption('💡 각 박스 헤더를 클릭하면 그 박스에 들어가는 모든 상품이 펼쳐집니다')
        for key, ent in sorted(box_summary.items(), key=_box_prog_sort_key):
            pct = (ent['scanned'] / ent['needed'] * 100) if ent['needed'] else 0
            is_done = (ent['needed'] > 0 and ent['scanned'] >= ent['needed'])
            if is_done:
                status = '✅ 완료'
            elif ent['scanned'] > 0:
                status = f'🔄 {pct:.0f}%'
            else:
                status = '⬜ 대기'
            _bn_key_p = str(ent['box_num']).strip().upper()
            _size_label, _size_emoji = box_size_lookup.get(_bn_key_p, ('', ''))
            _size_str = f'{_size_emoji}{_size_label}' if _size_label else ''
            exp_title = (
                f"{status} · **{ent['box_num']}번 박스** {_size_str} · "
                f"{ent['scanned']}/{ent['needed']}개 · "
                f"SKU {ent['sku_done']}/{ent['sku_total']} · "
                f"송장 {len(ent['ships'])}개"
            )
            # 미완료 박스는 기본 펼침
            with st.expander(exp_title, expanded=(not is_done and ent['scanned'] > 0)):
                detail_rows = []
                for _bc_d, _v_d in sort_state.items():
                    for _it_d in _v_d['items']:
                        if _it_d['box_key'] != key:
                            continue
                        _d_s = _it_d['scanned']
                        _d_n = _it_d['needed']
                        _d_status = '✅' if _d_s >= _d_n and _d_n > 0 else ('🔄' if _d_s > 0 else '⬜')
                        detail_rows.append({
                            '상태': _d_status,
                            '바코드': _bc_d,
                            '상품명': _v_d['상품명'],
                            '배대지박스': _it_d.get('dapae_box', '') or '-',
                            '송장': _it_d.get('ship', ''),
                            '필요': _d_n,
                            '스캔': _d_s,
                            '남음': max(0, _d_n - _d_s),
                        })
                detail_rows.sort(key=lambda r: (
                    _box_sort_key(r['배대지박스']) if r['배대지박스'] != '-' else ('ZZZ', 99999),
                    r['송장'], r['바코드'],
                ))
                if detail_rows:
                    st.dataframe(
                        _pd2.DataFrame(detail_rows),
                        use_container_width=True, hide_index=True,
                        height=min(500, len(detail_rows) * 38 + 40),
                    )
                else:
                    st.caption('이 박스에 해당하는 항목이 없습니다')

        # 미스캔 항목
        incomplete = []
        for bc, v in sort_state.items():
            for it in v['items']:
                if it['scanned'] < it['needed']:
                    incomplete.append({
                        '박스': f"{it['box_num']}번",
                        '바코드': bc,
                        '상품명': v['상품명'][:35],
                        '필요': it['needed'],
                        '스캔': it['scanned'],
                        '남음': it['needed'] - it['scanned'],
                    })
        if incomplete:
            with st.expander(f'⏳ 미스캔 ({len(incomplete)}건)', expanded=False):
                st.dataframe(_pd2.DataFrame(incomplete), use_container_width=True, hide_index=True)

